# -*- coding: utf-8 -*-
"""
Created on Wed Mar 25 10:28:26 2020

@author: Jamal Khan
"""

'''
Egg & Chicken latest forecasts:
    weekly and daily 2020 forecasts
    NOTE: to update first run DataPrep_CnE and WeeklyReturns scripts
'''
 

from sklearn.model_selection import train_test_split
from sklearn import preprocessing
from sklearn.pipeline import make_pipeline
from sklearn.model_selection import GridSearchCV
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.externals import joblib 

#############################################################
# Import regression models class
#############################################################
from IPython import get_ipython
get_ipython().magic('reset -sf')

import os
wd = os.getcwd()
os.chdir('//rschfiler2.baml.com/UK-NYCOMMODITIES/Jamal')
from regression_models import regression_models
os.chdir(wd)

#############################################################
# Load data
#############################################################
import pandas as pd
import numpy as np
import os
import _pickle as pickle
import time
### set paths 
### !!UPDATE!! before running
data_dir = '//rschfiler2.baml.com/UK-NYCOMMODITIES/Jamal/CnE/Data'
out_dir = '//rschfiler2.baml.com/UK-NYCOMMODITIES/Jamal/CnE/Output'

fname=os.path.join(data_dir, 'signal_returns_co.pickle')
signal_co = pickle.load( open( fname, "rb" ) )

fname=os.path.join(data_dir, 'signal_returns_fx.pickle')
signal_fx = pickle.load( open( fname, "rb" ) )

fname=os.path.join(data_dir, 'trade_returns_co.pickle')
trade_co = pickle.load( open( fname, "rb" ) )

fname=os.path.join(data_dir, 'trade_returns_fx.pickle')
trade_fx = pickle.load( open( fname, "rb" ) )


fname=os.path.join(data_dir, 'signal_returns_daily_co.pickle')
signal_d_co = pickle.load( open( fname, "rb" ) )

fname=os.path.join(data_dir, 'signal_returns_daily_fx.pickle')
signal_d_fx = pickle.load( open( fname, "rb" ) )

fname=os.path.join(data_dir, 'trade_returns_daily_co.pickle')
trade_d_co = pickle.load( open( fname, "rb" ) )

fname=os.path.join(data_dir, 'trade_returns_daily_fx.pickle')
trade_d_fx = pickle.load( open( fname, "rb" ) )


fname=os.path.join(data_dir, 'signal_returns_monthly_co.pickle')
signal_m_co = pickle.load( open( fname, "rb" ) )

fname=os.path.join(data_dir, 'signal_returns_monthly_fx.pickle')
signal_m_fx = pickle.load( open( fname, "rb" ) )

fname=os.path.join(data_dir, 'trade_returns_monthly_co.pickle')
trade_m_co = pickle.load( open( fname, "rb" ) )

fname=os.path.join(data_dir, 'trade_returns_monthly_fx.pickle')
trade_m_fx = pickle.load( open( fname, "rb" ) )



def _20_performance(model, version, sw, baskets, cols):
    
    # generate performance stats -- no signal weighting
    model.performance_stats()
   
    writer=pd.ExcelWriter(os.path.join(data_dir, version + '2020_single_co.xlsx'))
    writer.save()

    hr_df = pd.DataFrame(index = model.hit_ratio_by_year.keys(), columns = model.models)
    ir_df = pd.DataFrame(index = model.hit_ratio_by_year.keys(), columns = model.models)
    ar_df = pd.DataFrame(index = model.hit_ratio_by_year.keys(), columns = model.models)

    for p in model.hit_ratio_by_year.keys(): 
        hr_df.loc[p,:] = model.hit_ratio_by_year[p][model.hit_ratio_by_year[p].index==2020].values
        ir_df.loc[p,:] = model.info_ratio_by_year[p][model.info_ratio_by_year[p].index==2020].values
        ar_df.loc[p,:] = model.ann_returns_by_year[p][model.ann_returns_by_year[p].index==2020].values
    
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + '2020_single_co.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + '2020_single_co.xlsx'),engine='openpyxl')
    writer.book=book
    hr_df.to_excel(writer,'hitratio')
    ir_df.to_excel(writer,'inforatio')
    ar_df.to_excel(writer,'annreturn')
    writer.save()
          
    avg_info = pd.DataFrame(index = cols, columns = model.models)
    avg_ret =  pd.DataFrame(index = cols, columns = model.models)
    
    for b,c in zip(baskets,cols):    
    #construct strategies
        model.construct_strategy(basket = b, signal_weighting = sw)
        avg_info.loc[c,:] = model.strategy_eq_info[(model.strategy_eq_info.index==2020)].values  
        avg_ret.loc[c,:] = model.strategy_eq_ret[(model.strategy_eq_ret.index==2020)].values


    writer=pd.ExcelWriter(os.path.join(data_dir, version + '2020_baskets.xlsx'))
    writer.save()

    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + '2020_baskets.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + '2020_baskets.xlsx'),engine='openpyxl')
    writer.book=book
    avg_info.to_excel(writer,'info')
    avg_ret.to_excel(writer,'ret')
    writer.save()


#############################################################
# Run models -- CnE & EnC
#############################################################

'''
EnC
'''

energy =      ['Brent','WTI','Natural Gas','Heating Oil','Gasoline','Gasoil']
ags =         ['Cocoa','Coffee','Corn','Cotton','Sugar','Wheat','Kansas Wheat','Soybean Oil','Soybeans','Soybean Meal','Orange Juice']
industrials = ['Aluminum','Nickel','Zinc','Tin','Copper']
precious =    ['Gold','Silver','Palladium','Platinum']
livestock =   ['Lean Hogs','Live Cattle','Feeder Cattle'] 
_all =        energy + ags + industrials + precious + livestock + ['Softs','Grains','Industrial Metals','Livestock','Precious Metals','Petroleum']


#############################################################
# Run models -- Weekly
#############################################################

enc = regression_models(X = signal_fx.replace(np.nan, 0).iloc[-200:, :], Y = trade_co.replace(np.nan, 0).iloc[-200:, :], freq = 'weekly', window = 52, 
                        datevars_x = ['date', 'signalDate'], datevars_y = ['date', 'tradeDate'], mergevar = ['date'])

enc.fit()
enc.ensemble()

with open(os.path.join(data_dir, 'encModels_latest.pickle'), 'wb') as output:
    pickle.dump(enc, output)


fname=os.path.join(data_dir, 'encModels_latest.pickle')
model = pickle.load( open( fname, "rb" ))
    
_20_performance(model = model, version = 'enc_', sw = False,
                baskets = [_all, energy, ags, industrials, precious, livestock], 
                cols   = ['All', 'Energy', 'Agricultural', 'Industrial metals', 'Precious metals', 'Livestock'])


for q,r in zip(model.predictions, model.models):
    writer=pd.ExcelWriter(os.path.join(data_dir,'2020', r + '.xlsx'))
    writer.save()
    for p in q.keys():
        from openpyxl import load_workbook
        book= load_workbook(os.path.join(data_dir,'2020', r + '.xlsx'))
        writer=pd.ExcelWriter(os.path.join(data_dir,'2020', r + '.xlsx'),engine='openpyxl')
        writer.book=book
        q[p].to_excel(writer,(p))
        writer.save()


writer=pd.ExcelWriter(os.path.join(data_dir,'2020/Actual.xlsx'))
writer.save()
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir,'2020/Actual.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir,'2020/Actual.xlsx'),engine='openpyxl')
writer.book=book
trade_co.to_excel(writer,'Actual')
writer.save()

#############################################################

'''
CnE
'''

#############################################################
# Run models -- Weekly
#############################################################

cne = regression_models(X = signal_co.replace(np.nan, 0).iloc[-200:, :], Y = trade_fx.replace(np.nan, 0).iloc[-200:, :], freq = 'weekly', window = 52, 
                        datevars_x = ['date', 'signalDate'], datevars_y = ['date', 'tradeDate'], mergevar = ['date'])

cne.fit()

with open(os.path.join(data_dir, 'cneModels_latest.pickle'), 'wb') as output:
    pickle.dump(cne, output)

    
#############################################################
# Run models -- EnC Daily
#############################################################

enc = regression_models(X = signal_d_fx.replace(np.nan, 0).iloc[-600:,], Y = trade_d_co.replace(np.nan, 0).iloc[-600:,], freq = 'daily', window = 252, 
                        datevars_x = ['date', 'signalDate'], datevars_y = ['date', 'tradeDate'], mergevar = ['date'])

enc.fit()

with open(os.path.join(data_dir, 'encModels_daily_latest.pickle'), 'wb') as output:
    pickle.dump(enc, output)

# export -- daily
fname=os.path.join(data_dir, 'encModels_daily_latest.pickle')
model = pickle.load( open( fname, "rb" ))
    
    # generate performance stats -- no signal weighting
model.ensemble()

for q,r in zip(model.predictions, model.models):
    writer=pd.ExcelWriter(os.path.join(data_dir,'2020/daily', r + '.xlsx'))
    writer.save()
    for p in q.keys():
        from openpyxl import load_workbook
        book= load_workbook(os.path.join(data_dir,'2020/daily', r + '.xlsx'))
        writer=pd.ExcelWriter(os.path.join(data_dir,'2020/daily', r + '.xlsx'),engine='openpyxl')
        writer.book=book
        q[p].to_excel(writer,(p))
        writer.save()


writer=pd.ExcelWriter(os.path.join(data_dir,'2020/daily/Actual.xlsx'))
writer.save()
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir,'2020/daily/Actual.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir,'2020/daily/Actual.xlsx'),engine='openpyxl')
writer.book=book
trade_co.to_excel(writer,'Actual')
writer.save()
