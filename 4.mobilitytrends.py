# -*- coding: utf-8 -*-
from IPython import get_ipython
get_ipython().magic('reset -sf')

#############################################################
# Load/import data
#############################################################
import pandas as pd
import numpy as np
import os
import pickle
import datetime as dt

### set paths 
### !!UPDATE!! before running
data_dir = '//rschfiler2.baml.com/UK-NYCOMMODITIES/Jamal/CovidAnalysis/Mobility/Other'

#################
# import data
#################

# contact tracing
fname = os.path.join(data_dir, 'covid-contact-tracing.csv')
contact_tracing = pd.read_csv(fname,skiprows=range(0))
contact_tracing = contact_tracing[['Date','Entity','Contact tracing (OxBSG)']] 
contact_tracing.columns = ['date','location','contact_tracing']
contact_tracing['date'] = contact_tracing['date'].apply(lambda x: dt.datetime.strptime(x, '%b %d, %Y'))

# testing
fname = os.path.join(data_dir, 'full-list-total-tests-for-covid-19.csv')
testing = pd.read_csv(fname,skiprows=range(0))
testing = testing[['Date','Entity','Total tests']] 
testing.columns = ['date','location','total_tests']
testing['date'] = testing['date'].apply(lambda x: dt.datetime.strptime(x, '%b %d, %Y'))

# public gathering rules
fname = os.path.join(data_dir, 'public-gathering-rules-covid.csv')
public_rules = pd.read_csv(fname,skiprows=range(0))
public_rules = public_rules[['Date','Entity','Restrictions on gatherings (OxBSG)']]
public_rules.columns = ['date','location','public_rules']
public_rules['date'] = public_rules['date'].apply(lambda x: dt.datetime.strptime(x, '%b %d, %Y'))

# gdp per cap
fname = os.path.join(data_dir, 'gdppc.xlsx')
gdppc = pd.read_excel(fname,skiprows=range(0))
temp=gdppc.T.iloc[2:,:].fillna(method='ffill').tail(1).reset_index(drop=True).T
gdppc = pd.concat([gdppc['Country Name'],temp],axis=1).dropna().reset_index(drop=True)
gdppc.columns = ['location','gdppc']

# healthcare access
fname = os.path.join(data_dir, 'healthcare-access-and-quality-index.csv')
healthcare = pd.read_csv(fname,skiprows=range(0))
healthcare = healthcare[healthcare['Year']==2015].reset_index(drop=True) 
healthcare = healthcare[['Entity','HAQ Index (IHME (2017))']]
healthcare.columns = ['location','health_access']

# diabetes
fname = os.path.join(data_dir, 'diabetes.xlsx')
diabetes = pd.read_excel(fname,skiprows=range(0))
diabetes.columns = ['location','diabetes'] 

# GCI
fname = os.path.join(data_dir, 'GCI.xlsx')
gci = pd.read_excel(fname,skiprows=range(0))
gci.columns = ['location','gci'] 

# obesity
fname = os.path.join(data_dir, 'obesity.xlsx')
obesity = pd.read_excel(fname,skiprows=range(0))
obesity.columns = ['location','obesity'] 

# median age
fname = os.path.join(data_dir, 'medianage.xlsx')
age = pd.read_excel(fname,skiprows=range(1))
age.columns = ['location','median_age'] 

# population over 65
fname = os.path.join(data_dir, 'PopOver65.xlsx')
pop65 = pd.read_excel(fname,skiprows=range(0))
temp=pop65.T.iloc[2:,:].fillna(method='ffill').tail(1).reset_index(drop=True).T
pop65 = pd.concat([pop65['Country Name'],temp],axis=1).dropna().reset_index(drop=True)
pop65.columns = ['location','pop65'] 

# population level+density
fname = os.path.join(data_dir, 'population.xlsx')
pop = pd.read_excel(fname,sheet_name='country',skiprows=range(0))
pop = pop[['Country', 'per sq mile', 'population', 'Region']]
pop.columns = ['location', 'pop_per_sq_m', 'pop', 'region']

#urban population
fname = os.path.join(data_dir, 'urban.xlsx')
urban = pd.read_excel(fname,skiprows=range(0))
temp=urban.T.iloc[2:,:].fillna(method='ffill').tail(1).reset_index(drop=True).T
urban = pd.concat([urban['Country Name'],temp],axis=1).dropna().reset_index(drop=True)
urban.columns = ['location','urban'] 

#healthcare per capita
fname = os.path.join(data_dir, 'healthcare_pc.xlsx')
healthcare_pc = pd.read_excel(fname,skiprows=range(0))
temp=healthcare_pc.T.iloc[2:,:].fillna(method='ffill').tail(1).reset_index(drop=True).T
healthcare_pc = pd.concat([healthcare_pc['Country Name'],temp],axis=1).dropna().reset_index(drop=True)
healthcare_pc.columns = ['location','healthcare_pc'] 

#oop expenditure
fname = os.path.join(data_dir, 'oop.xlsx')
oop = pd.read_excel(fname,skiprows=range(0))
temp=oop.T.iloc[2:,:].fillna(method='ffill').tail(1).reset_index(drop=True).T
oop = pd.concat([oop['Country Name'],temp],axis=1).dropna().reset_index(drop=True)
oop.columns = ['location','oop'] 


#################
# merge
#################
data_dir = '//rschfiler2.baml.com/UK-NYCOMMODITIES/Jamal/CovidAnalysis/Mobility'

fname=os.path.join(data_dir, 'cases_country.pickle')
cases_country = pickle.load( open( fname, "rb" ) )

_0 = pd.DataFrame(cases_country['location'].unique())
_0.columns = ['cases_country']

_1 = pd.DataFrame(contact_tracing['location'].unique())
_1.columns = ['contact_tracing']

_2 = pd.DataFrame(testing['location'].unique())
_2.columns = ['testing']

_3 = pd.DataFrame(public_rules['location'].unique())
_3.columns = ['public_rules']

_4 = pd.DataFrame(gdppc['location'].unique())
_4.columns = ['gdppc']

_5 = pd.DataFrame(healthcare['location'].unique())
_5.columns = ['healthcare']

_6 = pd.DataFrame(age['location'].unique())
_6.columns = ['age']

_7 = pd.DataFrame(pop65['location'].unique())
_7.columns = ['pop65']

_8 = pd.DataFrame(pop['location'].unique())
_8.columns = ['pop']

_9 = pd.DataFrame(urban['location'].unique())
_9.columns = ['urban']

_10 = pd.DataFrame(healthcare_pc['location'].unique())
_10.columns = ['healthcare_pc']

_11 = pd.DataFrame(oop['location'].unique())
_11.columns = ['oop']

_12 = pd.DataFrame(diabetes['location'].unique())
_12.columns = ['diabetes']

_13 = pd.DataFrame(obesity['location'].unique())
_13.columns = ['obesity']

_14 = pd.DataFrame(gci['location'].unique())
_14.columns = ['gci']

fname=os.path.join(data_dir, 'mobility_google.pickle')
mobility_google = pickle.load( open( fname, "rb" ) )

fname=os.path.join(data_dir, 'mobility_apple.pickle')
mobility_apple = pickle.load( open( fname, "rb" ) )

fname=os.path.join(data_dir, 'case_stats.pickle')
cases = pickle.load( open( fname, "rb" ) )

_15 = pd.DataFrame(mobility_google.loc[(mobility_google['geo_level']=='country'),'location'].unique())
_15.columns = ['mobility_google']

_16 = pd.DataFrame(mobility_apple.loc[(mobility_apple['geo_level']=='country'),'location'].unique())
_16.columns = ['mobility_apple']

_17 = pd.DataFrame(cases.loc[(cases['geo_level']=='country'),'location'].unique())
_17.columns = ['case_stats']

country_list = pd.concat([_1,_2,_3,_4,_5,_6,_7,_8,_9,_10,_11,_12,_13,_14,_15,_16,_17],axis=1)

writer=pd.ExcelWriter(os.path.join(data_dir,'country_list.xlsx'))
writer.save()
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir,'country_list.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir,'country_list.xlsx'),engine='openpyxl')
writer.book=book
_0.to_excel(writer,'cases')
country_list.to_excel(writer,'country_list')
writer.save()

#################
# merge data
#################
# country master list
data_dir = '//rschfiler2.baml.com/UK-NYCOMMODITIES/Jamal/CovidAnalysis/Mobility'
fname = os.path.join(data_dir, 'country_match.xlsx')
countries = pd.read_excel(fname,skiprows=range(0))
countries.columns

fname=os.path.join(data_dir, 'cases_country.pickle')
cases_country = pickle.load( open( fname, "rb" ) )

# merge
def mergedat(dat,var):
    temp = countries[['cases_country', var]].dropna().reset_index(drop=True)
    temp.columns = ['cases_country','location']
    dat = pd.merge(dat,temp,on='location',how='left')
    dat['location'] = dat['cases_country']
    dat = dat.drop(['cases_country'],axis=1).dropna().reset_index(drop=True)
    return dat

age = mergedat(dat=age,var='age')
contact_tracing = mergedat(dat=contact_tracing,var='contact_tracing')
gdppc = mergedat(dat=gdppc,var='gdppc')
healthcare = mergedat(dat=healthcare,var='healthcare')
pop = mergedat(dat=pop,var='pop')
pop65 = mergedat(dat=pop65,var='pop65')
public_rules = mergedat(dat=public_rules,var='public_rules')
testing = mergedat(dat=testing,var='testing')
urban = mergedat(dat=urban,var='urban')
healthcare_pc = mergedat(dat=healthcare_pc,var='healthcare_pc')
oop = mergedat(dat=oop,var='oop')
diabetes = mergedat(dat=diabetes,var='diabetes')
obesity = mergedat(dat=obesity,var='obesity')
gci = mergedat(dat=gci,var='gci')

def mergevars(dat1,dat2,var):
    return pd.merge(dat1,dat2,on=var,how='left')

cases_country = mergevars(cases_country, contact_tracing, ['date','location'])
cases_country = mergevars(cases_country, testing, ['date','location'])
cases_country = mergevars(cases_country, public_rules, ['date','location'])
cases_country = mergevars(cases_country, age, ['location'])
cases_country = mergevars(cases_country, gdppc, ['location'])
cases_country = mergevars(cases_country, pop, ['location'])
cases_country = mergevars(cases_country, pop65, ['location'])
cases_country = mergevars(cases_country, urban, ['location'])
cases_country = mergevars(cases_country, healthcare, ['location'])
cases_country = mergevars(cases_country, healthcare_pc, ['location'])
cases_country = mergevars(cases_country, oop, ['location'])
cases_country = mergevars(cases_country, diabetes, ['location'])
cases_country = mergevars(cases_country, obesity, ['location'])
cases_country = mergevars(cases_country, gci, ['location'])

#fill missing
cases_country.loc[cases_country['total_tests'].isna(),'total_tests'] = 0.
cases_country.loc[cases_country['contact_tracing'].isna(),'contact_tracing'] = 1.0
cases_country.loc[cases_country['public_rules'].isna(),'public_rules'] = 1.0
cases_country.loc[cases_country['region'].isna(),'region'] = 'Asia'
#rename
cases_country.rename(columns = {'retail_and_recreation_percent_change_from_baseline' : 'retail_recreation',
                                'grocery_and_pharmacy_percent_change_from_baseline' : 'grocery_pharmacy',
                                'parks_percent_change_from_baseline' : 'parks',
                                'transit_stations_percent_change_from_baseline' : 'transit_stations',
                                'workplaces_percent_change_from_baseline' : 'workplaces',
                                'residential_percent_change_from_baseline' : 'residential'},
                    inplace=True)

#standardize mobility
cases_country.loc[cases_country['retail_recreation'].notna(),'retail_recreation'] += 100 
cases_country.loc[cases_country['grocery_pharmacy'].notna(),'grocery_pharmacy'] += 100 
cases_country.loc[cases_country['parks'].notna(),'parks'] += 100 
cases_country.loc[cases_country['transit_stations'].notna(),'transit_stations'] += 100 
cases_country.loc[cases_country['workplaces'].notna(),'workplaces'] += 100 
cases_country.loc[cases_country['residential'].notna(),'residential'] += 100 

with open(os.path.join(data_dir, 'cases_country.pickle'), 'wb') as output:
    pickle.dump(cases_country, output)

'''
writer=pd.ExcelWriter(os.path.join(data_dir,'cases_country_state.xlsx'))
writer.save()
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir,'cases_country_state.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir,'cases_country_state.xlsx'),engine='openpyxl')
writer.book=book
cases_country.to_excel(writer,'country')
writer.save()
'''

temp = cases_country[(cases_country['country']=='United States') & (cases_country['geo_level']=='country')].reset_index()

writer=pd.ExcelWriter(os.path.join(data_dir,'us_aggregate_ts.xlsx'))
writer.save()
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir,'us_aggregate_ts.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir,'us_aggregate_ts.xlsx'),engine='openpyxl')
writer.book=book
temp.to_excel(writer,'us')
writer.save()



