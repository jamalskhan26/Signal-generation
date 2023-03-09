# -*- coding: utf-8 -*-
from IPython import get_ipython
get_ipython().magic('reset -sf')

#############################################################
# Load/import data
#############################################################
import pandas as pd
import numpy as np
import os
import pickle
import statsmodels.api as sm
import datetime as dt
from datetime import timedelta  
from eqd import bbg
os.chdir('//rschfiler2.baml.com/UK-NYCOMMODITIES/Jamal')
import DataPrepFunctions as fns
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from linearmodels import PanelOLS
from linearmodels import RandomEffects

### set paths 
### !!UPDATE!! before running
data_dir = '//rschfiler2.baml.com/UK-NYCOMMODITIES/Jamal/CovidAnalysis/Mobility'

#################
# import data
#################

fname=os.path.join(data_dir, 'cases_state.pickle')
cases_state = pickle.load( open( fname, "rb" ) )

state_data_des = cases_state.describe()
state_data_des['N'] = len(cases_state.location.unique())
state_data_des['T'] = len(cases_state.date.unique())
state_data_des = state_data_des.T

writer=pd.ExcelWriter(os.path.join(data_dir, 'stateleveldescription.xlsx'))
writer.save() 
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir, 'stateleveldescription.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir, 'stateleveldescription.xlsx'),engine='openpyxl')
writer.book=book    
state_data_des.to_excel(writer,'des')
writer.save()


mobility_vars_g = ['retail_recreation','grocery_pharmacy','parks','transit_stations','workplaces','residential']
mobility_vars_a = ['driving']
control_vars = ['over65','median_income','median_age','pop_density','pop','household_size','tests']
lagged_dep = ['case_growth_1w']
#################
# data prep
#################

#transformations
cases_state['case_rate'] = cases_state['cases'] / cases_state['pop'] * 1000 
# death / recovery rates as of cases two weeks ago
cases_state['cases_2w'] =  cases_state[['location','cases']].groupby('location').shift(13)
cases_state['death_rate'] = cases_state['deaths'] / cases_state['cases_2w']
cases_state['cases_1w'] =  cases_state[['location','cases']].groupby('location').shift(8)
cases_state['case_growth_1w'] = (cases_state['cases'] / cases_state['cases_1w'])-1
cases_state['case_growth_1w'] = cases_state[['location','case_growth_1w']].groupby('location').shift(1)

#change in case rates, pct change in mobility
for c in cases_state['location'].unique():
    cases_state.loc[(cases_state['location']==c),'cases_d'] = (cases_state.loc[(cases_state['location']==c),'cases']).diff().values
    cases_state.loc[(cases_state['location']==c),'cases_ld'] = np.log(cases_state.loc[(cases_state['location']==c),'cases']).replace(np.inf,np.nan).replace(-np.inf,np.nan).diff().values
    cases_state.loc[(cases_state['location']==c),'case_rate_ld'] = np.log(cases_state.loc[(cases_state['location']==c),'case_rate']).replace(np.inf,np.nan).replace(-np.inf,np.nan).diff().values
    cases_state.loc[(cases_state['location']==c),'death_rate_d'] = cases_state.loc[(cases_state['location']==c),'death_rate'].diff().values
    cases_state.loc[(cases_state['location']==c),'death_rate_ld'] = np.log(cases_state.loc[(cases_state['location']==c),'death_rate']).replace(np.inf,np.nan).replace(-np.inf,np.nan).diff().values
#    cases_state.loc[(cases_state['location']==c), mobility_vars_g+mobility_vars_a] = np.log(cases_state.loc[(cases_state['location']==c),mobility_vars_g+mobility_vars_a]).diff().values

# run VIF test
from statsmodels.stats.outliers_influence import variance_inflation_factor
X = cases_state[control_vars].replace(np.inf,np.nan).replace(-np.inf,np.nan)
vif = pd.DataFrame()
vif["VIF Factor"] = [variance_inflation_factor(X.dropna().values, i) for i in range(X.shape[1])]
vif["features"] = X.columns
vif.round(1)

#remove median age, pop, gdppc
control_vars = ['const','over65','median_income','median_age','pop_density','household_size','tests']

#################
## regressions
#################

import sys

# shift mobility vars 
for i in range(1, 15, 1):
    lagged = cases_state[['location'] + mobility_vars_g+mobility_vars_a].groupby('location').shift(i)
    lagged.columns = [s + ('_' + str(i)) for s in list(lagged.columns)]
    colnames = list(cases_state.columns) + list(lagged.columns)
    cases_state = pd.concat([cases_state,lagged],axis=1, ignore_index=True)
    cases_state.columns = colnames

# set index for panel reg
cases_state['statecode'] = pd.Categorical(cases_state.location).codes
cases_state = cases_state.set_index(['date', 'statecode'])
cases_state['const']  = 1

'''
#to use later for testing
temp = cases_country.copy()
temp.index = temp['date']
split_index = round(len(temp.index.unique())*0.8)
split_date = temp.index.unique()[split_index]
df_train = temp.loc[temp.index <= split_date].copy()
df_test = temp.loc[temp.index > split_date].copy()
'''
'''
# determine optimal lag length
aic_df = {}
r2_fe = {}
r2_re = {}
lag_models_fe = {}
lag_models_re = {}
#Current minimum AIC score
min_aic = sys.float_info.max
#Current max R2
max_r2 = sys.float_info.min
#Lag for the best model seen so far
best_lag_aic = ''
best_lag_r2 = ''
best_mod_r2 = ''
#OLSResults objects for the best model seen so far
best_olsr_model_results = None

depvar_ = 'cases_ld'

#Run through each lag var   
for lag_num in range(0, 15):
    if lag_num == 0:
        colnames = mobility_vars_g+mobility_vars_a
    else:
        colnames = [s + ('_' + str(lag_num)) for s in mobility_vars_g+mobility_vars_a]
    
    print('Building model for lag: ' + str(lag_num))
    
    #Build and fit the OLSR model 
    olsr_results = sm.OLS(cases_state[depvar_], cases_state[colnames + control_vars + lagged_dep], missing = 'drop').fit()
    #Store away the model's AIC score and R2 
    aic_df[lag_num] = olsr_results.aic
    print('AIC='+str(aic_df[lag_num]))

   #If the model's AIC score is less than the current minimum score, update the current minimum AIC score and the current best model
    if olsr_results.aic < min_aic:
        min_aic = olsr_results.aic
        best_lag_aic = str(lag_num)

##Build and fit  fixed and random effects models 
    mod = PanelOLS(cases_state[depvar_], cases_state[colnames + control_vars + lagged_dep]).fit()
    r2_fe[lag_num] = mod.rsquared_overall
    lag_models_fe[lag_num] = mod

   #If the model's R2 is greater than the current max, update the current max and the current best model
    if mod.rsquared_overall > max_r2:
        max_r2 = mod.rsquared_overall
        best_lag_r2 = str(lag_num)
        best_mod_r2 = 'fe'

    mod = RandomEffects(cases_state[depvar_], cases_state[colnames + control_vars + lagged_dep]).fit()
    r2_re[lag_num] = mod.rsquared_overall    
    lag_models_re[lag_num] = mod

   #If the model's R2 is greater than the current max, update the current max and the current best model
    if mod.rsquared_overall > max_r2:
        max_r2 = mod.rsquared_overall
        best_lag_r2 = str(lag_num)
        best_mod_r2 = 're'
'''

#######################################################################

'''
writer=pd.ExcelWriter(os.path.join(data_dir,'cases_country.xlsx'))
writer.save()
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir,'cases_country.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir,'cases_country.xlsx'),engine='openpyxl')
writer.book=book
cases_country.to_excel(writer,'country')
writer.save()

writer=pd.ExcelWriter(os.path.join(data_dir,'correl.xlsx'))
writer.save()
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir,'correl.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir,'correl.xlsx'),engine='openpyxl')
writer.book=book
cases_country[mobility_vars_g+mobility_vars_a+control_vars].corr().to_excel(writer,'exog_corr')
writer.save()
'''

############################
# iterative models for mobility
############################

dep_vars = ['case_rate_ld','death_rate_ld']

mobility_vars_g = ['retail_recreation','grocery_pharmacy','parks','transit_stations','workplaces','residential']
mobility_vars_a = ['driving']
mobility_vars_g_7 = [s + ('_' + str(7)) for s in mobility_vars_g]
mobility_vars_a_7 = [s + ('_' + str(7)) for s in mobility_vars_a]

def run_reg(dep_var,mob_g,mob_a):
    reg_results = {}
    for mob_var in (mob_g+mob_a):
        mod = PanelOLS(cases_state[dep_var], cases_state[lagged_dep+control_vars+[mob_var]]).fit()
        temp = mod.summary.tables[1].as_html()
        reg_results[mob_var] = pd.read_html(temp, header=0, index_col=0)[0]           
        reg_results[mob_var].loc['R2_within','Parameter'] =  mod.rsquared_within
        reg_results[mob_var].loc['R2_between','Parameter'] = mod.rsquared_between
        reg_results[mob_var].loc['R2_overall','Parameter'] = mod.rsquared_overall

    mod = PanelOLS(cases_state[dep_var], cases_state[lagged_dep+control_vars+mob_g+mob_a]).fit()
    temp = mod.summary.tables[1].as_html()
    reg_results['all'] = pd.read_html(temp, header=0, index_col=0)[0]            
    reg_results['all'].loc['R2_within','Parameter'] = mod.rsquared_within
    reg_results['all'].loc['R2_between','Parameter'] = mod.rsquared_between
    reg_results['all'].loc['R2_overall','Parameter'] = mod.rsquared_overall

    mod = PanelOLS(cases_state[dep_var], cases_state[lagged_dep+control_vars+mob_g]).fit()
    temp = mod.summary.tables[1].as_html()
    reg_results['google'] = pd.read_html(temp, header=0, index_col=0)[0]
    reg_results['google'].loc['R2_within','Parameter'] = mod.rsquared_within
    reg_results['google'].loc['R2_between','Parameter'] = mod.rsquared_between
    reg_results['google'].loc['R2_overall','Parameter'] = mod.rsquared_overall

    mod = PanelOLS(cases_state[dep_var], cases_state[lagged_dep+control_vars+mob_a]).fit()
    temp = mod.summary.tables[1].as_html()
    reg_results['apple'] = pd.read_html(temp, header=0, index_col=0)[0]            
    reg_results['apple'].loc['R2_within','Parameter'] = mod.rsquared_within
    reg_results['apple'].loc['R2_between','Parameter'] = mod.rsquared_between
    reg_results['apple'].loc['R2_overall','Parameter'] = mod.rsquared_overall
            
    return reg_results

reg_cases =      run_reg(dep_var = dep_vars[0], mob_g = mobility_vars_g, mob_a = mobility_vars_a)
reg_cases_6 =  run_reg(dep_var = dep_vars[0], mob_g = mobility_vars_g_7, mob_a = mobility_vars_a_7)

mobility_vars_a_g = ['transit_stations','driving','residential']
mobility_vars_a_g_7 = [s + ('_' + str(7)) for s in mobility_vars_a_g]

def run_reg_custom(dep_var,mob):
    mod = PanelOLS(cases_state[dep_var], cases_state[lagged_dep+control_vars+mob]).fit()
    temp = mod.summary.tables[1].as_html()
    reg_results = pd.read_html(temp, header=0, index_col=0)[0]            
    reg_results.loc['R2_within','Parameter'] = mod.rsquared_within
    reg_results.loc['R2_between','Parameter'] = mod.rsquared_between
    reg_results.loc['R2_overall','Parameter'] = mod.rsquared_overall
    return reg_results

reg_cases['driving_transit'] = run_reg_custom(dep_vars[0],mobility_vars_a_g)
reg_cases['driving_transit1'] = run_reg_custom(dep_vars[0],mobility_vars_a_g_7)

writer=pd.ExcelWriter(os.path.join(data_dir, 'reg_cases_states.xlsx'))
writer.save() 
for p in reg_cases.keys(): 
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, 'reg_cases_states.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, 'reg_cases_states.xlsx'),engine='openpyxl')
    writer.book=book    
    reg_cases[p].to_excel(writer,p)
    writer.save()
for p in reg_cases_6.keys(): 
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, 'reg_cases_states.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, 'reg_cases_states.xlsx'),engine='openpyxl')
    writer.book=book    
    reg_cases_6[p].to_excel(writer,p)
    writer.save()

