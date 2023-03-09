# -*- coding: utf-8 -*-
from IPython import get_ipython
get_ipython().magic('reset -sf')

#############################################################
# Load/import data
#############################################################
import pandas as pd
import numpy as np
import os
import pickle
import statsmodels.api as sm
import datetime as dt
from datetime import timedelta  
from eqd import bbg
os.chdir('//rschfiler2.baml.com/UK-NYCOMMODITIES/Jamal')
import DataPrepFunctions as fns
import matplotlib.pyplot as plt

import numpy as np
import pandas as pd
from linearmodels import PanelOLS
from linearmodels import RandomEffects


### set paths 
### !!UPDATE!! before running
data_dir = '//rschfiler2.baml.com/UK-NYCOMMODITIES/Jamal/CovidAnalysis/Mobility'

#################
# import data
#################

fname=os.path.join(data_dir, 'cases_country.pickle')
cases_country = pickle.load( open( fname, "rb" ) )

mobility_vars_g = ['retail_recreation','grocery_pharmacy','parks','transit_stations','workplaces','residential']
mobility_vars_a = ['driving','transit','walking']
control_vars = ['median_age','gdppc','pop_per_sq_m','pop','pop65','urban','health_access','healthcare_pc','oop','diabetes','obesity','gci','contact_tracing','total_tests','public_rules','Lat','Long']
lagged_dep = ['case_growth_1w']

stdev = cases_country[['location','median_age','gdppc','pop_per_sq_m','pop','pop65','urban','health_access','healthcare_pc','oop','diabetes','obesity','gci']].drop_duplicates()
stdev = stdev.iloc[:,1:].std() 

writer=pd.ExcelWriter(os.path.join(data_dir, 'stdev_controls.xlsx'))
writer.save() 
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir, 'stdev_controls.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir, 'stdev_controls.xlsx'),engine='openpyxl')
writer.book=book    
stdev.to_excel(writer,'stdev_controls')
writer.save()


#################
# data prep
#################

#transformations
cases_country['case_rate'] = cases_country['cases'] / cases_country['pop'] * 1000 
# death / recovery rates as of cases two weeks ago
cases_country['cases_2w'] =  cases_country[['location','cases']].groupby('location').shift(13)
cases_country['death_rate'] = cases_country['deaths'] / cases_country['cases']
cases_country['recovery_rate'] = cases_country['recovered'] / cases_country['cases']
cases_country['cases_1w'] =  cases_country[['location','cases']].groupby('location').shift(8)
cases_country['case_growth_1w'] = (cases_country['cases'] / cases_country['cases_1w'])-1
cases_country['case_growth_1w'] = cases_country[['location','case_growth_1w']].groupby('location').shift(1)
cases_country['case_growth_1w'] = cases_country['case_growth_1w'].replace(np.inf,np.nan).replace(-np.inf,np.nan)

country_data_des = cases_country.describe()
country_data_des['countries'] = str(cases_country.location.unique())
country_data_des['N'] = len(cases_country.location.unique())
country_data_des['T'] = len(cases_country.date.unique())
country_data_des = country_data_des.T

writer=pd.ExcelWriter(os.path.join(data_dir, 'countryleveldescription.xlsx'))
writer.save() 
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir, 'countryleveldescription.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir, 'countryleveldescription.xlsx'),engine='openpyxl')
writer.book=book    
country_data_des.to_excel(writer,'des')
writer.save()


top10_cases = ['United States','United Kingdom','Germany'	,'France'	,'Italy','Spain',	'Brazil',	'Russia','India','Turkey','Mexico']
top10_cases_em = ['Brazil',	'Russia',	'India',	'Turkey',	'Iran',	'Peru',	'Saudi Arabia',	'Chile',	'Mexico',	'Pakistan']
bestresponse_cases = ['Korea, South', 'Singapore',	'Hong Kong','Australia','New Zealand','Taiwan','Finland','Denmark','Belgium','Sweden']

top10_dat = cases_country[cases_country['location'].isin(top10_cases)].reset_index(drop=True)
top10_em_dat = cases_country[cases_country['location'].isin(top10_cases_em)].reset_index(drop=True)
best_dat = cases_country[cases_country['location'].isin(bestresponse_cases)].reset_index(drop=True)

writer=pd.ExcelWriter(os.path.join(data_dir, 'country_mob_trends_top10.xlsx'))
writer.save() 
for mob in mobility_vars_g+mobility_vars_a:
    temp = pd.pivot_table(top10_dat[['date','location']+[mob]], index=['date'], columns=['location'], values=[mob], fill_value=np.nan).dropna(how='all')
    temp.iloc[7:,:] = temp.rolling(7).mean().iloc[7:,:]
    book= load_workbook(os.path.join(data_dir, 'country_mob_trends_top10.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, 'country_mob_trends_top10.xlsx'),engine='openpyxl')
    writer.book=book    
    temp.to_excel(writer,mob)
    writer.save()

writer=pd.ExcelWriter(os.path.join(data_dir, 'country_mob_trends_best10.xlsx'))
writer.save() 
for mob in mobility_vars_g+mobility_vars_a:
    temp = pd.pivot_table(best_dat[['date','location']+[mob]], index=['date'], columns=['location'], values=[mob], fill_value=np.nan).dropna(how='all')
    temp.iloc[7:,:] = temp.rolling(7).mean().iloc[7:,:]
    book= load_workbook(os.path.join(data_dir, 'country_mob_trends_best10.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, 'country_mob_trends_best10.xlsx'),engine='openpyxl')
    writer.book=book    
    temp.to_excel(writer,mob)
    writer.save()


#change in case rates, pct change in mobility
for c in cases_country['location'].unique():
    cases_country.loc[(cases_country['location']==c) & (cases_country['geo_level']=='country'),'cases_d'] = (cases_country.loc[(cases_country['location']==c) & (cases_country['geo_level']=='country'),'cases']).diff().values
    cases_country.loc[(cases_country['location']==c) & (cases_country['geo_level']=='country'),'cases_ld'] = np.log(cases_country.loc[(cases_country['location']==c) & (cases_country['geo_level']=='country'),'cases']).replace(np.inf,np.nan).replace(-np.inf,np.nan).diff().values
    cases_country.loc[(cases_country['location']==c) & (cases_country['geo_level']=='country'),'case_rate_ld'] = np.log(cases_country.loc[(cases_country['location']==c) & (cases_country['geo_level']=='country'),'case_rate']).replace(np.inf,np.nan).replace(-np.inf,np.nan).diff().values
    cases_country.loc[(cases_country['location']==c) & (cases_country['geo_level']=='country'),'death_rate_d'] = cases_country.loc[(cases_country['location']==c) & (cases_country['geo_level']=='country'),'death_rate'].diff().values
    cases_country.loc[(cases_country['location']==c) & (cases_country['geo_level']=='country'),'death_rate_ld'] = np.log(cases_country.loc[(cases_country['location']==c) & (cases_country['geo_level']=='country'),'death_rate']).replace(np.inf,np.nan).replace(-np.inf,np.nan).diff().values
    cases_country.loc[(cases_country['location']==c) & (cases_country['geo_level']=='country'),'recovery_rate_d'] = cases_country.loc[(cases_country['location']==c) & (cases_country['geo_level']=='country'),'recovery_rate'].diff().values
#    cases_country.loc[(cases_country['location']==c) & (cases_country['geo_level']=='country'), mobility_vars_g+mobility_vars_a] = (cases_country.loc[(cases_country['location']==c) & (cases_country['geo_level']=='country'),mobility_vars_g+mobility_vars_a]).div((cases_country.loc[(cases_country['location']==c) & (cases_country['geo_level']=='country'),mobility_vars_g+mobility_vars_a]).shift(1).values)

# run VIF test
from statsmodels.stats.outliers_influence import variance_inflation_factor
X = cases_country[['median_age','gdppc','pop_per_sq_m','pop','pop65','urban','health_access','healthcare_pc','oop','diabetes','obesity','gci','contact_tracing','total_tests','public_rules','Lat','Long']].replace(np.inf,np.nan).replace(-np.inf,np.nan)
vif = pd.DataFrame()
vif["VIF Factor"] = [variance_inflation_factor(X.dropna().values, i) for i in range(X.shape[1])]
vif["features"] = X.columns
vif.round(1)

#remove median age, pop, gdppc
control_vars = ['const','pop_per_sq_m','pop65','urban','health_access','healthcare_pc','oop','diabetes','obesity','gci','contact_tracing','total_tests','public_rules','Lat','Long']

#################
## regressions
#################

import sys

# shift mobility vars 
for i in range(1, 10, 1):
    lagged = cases_country[['location'] + mobility_vars_g+mobility_vars_a].groupby('location').shift(i)
    lagged.columns = [s + ('_' + str(i)) for s in list(lagged.columns)]
    colnames = list(cases_country.columns) + list(lagged.columns)
    cases_country = pd.concat([cases_country,lagged],axis=1, ignore_index=True)
    cases_country.columns = colnames

cases_country_testing = cases_country[cases_country['total_tests']!=0].reset_index(drop=True)

# set index for panel reg
cases_country['countrycode'] = pd.Categorical(cases_country.location).codes
cases_country = cases_country.set_index(['date', 'countrycode'])
cases_country['const']  = 1

cases_country_testing['countrycode'] = pd.Categorical(cases_country_testing.location).codes
cases_country_testing = cases_country_testing.set_index(['date', 'countrycode'])
cases_country_testing['const']  = 1

'''
#to use later for testing
temp = cases_country.copy()
temp.index = temp['date']
split_index = round(len(temp.index.unique())*0.8)
split_date = temp.index.unique()[split_index]
df_train = temp.loc[temp.index <= split_date].copy()
df_test = temp.loc[temp.index > split_date].copy()
'''

'''
# determine optimal lag length
aic_df = {}
r2_fe = {}
r2_re = {}
lag_models_fe = {}
lag_models_re = {}
#Current minimum AIC score
min_aic = sys.float_info.max
#Current max R2
max_r2 = sys.float_info.min
#Lag for the best model seen so far
best_lag_aic = ''
best_lag_r2 = ''
best_mod_r2 = ''
#OLSResults objects for the best model seen so far
best_olsr_model_results = None

depvar_ = 'cases_ld'

#Run through each lag var   
for lag_num in range(0, 10):
    if lag_num == 0:
        colnames = mobility_vars_g+mobility_vars_a
    else:
        colnames = [s + ('_' + str(lag_num)) for s in mobility_vars_g+mobility_vars_a]
    
    print('Building model for lag: ' + str(lag_num))
    
    #Build and fit the OLSR model 
    olsr_results = sm.OLS(cases_country[depvar_], cases_country[colnames + control_vars + lagged_dep], missing = 'drop').fit()
    #Store away the model's AIC score and R2 
    aic_df[lag_num] = olsr_results.aic
    print('AIC='+str(aic_df[lag_num]))

   #If the model's AIC score is less than the current minimum score, update the current minimum AIC score and the current best model
    if olsr_results.aic < min_aic:
        min_aic = olsr_results.aic
        best_lag_aic = str(lag_num)

##Build and fit  fixed and random effects models 
    mod = PanelOLS(cases_country[depvar_], cases_country[colnames + control_vars + lagged_dep]).fit()
    r2_fe[lag_num] = mod.rsquared_overall
    lag_models_fe[lag_num] = mod

   #If the model's R2 is greater than the current max, update the current max and the current best model
    if mod.rsquared_overall > max_r2:
        max_r2 = mod.rsquared_overall
        best_lag_r2 = str(lag_num)
        best_mod_r2 = 'fe'

    mod = RandomEffects(cases_country[depvar_], cases_country[colnames + control_vars + lagged_dep]).fit()
    r2_re[lag_num] = mod.rsquared_overall    
    lag_models_re[lag_num] = mod

   #If the model's R2 is greater than the current max, update the current max and the current best model
    if mod.rsquared_overall > max_r2:
        max_r2 = mod.rsquared_overall
        best_lag_r2 = str(lag_num)
        best_mod_r2 = 're'
'''

#######################################################################

'''
writer=pd.ExcelWriter(os.path.join(data_dir,'cases_country.xlsx'))
writer.save()
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir,'cases_country.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir,'cases_country.xlsx'),engine='openpyxl')
writer.book=book
cases_country.to_excel(writer,'country')
writer.save()

writer=pd.ExcelWriter(os.path.join(data_dir,'correl.xlsx'))
writer.save()
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir,'correl.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir,'correl.xlsx'),engine='openpyxl')
writer.book=book
cases_country[mobility_vars_g+mobility_vars_a+control_vars].corr().to_excel(writer,'exog_corr')
writer.save()
'''

############################
# iterative models for mobility
############################

dep_vars = ['case_rate_ld','death_rate_ld']

mobility_vars_g = ['retail_recreation','grocery_pharmacy','parks','transit_stations','workplaces','residential']
mobility_vars_a = ['driving','transit','walking']
mobility_vars_g_7 = [s + ('_' + str(7)) for s in mobility_vars_g]
mobility_vars_a_7 = [s + ('_' + str(7)) for s in mobility_vars_a]

def run_reg(dat,dep_var,mob_g,mob_a):
    reg_results = {}
    for mob_var in (mob_g+mob_a):
        mod = PanelOLS(dat[dep_var], dat[control_vars + lagged_dep +[mob_var]]).fit()
        temp = mod.summary.tables[1].as_html()
        reg_results[mob_var] = pd.read_html(temp, header=0, index_col=0)[0]            
        reg_results[mob_var].loc['R2_within','Parameter'] =  mod.rsquared_within
        reg_results[mob_var].loc['R2_between','Parameter'] = mod.rsquared_between
        reg_results[mob_var].loc['R2_overall','Parameter'] = mod.rsquared_overall

    mod = PanelOLS(dat[dep_var], dat[control_vars + lagged_dep +mob_g+mob_a]).fit()
    temp = mod.summary.tables[1].as_html()
    reg_results['all'] = pd.read_html(temp, header=0, index_col=0)[0]            
    reg_results['all'].loc['R2_within','Parameter'] = mod.rsquared_within
    reg_results['all'].loc['R2_between','Parameter'] = mod.rsquared_between
    reg_results['all'].loc['R2_overall','Parameter'] = mod.rsquared_overall

    mod = PanelOLS(dat[dep_var], dat[control_vars + lagged_dep +mob_g]).fit()
    temp = mod.summary.tables[1].as_html()
    reg_results['google'] = pd.read_html(temp, header=0, index_col=0)[0]
    reg_results['google'].loc['R2_within','Parameter'] = mod.rsquared_within
    reg_results['google'].loc['R2_between','Parameter'] = mod.rsquared_between
    reg_results['google'].loc['R2_overall','Parameter'] = mod.rsquared_overall

    mod = PanelOLS(dat[dep_var], dat[control_vars + lagged_dep +mob_a]).fit()
    temp = mod.summary.tables[1].as_html()
    reg_results['apple'] = pd.read_html(temp, header=0, index_col=0)[0]            
    reg_results['apple'].loc['R2_within','Parameter'] = mod.rsquared_within
    reg_results['apple'].loc['R2_between','Parameter'] = mod.rsquared_between
    reg_results['apple'].loc['R2_overall','Parameter'] = mod.rsquared_overall
            
    return reg_results

reg_cases =    run_reg(dat = cases_country, dep_var = dep_vars[0], mob_g = mobility_vars_g, mob_a = mobility_vars_a)
reg_cases_6 =  run_reg(dat = cases_country, dep_var = dep_vars[0], mob_g = mobility_vars_g_7, mob_a = mobility_vars_a_7)

writer=pd.ExcelWriter(os.path.join(data_dir, 'reg_cases.xlsx'))
writer.save() 
for p in reg_cases.keys(): 
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, 'reg_cases.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, 'reg_cases.xlsx'),engine='openpyxl')
    writer.book=book    
    reg_cases[p].to_excel(writer,p)
    writer.save()
for p in reg_cases_6.keys(): 
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, 'reg_cases.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, 'reg_cases.xlsx'),engine='openpyxl')
    writer.book=book    
    reg_cases_6[p].to_excel(writer,p)
    writer.save()

reg_deaths =  run_reg(dat = cases_country, dep_var = dep_vars[1], mob_g = mobility_vars_g, mob_a = mobility_vars_a)
reg_deaths_6 =  run_reg(dat = cases_country, dep_var = dep_vars[1], mob_g = mobility_vars_g_7, mob_a = mobility_vars_a_7)

writer=pd.ExcelWriter(os.path.join(data_dir, 'reg_deaths.xlsx'))
writer.save() 
for p in reg_deaths.keys(): 
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, 'reg_deaths.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, 'reg_deaths.xlsx'),engine='openpyxl')
    writer.book=book    
    reg_deaths[p].to_excel(writer,p)
    writer.save()
for p in reg_deaths_6.keys(): 
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, 'reg_deaths.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, 'reg_deaths.xlsx'),engine='openpyxl')
    writer.book=book    
    reg_deaths_6[p].to_excel(writer,p)
    writer.save()


####################################################################################

def run_reg_re(dat,dep_var,mob_g,mob_a):
    reg_results = {}
    for mob_var in (mob_g+mob_a):
        mod = RandomEffects(dat[dep_var], dat[control_vars + lagged_dep +[mob_var]]).fit()
        temp = mod.summary.tables[1].as_html()
        reg_results[mob_var] = pd.read_html(temp, header=0, index_col=0)[0]            
        reg_results[mob_var].loc['R2_within','Parameter'] =  mod.rsquared_within
        reg_results[mob_var].loc['R2_between','Parameter'] = mod.rsquared_between
        reg_results[mob_var].loc['R2_overall','Parameter'] = mod.rsquared_overall

    mod = RandomEffects(dat[dep_var], dat[control_vars + lagged_dep +mob_g+mob_a]).fit()
    temp = mod.summary.tables[1].as_html()
    reg_results['all'] = pd.read_html(temp, header=0, index_col=0)[0]            
    reg_results['all'].loc['R2_within','Parameter'] = mod.rsquared_within
    reg_results['all'].loc['R2_between','Parameter'] = mod.rsquared_between
    reg_results['all'].loc['R2_overall','Parameter'] = mod.rsquared_overall

    mod = RandomEffects(dat[dep_var], dat[control_vars + lagged_dep +mob_g]).fit()
    temp = mod.summary.tables[1].as_html()
    reg_results['google'] = pd.read_html(temp, header=0, index_col=0)[0]
    reg_results['google'].loc['R2_within','Parameter'] = mod.rsquared_within
    reg_results['google'].loc['R2_between','Parameter'] = mod.rsquared_between
    reg_results['google'].loc['R2_overall','Parameter'] = mod.rsquared_overall

    mod = RandomEffects(dat[dep_var], dat[control_vars + lagged_dep +mob_a]).fit()
    temp = mod.summary.tables[1].as_html()
    reg_results['apple'] = pd.read_html(temp, header=0, index_col=0)[0]            
    reg_results['apple'].loc['R2_within','Parameter'] = mod.rsquared_within
    reg_results['apple'].loc['R2_between','Parameter'] = mod.rsquared_between
    reg_results['apple'].loc['R2_overall','Parameter'] = mod.rsquared_overall
            
    return reg_results

reg_cases =    run_reg_re(dat = cases_country, dep_var = dep_vars[0], mob_g = mobility_vars_g, mob_a = mobility_vars_a)
reg_cases_6 =  run_reg_re(dat = cases_country, dep_var = dep_vars[0], mob_g = mobility_vars_g_7, mob_a = mobility_vars_a_7)

writer=pd.ExcelWriter(os.path.join(data_dir, 'reg_cases_re.xlsx'))
writer.save() 
for p in reg_cases.keys(): 
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, 'reg_cases_re.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, 'reg_cases_re.xlsx'),engine='openpyxl')
    writer.book=book    
    reg_cases[p].to_excel(writer,p)
    writer.save()
for p in reg_cases_6.keys(): 
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, 'reg_cases_re.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, 'reg_cases_re.xlsx'),engine='openpyxl')
    writer.book=book    
    reg_cases_6[p].to_excel(writer,p)
    writer.save()

####################################################################################

top10_cases = ['United States','United Kingdom','Germany'	,'France'	,'Italy','Spain',	'Brazil',	'Russia','India','Turkey']
top10_cases_em = ['Brazil',	'Russia',	'India',	'Turkey',	'Iran',	'Peru',	'Saudi Arabia',	'Chile',	'Mexico',	'Pakistan']
bestresponse_cases = ['Korea, South', 'Singapore',	'Hong Kong','Australia','New Zealand','Taiwan','Finland','Denmark','Belgium','Sweden']

top10_dat = cases_country[cases_country['location'].isin(top10_cases)].reset_index()
top10_em_dat = cases_country[cases_country['location'].isin(top10_cases_em)].reset_index()
best_dat = cases_country[cases_country['location'].isin(bestresponse_cases)].reset_index()

# set index for panel reg
top10_dat['countrycode'] = pd.Categorical(top10_dat.location).codes
top10_dat = top10_dat.set_index(['date', 'countrycode'])
top10_dat['const']  = 1

# set index for panel reg
best_dat['countrycode'] = pd.Categorical(best_dat.location).codes
best_dat = best_dat.set_index(['date', 'countrycode'])
best_dat['const']  = 1

top10_dat['location'].unique()
best_dat['location'].unique()

reg_cases_top10 =  run_reg(dat = top10_dat, dep_var = dep_vars[0], mob_g = mobility_vars_g_7, mob_a = mobility_vars_a_7)
reg_cases_best =  run_reg(dat = best_dat, dep_var = dep_vars[0], mob_g = mobility_vars_g_7, mob_a = mobility_vars_a_7)

####################################################################################

reg_cases_testing = run_reg(dat = cases_country_testing, dep_var = dep_vars[0], mob_g = mobility_vars_g, mob_a = mobility_vars_a)

writer=pd.ExcelWriter(os.path.join(data_dir, 'reg_cases_testing.xlsx'))
writer.save() 
for p in reg_cases_testing .keys(): 
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, 'reg_cases_testing.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, 'reg_cases_testing.xlsx'),engine='openpyxl')
    writer.book=book    
    reg_cases_testing[p].to_excel(writer,p)
    writer.save()
