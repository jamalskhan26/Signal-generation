# -*- coding: utf-8 -*-
from IPython import get_ipython
get_ipython().magic('reset -sf')

#############################################################
# Load/import data
#############################################################
import pandas as pd
import numpy as np
import os
import pickle
import statsmodels.api as sm
import datetime as dt
from datetime import timedelta  
from eqd import bbg
os.chdir('//rschfiler2.baml.com/UK-NYCOMMODITIES/Jamal')
import DataPrepFunctions as fns
import numpy as np
import pandas as pd
from linearmodels import PanelOLS
from linearmodels import RandomEffects

### set paths 
### !!UPDATE!! before running
data_dir = '//rschfiler2.baml.com/UK-NYCOMMODITIES/Jamal/CovidAnalysis/Mobility'

#################
# import data
#################

fname=os.path.join(data_dir, 'cases_county.pickle')
cases_county = pickle.load( open( fname, "rb" ) )

cases_county_des = cases_county.describe()
cases_county_des['N'] = len(cases_county.location.unique())
cases_county_des['T'] = len(cases_county.date.unique())
cases_county_des = cases_county_des.T

writer=pd.ExcelWriter(os.path.join(data_dir, 'countyleveldescription.xlsx'))
writer.save() 
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir, 'countyleveldescription.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir, 'countyleveldescription.xlsx'),engine='openpyxl')
writer.book=book    
cases_county_des.to_excel(writer,'des')
writer.save()


mobility_vars_g = ['retail_recreation','grocery_pharmacy','parks','transit_stations','workplaces','residential']
mobility_vars_a = ['driving']

#transformations

cases_county['case_rate'] = cases_county['cases'] / cases_county['county_pop'] * 1000 

# death / recovery rates / case growth
cases_county['cases_2w'] =  cases_county[['location','cases']].groupby('location').shift(14)
cases_county['death_rate'] = cases_county['deaths'] / cases_county['cases']
cases_county['recovery_rate'] = cases_county['recovered'] / cases_county['cases']
cases_county['cases_1w'] =  cases_county[['location','cases']].groupby('location').shift(7)
cases_county['case_growth_1w'] = (np.log(cases_county['cases']) - np.log(cases_county['cases_1w'])).replace(np.inf,np.nan).replace(-np.inf,np.nan)
cases_county['case_growth_1w'] = cases_county[['location','case_growth_1w']].groupby('location').shift(1)

county_data_des = cases_county.describe()
county_data_des['countries'] = str(cases_county.location.unique())
county_data_des['N'] = len(cases_county.location.unique())
county_data_des['T'] = len(cases_county.date.unique())
county_data_des = county_data_des.T

writer=pd.ExcelWriter(os.path.join(data_dir, 'countyleveldescription.xlsx'))
writer.save() 
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir, 'countyleveldescription.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir, 'countyleveldescription.xlsx'),engine='openpyxl')
writer.book=book    
county_data_des.to_excel(writer,'des')
writer.save()


for c in cases_county['location'].unique():
    cases_county.loc[(cases_county['location']==c) & (cases_county['geo_level']=='county'),'cases_d'] = (cases_county.loc[(cases_county['location']==c) & (cases_county['geo_level']=='county'),'cases']).diff().values
    cases_county.loc[(cases_county['location']==c) & (cases_county['geo_level']=='county'),'case_rate_ld'] = np.log(cases_county.loc[(cases_county['location']==c) & (cases_county['geo_level']=='county'),'case_rate']).replace(np.inf,np.nan).replace(-np.inf,np.nan).diff().values
    cases_county.loc[(cases_county['location']==c) & (cases_county['geo_level']=='county'),'deaths_d'] = (cases_county.loc[(cases_county['location']==c) & (cases_county['geo_level']=='county'),'deaths']).diff().values
    cases_county.loc[(cases_county['location']==c) & (cases_county['geo_level']=='county'),'death_rate_ld'] = np.log(cases_county.loc[(cases_county['location']==c) & (cases_county['geo_level']=='county'),'deaths']).replace(np.inf,np.nan).replace(-np.inf,np.nan).diff().values


with open(os.path.join(data_dir, 'cases_county.pickle'), 'wb') as output:
    pickle.dump(cases_county, output)

#dim 1
county_summary = cases_county.groupby('location')[['location','case_rate']].tail(1).reset_index(drop=True)

#dim 2
temp = cases_county.groupby('location')[['location','case_rate_ld']].tail(30).reset_index(drop=True)
#temp = temp.groupby('location')[['location','case_rate_ld']].mean().reset_index()
temp = temp.groupby('location')[['location','case_rate_ld']].apply(lambda x: x.ewm(alpha=0.1, adjust=False).mean()).reset_index(drop=True)
temp = temp.groupby('location')[['location','case_rate_ld']].tail(1).reset_index(drop=True)

county_summary = pd.merge(county_summary,temp,on='location',how='left')

#dim 3
temp = cases_county.groupby('location')[['location','death_rate']].tail(1).reset_index(drop=True)
county_summary = pd.merge(county_summary,temp,on='location',how='left')

#dim 4
temp = cases_county.groupby('location')[['location','death_rate_ld']].tail(30).reset_index(drop=True)
#temp = temp.groupby('location')[['location','death_rate_ld']].mean().reset_index()
temp = temp.groupby('location')[['location','death_rate_ld']].apply(lambda x: x.ewm(alpha=0.1, adjust=False).mean()).reset_index(drop=True)
temp = temp.groupby('location')[['location','death_rate_ld']].tail(1).reset_index(drop=True)

county_summary = pd.merge(county_summary,temp,on='location',how='left')


#import FIPS
data_dir = '//rschfiler2.baml.com/UK-NYCOMMODITIES/Jamal/CovidAnalysis/Mobility'
fname = os.path.join(data_dir, 'county_match.xlsx')
fips = pd.read_excel(fname,skiprows=range(0))
fips.columns
fips = fips[['case_stats','FIPS']]
fips.columns = ['location','FIPS']
fips['FIPS'] = fips['FIPS'].map(lambda x: str(x).lstrip("'"))

#county_summary = county_summary.drop(['FIPS'],axis=1) 
county_summary = pd.merge(county_summary, fips, on ='location', how = 'left')    
    
with open(os.path.join(data_dir, 'county_summary.pickle'), 'wb') as output:
    pickle.dump(county_summary, output)

##############DRIVING

fname=os.path.join(data_dir, 'county_summary.pickle')
county_summary = pickle.load( open( fname, "rb" ) )

fname=os.path.join(data_dir, 'cases_county.pickle')
cases_county = pickle.load( open( fname, "rb" ) )


driving_county = cases_county[['date','location','geo_level','apple','driving']]
driving_county = driving_county[driving_county['apple']==1].reset_index(drop=True)

driving_county_1 = driving_county[(driving_county['date']>='2020-03-01') & (driving_county['date']<='2020-04-10')].reset_index(drop=True)
driving_county_2 = driving_county[(driving_county['date']>'2020-04-10')].reset_index(drop=True)

driving_county_avg_1 = pd.DataFrame()
for c in driving_county_1['location'].unique():
    driving_county_avg_1.loc[c,'average_driving'] = (driving_county_1.loc[(driving_county_1['location']==c) & (driving_county_1['geo_level']=='county'),'driving']).mean()
driving_county_avg_1 = driving_county_avg_1.reset_index()
driving_county_avg_1.columns = ['location','driving'] 
driving_county_avg_1 = pd.merge(driving_county_avg_1,county_summary[['location','FIPS']],on='location',how='left')
with open(os.path.join(data_dir, 'driving_county_avg_1.pickle'), 'wb') as output:
    pickle.dump(driving_county_avg_1, output)

driving_county_avg_2 = pd.DataFrame()
for c in driving_county_2['location'].unique():
    driving_county_avg_2.loc[c,'average_driving'] = (driving_county_2.loc[(driving_county_1['location']==c) & (driving_county_2['geo_level']=='county'),'driving']).mean()
driving_county_avg_2 = driving_county_avg_2.reset_index()
driving_county_avg_2.columns = ['location','driving'] 
driving_county_avg_2 = pd.merge(driving_county_avg_2,county_summary[['location','FIPS']],on='location',how='left')
with open(os.path.join(data_dir, 'driving_county_avg_2.pickle'), 'wb') as output:
    pickle.dump(driving_county_avg_2, output)

#######################################################

    
'''
#K-means

from sklearn.model_selection import train_test_split
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.20)

from sklearn.preprocessing import StandardScaler
scaler = StandardScaler()
scaler.fit(X_train)

X_train = scaler.transform(X_train)
X_test = scaler.transform(X_test)

from sklearn.neighbors import KNeighborsClassifier
classifier = KNeighborsClassifier(n_neighbors=5)
classifier.fit(X_train, y_train)

y_pred = classifier.predict(X_test)
'''

import matplotlib.pyplot as plt

clusterdata = county_summary.dropna().reset_index(drop=True) 
#remove outliers
from scipy import stats
clusterdata['z_score_cases']=stats.zscore(clusterdata['case_rate_ld'])
clusterdata = clusterdata.loc[clusterdata['z_score_cases'].abs()<=3].reset_index(drop=True)
clusterdata['z_score_deaths']=stats.zscore(clusterdata['death_rate_ld'])
clusterdata = clusterdata.loc[clusterdata['z_score_deaths'].abs()<=3].reset_index(drop=True)

X = clusterdata[['case_rate_ld','death_rate_ld']]
y = clusterdata['location']


from sklearn.cluster import AgglomerativeClustering

cluster = AgglomerativeClustering(n_clusters=9, affinity='euclidean', linkage='ward')
cluster.fit_predict(X)
plt.scatter(X.iloc[:,1],X.iloc[:,0], c=cluster.labels_, cmap='rainbow')

len(cluster.labels_)
cluster_labels = pd.concat([clusterdata,pd.DataFrame(cluster.labels_)],axis=1)
cluster_labels.rename(columns = {0:'cluster'},inplace=True)

#temp = cluster_labels[cluster_labels[0]==0]['location'].reset_index(drop=True)
#temp['cluster'] = '0' 

#cases_county = cases_county.drop(['cluster'],axis=1)
cases_county = pd.merge(cases_county,cluster_labels[['location','cluster']],on='location',how='left')

with open(os.path.join(data_dir, 'cases_county.pickle'), 'wb') as output:
    pickle.dump(cases_county, output)

with open(os.path.join(data_dir, 'cluster_labels.pickle'), 'wb') as output:
    pickle.dump(cluster_labels, output)


'''
from sklearn.cluster import KMeans
plt.scatter(X['death_rate_ld'],X['case_rate_ld'])

wcss = []
for i in range(1, 15):
    kmeans = KMeans(n_clusters=i, init='k-means++', max_iter=1000, n_init=10, random_state=0)
    kmeans.fit(X)
    wcss.append(kmeans.inertia_)

plt.plot(range(1, 15), wcss)
plt.title('Elbow Method')
plt.xlabel('Number of clusters')
plt.ylabel('WCSS')
plt.show()
'''
'''
kmeans = KMeans(n_clusters=7, init='k-means++', max_iter=1000, n_init=10, random_state=0)
pred_y = kmeans.fit_predict(X)

plt.scatter(clusterdata['death_rate_ld'],clusterdata['case_rate_ld'])
plt.scatter(kmeans.cluster_centers_[:, 1], kmeans.cluster_centers_[:, 0], s=300, c='red')
plt.show()

from collections import Counter, defaultdict
print(Counter(kmeans.labels_))

clusters_indices = defaultdict(list)
for index, c  in enumerate(kmeans.labels_):
    clusters_indices[c].append(index)
    
print(clusters_indices[0])    


cluster1 = clusterdata.iloc[clusters_indices[0],:].reset_index(drop=True)
'''

###########################RF FEATURE SELECTION#################################### 

from sklearn.ensemble import RandomForestRegressor

mob_features_case = pd.DataFrame(index=cases_county['location'].unique(),columns=mobility_vars_g)
mob_features_death= pd.DataFrame(index=cases_county['location'].unique(),columns=mobility_vars_g) 

for c in cases_county['location'].unique():
    ## CHANGE IN CASE RATE
    temp = cases_county.loc[(cases_county['location']==c),['case_rate_ld']+mobility_vars_g].dropna().reset_index(drop=True)
    X = temp.iloc[:,1:]
    y = temp.iloc[:,0]
    # Create a random forest classifier
    rff = RandomForestRegressor(n_estimators = 100,random_state = 0, n_jobs = -1)   
    # Train the classifier
    try:
        rff.fit(X, y)
        # get feature importance
        mob_features_case.loc[c,:] = rff.feature_importances_
    except:
        continue    
    ## CHANGE IN DEATH RATE
    temp = cases_county.loc[(cases_county['location']==c),['death_rate_ld']+mobility_vars_g].dropna().reset_index(drop=True)
    X = temp.iloc[:,1:]
    y = temp.iloc[:,0]
    # Create a random forest classifier
    rff = RandomForestRegressor(n_estimators = 100,random_state = 0, n_jobs = -1)   
    # Train the classifier
    try:
        rff.fit(X, y)
        # get feature importance
        mob_features_death.loc[c,:] = rff.feature_importances_
    except:
        continue    

with open(os.path.join(data_dir, 'mob_features_case.pickle'), 'wb') as output:
    pickle.dump(mob_features_case, output)
with open(os.path.join(data_dir, 'mob_features_death.pickle'), 'wb') as output:
    pickle.dump(mob_features_death, output)
    
#######################################

from sklearn.ensemble import RandomForestRegressor
mob_features_case_lag = pd.DataFrame(index=cases_county['location'].unique(),columns=mobility_vars_g)
mob_features_death_lag= pd.DataFrame(index=cases_county['location'].unique(),columns=mobility_vars_g) 

for c in cases_county['location'].unique():
    ## CHANGE IN CASE RATE
    temp = cases_county.loc[(cases_county['location']==c),['case_rate_ld']+mobility_vars_g].dropna().reset_index(drop=True)
    X = temp.iloc[:,1:].shift(7).iloc[7:,:].reset_index(drop=True)
    y = temp.iloc[:,0].iloc[7:].reset_index(drop=True)
    # Create a random forest classifier
    rff = RandomForestRegressor(n_estimators = 1000,random_state = 0, n_jobs = -1)   
    # Train the classifier
    try:
        rff.fit(X, y)
        # get feature importance
        mob_features_case_lag.loc[c,:] = rff.feature_importances_
    except:
        continue    
    ## CHANGE IN DEATH RATE
    temp = cases_county.loc[(cases_county['location']==c),['death_rate_ld']+mobility_vars_g].dropna().reset_index(drop=True)
    X = temp.iloc[:,1:].shift(7).iloc[7:,:].reset_index(drop=True)
    y = temp.iloc[:,0].iloc[7:].reset_index(drop=True)
    # Create a random forest classifier
    rff = RandomForestRegressor(n_estimators = 1000,random_state = 0, n_jobs = -1)   
    # Train the classifier
    try:
        rff.fit(X, y)
        # get feature importance
        mob_features_death_lag.loc[c,:] = rff.feature_importances_
    except:
        continue    

with open(os.path.join(data_dir, 'mob_features_case_lag.pickle'), 'wb') as output:
    pickle.dump(mob_features_case_lag, output)
with open(os.path.join(data_dir, 'mob_features_death_lag.pickle'), 'wb') as output:
    pickle.dump(mob_features_death_lag, output)    

#######################################

mob_features_case_lag_2 = pd.DataFrame(index=cases_county['location'].unique(),columns=mobility_vars_g)
mob_features_death_lag_2= pd.DataFrame(index=cases_county['location'].unique(),columns=mobility_vars_g) 

for c in cases_county['location'].unique():
    ## CHANGE IN CASE RATE
    temp = cases_county.loc[(cases_county['location']==c),['case_rate_ld']+mobility_vars_g].dropna().reset_index(drop=True)
    X = temp.iloc[:,1:].shift(14).iloc[14:,:].reset_index(drop=True)
    y = temp.iloc[:,0].iloc[14:].reset_index(drop=True)
    # Create a random forest classifier
    rff = RandomForestRegressor(n_estimators = 100,random_state = 0, n_jobs = -1)   
    # Train the classifier
    try:
        rff.fit(X, y)
        # get feature importance
        mob_features_case_lag_2.loc[c,:] = rff.feature_importances_
    except:
        continue    
    ## CHANGE IN DEATH RATE
    temp = cases_county.loc[(cases_county['location']==c),['death_rate_ld']+mobility_vars_g].dropna().reset_index(drop=True)
    X = temp.iloc[:,1:].shift(14).iloc[14:,:].reset_index(drop=True)
    y = temp.iloc[:,0].iloc[14:].reset_index(drop=True)
    # Create a random forest classifier
    rff = RandomForestRegressor(n_estimators = 100,random_state = 0, n_jobs = -1)   
    # Train the classifier
    try:
        rff.fit(X, y)
        # get feature importance
        mob_features_death_lag_2.loc[c,:] = rff.feature_importances_
    except:
        continue    

with open(os.path.join(data_dir, 'mob_features_case_lag_2.pickle'), 'wb') as output:
    pickle.dump(mob_features_case_lag_2, output)
with open(os.path.join(data_dir, 'mob_features_death_lag_2.pickle'), 'wb') as output:
    pickle.dump(mob_features_death_lag_2, output)    
    

########################### RIDGE #################################### 

from sklearn.model_selection import GridSearchCV
from sklearn.linear_model import Ridge

ridge = Ridge()
parameters = {'alpha' : [1e-15,1e-10,1e-8,1e-4,1e-3,1e-2,1,5,10,20]}

l_mob_features_case_lag = pd.DataFrame(index=cases_county['location'].unique(),columns=mobility_vars_g)
l_mob_features_death_lag= pd.DataFrame(index=cases_county['location'].unique(),columns=mobility_vars_g) 

for c in cases_county['location'].unique():
    ## CHANGE IN CASE RATE
    temp = cases_county.loc[(cases_county['location']==c),['case_rate_ld']+mobility_vars_g].dropna().reset_index(drop=True)
    X = temp.iloc[:,1:].shift(7).iloc[7:,:].reset_index(drop=True)
    y = temp.iloc[:,0].iloc[7:].reset_index(drop=True)
    # Create a ridge classifier
    ridge_reg = GridSearchCV(ridge,parameters,scoring='neg_mean_squared_error',cv=5)   
    # Train the classifier
    try:
        ridge_reg.fit(X, y)
        # get feature importance
        l_mob_features_case_lag.loc[c,:] = ridge_reg.best_estimator_.coef_
    except:
        continue    
    ## CHANGE IN DEATH RATE
    temp = cases_county.loc[(cases_county['location']==c),['death_rate_ld']+mobility_vars_g].dropna().reset_index(drop=True)
    X = temp.iloc[:,1:].shift(7).iloc[7:,:].reset_index(drop=True)
    y = temp.iloc[:,0].iloc[7:].reset_index(drop=True)

    ridge_reg = GridSearchCV(ridge,parameters,scoring='neg_mean_squared_error',cv=5)   
    # Train the classifier
    try:
        ridge_reg.fit(X, y)
        # get feature importance
        l_mob_features_death_lag.loc[c,:] = ridge_reg.best_estimator_.coef_
    except:
        continue    

with open(os.path.join(data_dir, 'l_mob_features_case_lag.pickle'), 'wb') as output:
    pickle.dump(l_mob_features_case_lag, output)
with open(os.path.join(data_dir, 'l_mob_features_death_lag.pickle'), 'wb') as output:
    pickle.dump(l_mob_features_death_lag, output)    


writer=pd.ExcelWriter(os.path.join(data_dir,'mob_features_ranking_ridge.xlsx'))
writer.save()
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir, 'mob_features_ranking_ridge.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir, 'mob_features_ranking_ridge.xlsx'),engine='openpyxl')
writer.book=book
l_mob_features_case_lag.to_excel(writer,'cases')
l_mob_features_death_lag.to_excel(writer,'deaths')
writer.save()
   