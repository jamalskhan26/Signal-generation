# -*- coding: utf-8 -*-
from IPython import get_ipython
get_ipython().magic('reset -sf')

#############################################################
# Load/import data
#############################################################
import pandas as pd
import numpy as np
import os
import pickle

data_dir = '//rschfiler2.baml.com/UK-NYCOMMODITIES/Jamal/CovidAnalysis/Mobility'

fname=os.path.join(data_dir, 'county_summary.pickle')
county_summary = pickle.load( open( fname, "rb" ) )

fname=os.path.join(data_dir, 'cluster_labels.pickle')
cluster_labels = pickle.load( open( fname, "rb" ) )

clusterdata = cluster_labels[['cluster','case_rate_ld','death_rate_ld']]

writer=pd.ExcelWriter(os.path.join(data_dir,'county_cluster_data.xlsx'))
writer.save()
for c in clusterdata['cluster'].unique():
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, 'county_cluster_data.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, 'county_cluster_data.xlsx'),engine='openpyxl')
    writer.book=book
    clusterdata[clusterdata['cluster']==c].reset_index(drop=True).to_excel(writer,str(c))
    writer.save()
    
fname=os.path.join(data_dir, 'cases_county.pickle')
cases_county = pickle.load( open( fname, "rb" ) )

mobility_vars = ['retail_recreation','grocery_pharmacy','parks','transit_stations','workplaces','residential','driving']

temp = (cases_county[['location']+mobility_vars].groupby('location').mean() - 100).reset_index()
cluster_mobility_avg = pd.merge(temp,cluster_labels[['location','cluster']],on='location',how='left').drop(['location'],axis=1)
cluster_mobility_avg = cluster_mobility_avg.groupby('cluster').mean()
cluster_mobility_avg = cluster_mobility_avg

temp = (cases_county[['location']+mobility_vars].groupby('location').quantile(0.9) - 100).reset_index()
cluster_mobility_90q = pd.merge(temp,cluster_labels[['location','cluster']],on='location',how='left').drop(['location'],axis=1)
cluster_mobility_90q = cluster_mobility_90q.groupby('cluster').mean()
cluster_mobility_90q = cluster_mobility_90q

temp = (cases_county[['location']+mobility_vars].groupby('location').quantile(0.5) - 100).reset_index()
cluster_mobility_50q = pd.merge(temp,cluster_labels[['location','cluster']],on='location',how='left').drop(['location'],axis=1)
cluster_mobility_50q = cluster_mobility_50q.groupby('cluster').mean()
cluster_mobility_50q = cluster_mobility_50q

writer=pd.ExcelWriter(os.path.join(data_dir,'cluster_mobility_avg.xlsx'))
writer.save()
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir, 'cluster_mobility_avg.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir, 'cluster_mobility_avg.xlsx'),engine='openpyxl')
writer.book=book
cluster_mobility_avg.to_excel(writer,'cluster_mobility_avg')
cluster_mobility_90q.to_excel(writer,'cluster_mobility_90q')
cluster_mobility_50q.to_excel(writer,'cluster_mobility_50q')
writer.save()
       
############################################################################################################

## map data
from bs4 import BeautifulSoup

fname=os.path.join(data_dir, 'driving_county_avg_1.pickle')
driving_county_avg_1 = pickle.load( open( fname, "rb" ) )
driving_county_dict_1 = {}
for c in driving_county_avg_1['FIPS'].unique():
    try:
        driving_county_dict_1[c] = float(driving_county_avg_1.loc[driving_county_avg_1['FIPS'] == c, 'driving'])
    except:
        continue

fname=os.path.join(data_dir, 'driving_county_avg_2.pickle')
driving_county_avg_2 = pickle.load( open( fname, "rb" ) )
driving_county_dict_2 = {}
for c in driving_county_avg_2['FIPS'].unique():
    try:
        driving_county_dict_2[c] = float(driving_county_avg_2.loc[driving_county_avg_2['FIPS'] == c, 'driving'])
    except:
        continue


############################################################################################################

# Load the SVG map
fname=os.path.join(data_dir, 'counties.svg')
svg = open(fname, 'r').read()

# Load into Beautiful Soup
soup = BeautifulSoup(svg, selfClosingTags=['defs','sodipodi:namedview'])

# Find counties
paths = soup.findAll('path')

# map colors
colors_cases =  ['#f7fbff', '#deebf7', '#c6dbef', '#9ecae1', '#6baed6', '#4292c6', '#2171b5', '#08519c', '#08306b']
                
# County style
path_style = 'font-size:12px;fill-rule:nonzero;stroke:#FFFFFF;stroke-opacity:1;stroke-width:0.1;stroke-miterlimit:4;stroke-dasharray:none;stroke-linecap:butt;marker-start:none;stroke-linejoin:bevel;fill:'

    
# Color the counties based on case rate 
for p in paths:
     
    if p['id'] not in ["State_Lines", "separator"]:
        # pass
        try:
            rate = driving_county_dict_1[p['id']]
        except:
            continue
             
        if rate > np.quantile(driving_county_avg_1['driving'].dropna(),0.95):
            color_class = 8
        elif rate > np.quantile(driving_county_avg_1['driving'].dropna(),0.90):
            color_class = 7
        elif rate > np.quantile(driving_county_avg_1['driving'].dropna(),0.75):
            color_class = 6
        elif rate > np.quantile(driving_county_avg_1['driving'].dropna(),0.6):
            color_class = 5
        elif rate > np.quantile(driving_county_avg_1['driving'].dropna(),0.5):
            color_class = 4
        elif rate > np.quantile(driving_county_avg_1['driving'].dropna(),0.25):
            color_class = 3
        elif rate > np.quantile(driving_county_avg_1['driving'].dropna(),0.1):
            color_class = 2
        elif rate > np.quantile(driving_county_avg_1['driving'].dropna(),0.05):
            color_class = 1
        else:
            color_class = 0
        
        color = colors_cases[color_class]
        p['style'] = path_style + color
    

# Output map
fname = os.path.join(data_dir, 'countymap_driving_1.svg')      
with open(fname, 'w') as f:
    print(soup.prettify(),file=f)
    
# Load the SVG map
fname=os.path.join(data_dir, 'counties.svg')
svg = open(fname, 'r').read()

# Load into Beautiful Soup
soup = BeautifulSoup(svg, selfClosingTags=['defs','sodipodi:namedview'])

# Find counties
paths = soup.findAll('path')

# map colors
colors_cases =  ['#f7fbff', '#deebf7', '#c6dbef', '#9ecae1', '#6baed6', '#4292c6', '#2171b5', '#08519c', '#08306b']
                
# County style
path_style = 'font-size:12px;fill-rule:nonzero;stroke:#FFFFFF;stroke-opacity:1;stroke-width:0.1;stroke-miterlimit:4;stroke-dasharray:none;stroke-linecap:butt;marker-start:none;stroke-linejoin:bevel;fill:'

    
# Color the counties based on case rate 
for p in paths:
     
    if p['id'] not in ["State_Lines", "separator"]:
        # pass
        try:
            rate = driving_county_dict_2[p['id']]
        except:
            continue
             
        if rate > np.quantile(driving_county_avg_2['driving'].dropna(),0.95):
            color_class = 8
        elif rate > np.quantile(driving_county_avg_2['driving'].dropna(),0.90):
            color_class = 7
        elif rate > np.quantile(driving_county_avg_2['driving'].dropna(),0.75):
            color_class = 6
        elif rate > np.quantile(driving_county_avg_2['driving'].dropna(),0.6):
            color_class = 5
        elif rate > np.quantile(driving_county_avg_2['driving'].dropna(),0.5):
            color_class = 4
        elif rate > np.quantile(driving_county_avg_2['driving'].dropna(),0.25):
            color_class = 3
        elif rate > np.quantile(driving_county_avg_2['driving'].dropna(),0.1):
            color_class = 2
        elif rate > np.quantile(driving_county_avg_2['driving'].dropna(),0.05):
            color_class = 1
        else:
            color_class = 0
        
        color = colors_cases[color_class]
        p['style'] = path_style + color
    

# Output map
fname = os.path.join(data_dir, 'countymap_driving_2.svg')      
with open(fname, 'w') as f:
    print(soup.prettify(),file=f)
    
    
### CHANGE STROKE IN LAST TWO LINES TO #FFFFFF    

###############################################################
        
import svgutils.compose as sc
from IPython.display import SVG # /!\ note the 'SVG' function also in svgutils.compose
import numpy as np

# print image then save as png -- word to excel and RSCH copy into word with title
fname=os.path.join(data_dir, 'countymap_driving_1.svg')
SVG(fname)

# print image then save as png -- word to excel and RSCH copy into word with title
fname=os.path.join(data_dir, 'countymap_driving_2.svg')
SVG(fname)

############################################################################################################

############################################################################################################

## map data

from bs4 import BeautifulSoup

map_dict_cases = {}
map_dict_deaths = {}
for c in county_summary['FIPS'].unique():
    try:
        map_dict_cases[c] = float(county_summary.loc[county_summary['FIPS'] == c, 'case_rate'])
    except:
        continue
    try:
        map_dict_deaths[c] = float(county_summary.loc[county_summary['FIPS'] == c, 'death_rate'])
    except:
        continue

############################################################################################################

# Load the SVG map
fname=os.path.join(data_dir, 'counties.svg')
svg = open(fname, 'r').read()

# Load into Beautiful Soup
soup = BeautifulSoup(svg, selfClosingTags=['defs','sodipodi:namedview'])

# Find counties
paths = soup.findAll('path')

# map colors
colors_cases =  ['#fcfbfd', '#efedf5', '#dadaeb', '#dadaeb', '#9e9ac8', '#807dba', '#6a51a3', '#54278f', '#3f007d']
                
# County style
path_style = 'font-size:12px;fill-rule:nonzero;stroke:#FFFFFF;stroke-opacity:1;stroke-width:0.1;stroke-miterlimit:4;stroke-dasharray:none;stroke-linecap:butt;marker-start:none;stroke-linejoin:bevel;fill:'

    
# Color the counties based on case rate 
for p in paths:
     
    if p['id'] not in ["State_Lines", "separator"]:
        # pass
        try:
            rate = map_dict_cases[p['id']]
        except:
            continue
             
        if rate > np.quantile(county_summary['case_rate'].dropna(),0.95):
            color_class = 8
        elif rate > np.quantile(county_summary['case_rate'].dropna(),0.90):
            color_class = 7
        elif rate > np.quantile(county_summary['case_rate'].dropna(),0.75):
            color_class = 6
        elif rate > np.quantile(county_summary['case_rate'].dropna(),0.6):
            color_class = 5
        elif rate > np.quantile(county_summary['case_rate'].dropna(),0.4):
            color_class = 4
        elif rate > np.quantile(county_summary['case_rate'].dropna(),0.25):
            color_class = 3
        elif rate > np.quantile(county_summary['case_rate'].dropna(),0.1):
            color_class = 2
        elif rate > np.quantile(county_summary['case_rate'].dropna(),0.01):
            color_class = 1
        else:
            color_class = 0
        
        color = colors_cases[color_class]
        p['style'] = path_style + color
    

# Output map
fname = os.path.join(data_dir, 'countymap_cases.svg')      
with open(fname, 'w') as f:
    print(soup.prettify(),file=f)
    
    
### CHANGE STROKE IN LAST TWO LINES TO #FFFFFF    

###############################################################
        
import svgutils.compose as sc
from IPython.display import SVG # /!\ note the 'SVG' function also in svgutils.compose
import numpy as np

# print image then save as png -- word to excel and RSCH copy into word with title
fname=os.path.join(data_dir, 'countymap_cases.svg')
SVG(fname)

############################################################################################################

# Load the SVG map
fname=os.path.join(data_dir, 'counties.svg')
svg = open(fname, 'r').read()

# Load into Beautiful Soup
soup = BeautifulSoup(svg, selfClosingTags=['defs','sodipodi:namedview'])

# Find counties
paths = soup.findAll('path')

# map colors
colors_deaths = ['#fff5f0', '#fee0d2', '#fcbba1', '#fc9272', '#fb6a4a', '#ef3b2c', '#cb181d', '#a50f15', '#67000d']
                
# County style
path_style = 'font-size:12px;fill-rule:nonzero;stroke:#FFFFFF;stroke-opacity:1;stroke-width:0.1;stroke-miterlimit:4;stroke-dasharray:none;stroke-linecap:butt;marker-start:none;stroke-linejoin:bevel;fill:'

    
# Color the counties based on case rate 
for p in paths:
     
    if p['id'] not in ["State_Lines", "separator"]:
        # pass
        try:
            rate = map_dict_deaths[p['id']]
        except:
            continue
             
        if rate > np.quantile(county_summary['death_rate'].dropna(),0.99):
            color_class = 8
        elif rate > np.quantile(county_summary['death_rate'].dropna(),0.95):
            color_class = 7
        elif rate > np.quantile(county_summary['death_rate'].dropna(),0.9):
            color_class = 6
        elif rate > np.quantile(county_summary['death_rate'].dropna(),0.8):
            color_class = 5
        elif rate > np.quantile(county_summary['death_rate'].dropna(),0.7):
            color_class = 4
        elif rate > np.quantile(county_summary['death_rate'].dropna(),0.6):
            color_class = 3
        elif rate > np.quantile(county_summary['death_rate'].dropna(),0.5):
            color_class = 2
        elif rate > np.quantile(county_summary['death_rate'].dropna(),0.4):
            color_class = 1
        else:
            color_class = 0
        
        color = colors_deaths[color_class]
        p['style'] = path_style + color
    

# Output map
fname = os.path.join(data_dir, 'countymap_deaths.svg')      
with open(fname, 'w') as f:
    print(soup.prettify(),file=f)
    
    
### CHANGE STROKE IN LAST TWO LINES TO #FFFFFF    

###############################################################

# print image then save as png -- word to excel and RSCH copy into word with title
fname=os.path.join(data_dir, 'countymap_deaths.svg')
SVG(fname)


###############################################################

fname=os.path.join(data_dir, 'cluster_labels.pickle')
cluster_labels = pickle.load( open( fname, "rb" ) )

cluster_labels = pd.merge(cluster_labels, county_summary[['location','FIPS']],on='location',how='left')

cluster_labels_dict = {}
for c in cluster_labels['FIPS'].unique():
    try:
        cluster_labels_dict[c] = int(cluster_labels.loc[cluster_labels['FIPS'] == c, 'cluster'])
    except:
        continue


# Load the SVG map
fname=os.path.join(data_dir, 'counties.svg')
svg = open(fname, 'r').read()

# Load into Beautiful Soup
soup = BeautifulSoup(svg, selfClosingTags=['defs','sodipodi:namedview'])

# Find counties
paths = soup.findAll('path')

# map colors
colors_clusters = ['#a6cee3', '#1f78b4', '#b2df8a', '#33a02c', '#fb9a99', '#e31a1c', '#fdbf6f', '#ff7f00', '#cab2d6']
                
# County style
path_style = 'font-size:12px;fill-rule:nonzero;stroke:#FFFFFF;stroke-opacity:1;stroke-width:0.1;stroke-miterlimit:4;stroke-dasharray:none;stroke-linecap:butt;marker-start:none;stroke-linejoin:bevel;fill:'

    
# Color the counties based on case rate 
for p in paths:
     
    if p['id'] not in ["State_Lines", "separator"]:
        # pass
        try:
            rate = cluster_labels_dict[p['id']]
        except:
            continue
             
        color_class = rate
        
        color = colors_clusters[color_class]
        p['style'] = path_style + color
    

# Output map
fname = os.path.join(data_dir, 'countymap_clusters.svg')      
with open(fname, 'w') as f:
    print(soup.prettify(),file=f)
        
### CHANGE STROKE IN LAST TWO LINES TO #FFFFFF    

###############################################################

# print image then save as png -- word to excel and RSCH copy into word with title
fname=os.path.join(data_dir, 'countymap_clusters.svg')
SVG(fname)

###############################################################

fname=os.path.join(data_dir, 'mob_features_case_lag.pickle')
mob_features_case = pickle.load( open( fname, "rb" ) )
mob_features_case = mob_features_case.fillna(0)

top_features_case = mob_features_case.eq(mob_features_case.max(1), axis=0).dot(mob_features_case.columns).reset_index()
top_features_case.columns = ['location','type'] 
top_features_case['topmob'] = int(0)

top_features_case = pd.merge(top_features_case, county_summary[['location','FIPS']],on='location',how='left')

top_features_case.loc[top_features_case['type']=='retail_recreation','topmob'] = 1
top_features_case.loc[top_features_case['type']=='grocery_pharmacy','topmob'] = 2
top_features_case.loc[top_features_case['type']=='parks','topmob'] = 3
top_features_case.loc[top_features_case['type']=='transit_stations','topmob'] = 4
top_features_case.loc[top_features_case['type']=='workplaces','topmob'] = 5
top_features_case.loc[top_features_case['type']=='residential','topmob'] = 6

top_features_case_dict  = {}
for c in top_features_case['FIPS'].unique():
    try:
        top_features_case_dict[c] = top_features_case.loc[top_features_case['FIPS'] == c, 'topmob'].values[0]
    except:
        continue

# Load the SVG map
fname=os.path.join(data_dir, 'counties.svg')
svg = open(fname, 'r').read()

# Load into Beautiful Soup
soup = BeautifulSoup(svg, selfClosingTags=['defs','sodipodi:namedview'])

# Find counties
paths = soup.findAll('path')

# map colors
colors_topmob = ['#d0d0d0', '#66c2a5', '#fc8d62', '#8da0cb', '#e78ac3', '#a6d854', '#ffd92f']
                
# County style
path_style = 'font-size:12px;fill-rule:nonzero;stroke:#FFFFFF;stroke-opacity:1;stroke-width:0.1;stroke-miterlimit:4;stroke-dasharray:none;stroke-linecap:butt;marker-start:none;stroke-linejoin:bevel;fill:'

    
# Color the counties based on case rate 
for p in paths:
     
    if p['id'] not in ["State_Lines", "separator"]:
        # pass
        try:
            rate = int(top_features_case_dict[p['id']])
        except:
            continue
                 
        color_class = rate
        
        color = colors_topmob[color_class]
        p['style'] = path_style + color
    

# Output map
fname = os.path.join(data_dir, 'countymap_mob_cases.svg')      
with open(fname, 'w') as f:
    print(soup.prettify(),file=f)
        
### CHANGE STROKE IN LAST TWO LINES TO #FFFFFF    

###############################################################

# print image then save as png -- word to excel and RSCH copy into word with title
fname=os.path.join(data_dir, 'countymap_mob_cases.svg')
SVG(fname)

###############################################################

fname=os.path.join(data_dir, 'mob_features_death_lag.pickle')
mob_features_death = pickle.load( open( fname, "rb" ) )
mob_features_death = mob_features_death.fillna(0)

top_features_death = mob_features_death.eq(mob_features_death.max(1), axis=0).dot(mob_features_death.columns).reset_index()
top_features_death.columns = ['location','type'] 
top_features_death['topmob'] = int(0)

top_features_death = pd.merge(top_features_death, county_summary[['location','FIPS']],on='location',how='left')

top_features_death.loc[top_features_death['type']=='retail_recreation','topmob'] = 1
top_features_death.loc[top_features_death['type']=='grocery_pharmacy','topmob'] = 2
top_features_death.loc[top_features_death['type']=='parks','topmob'] = 3
top_features_death.loc[top_features_death['type']=='transit_stations','topmob'] = 4
top_features_death.loc[top_features_death['type']=='workplaces','topmob'] = 5
top_features_death.loc[top_features_death['type']=='residential','topmob'] = 6

top_features_death_dict  = {}
for c in top_features_death['FIPS'].unique():
    try:
        top_features_death_dict[c] = top_features_death.loc[top_features_death['FIPS'] == c, 'topmob'].values[0]
    except:
        continue

# Load the SVG map
fname=os.path.join(data_dir, 'counties.svg')
svg = open(fname, 'r').read()

# Load into Beautiful Soup
soup = BeautifulSoup(svg, selfClosingTags=['defs','sodipodi:namedview'])

# Find counties
paths = soup.findAll('path')

# map colors
colors_topmob = ['#d0d0d0', '#66c2a5', '#fc8d62', '#8da0cb', '#e78ac3', '#a6d854', '#ffd92f']
                
# County style
path_style = 'font-size:12px;fill-rule:nonzero;stroke:#FFFFFF;stroke-opacity:1;stroke-width:0.1;stroke-miterlimit:4;stroke-dasharray:none;stroke-linecap:butt;marker-start:none;stroke-linejoin:bevel;fill:'

    
# Color the counties based on case rate 
for p in paths:
     
    if p['id'] not in ["State_Lines", "separator"]:
        # pass
        try:
            rate = int(top_features_death_dict[p['id']])
        except:
            continue
                 
        color_class = rate
        
        color = colors_topmob[color_class]
        p['style'] = path_style + color
    

# Output map
fname = os.path.join(data_dir, 'countymap_mob_deaths.svg')      
with open(fname, 'w') as f:
    print(soup.prettify(),file=f)
        
### CHANGE STROKE IN LAST TWO LINES TO #FFFFFF    

###############################################################

# print image then save as png -- word to excel and RSCH copy into word with title
fname=os.path.join(data_dir, 'countymap_mob_deaths.svg')
SVG(fname)

###############################################################

fname=os.path.join(data_dir, 'mob_features_case_lag.pickle')
mob_features_case = pickle.load( open( fname, "rb" ) )
mob_features_case = mob_features_case.fillna(0)
mob_features_case = mob_features_case.reset_index() 
mob_features_case = mob_features_case.rename(columns = {'index':'location'})

fname=os.path.join(data_dir, 'cluster_labels.pickle')
cluster_labels = pickle.load( open( fname, "rb" ) )

mob_features_ranking_case = pd.merge(mob_features_case,cluster_labels[['location','cluster']],on='location',how='left')

writer=pd.ExcelWriter(os.path.join(data_dir,'mob_features_ranking_case.xlsx'))
writer.save()
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir, 'mob_features_ranking_case.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir, 'mob_features_ranking_case.xlsx'),engine='openpyxl')
writer.book=book
mob_features_ranking_case.to_excel(writer,'mob_features_ranking_case')
writer.save()



fname=os.path.join(data_dir, 'mob_features_death_lag.pickle')
mob_features_death = pickle.load( open( fname, "rb" ) )
mob_features_death = mob_features_death.fillna(0)
mob_features_death = mob_features_death.reset_index() 
mob_features_death = mob_features_death.rename(columns = {'index':'location'})

fname=os.path.join(data_dir, 'cluster_labels.pickle')
cluster_labels = pickle.load( open( fname, "rb" ) )

mob_features_ranking_death = pd.merge(mob_features_death,cluster_labels[['location','cluster']],on='location',how='left')
mob_features_ranking_death = mob_features_ranking_death.dropna().reset_index(drop=True)
temp = mob_features_ranking_death.iloc[:,1:-1]
temp['sum'] = temp.iloc[:,:6].sum(1) 
temp['mean']= temp.iloc[:,:6].mean(1) 
temp['mean'].mean()

temp[temp['sum'] != 0]

writer=pd.ExcelWriter(os.path.join(data_dir,'mob_features_ranking_death.xlsx'))
writer.save()
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir, 'mob_features_ranking_death.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir, 'mob_features_ranking_death.xlsx'),engine='openpyxl')
writer.book=book
mob_features_ranking_death.to_excel(writer,'mob_features_ranking_death')
writer.save()
