# -*- coding: utf-8 -*-
"""
Created on Mon Sep 30 13:09:08 2019

@author: ZK463GK
"""

from IPython import get_ipython
get_ipython().magic('reset -sf')
import pandas as pd
import numpy as np
import os
import pickle
import statsmodels.api as sm
import datetime as dt
from datetime import timedelta as td
import matplotlib.backends.backend_pdf
import matplotlib.pyplot as plt
from dateutil.relativedelta import relativedelta

###import helper functions
os.chdir('//corp.bankofamerica.com/london/Researchshared/Share4/Sector/Commodities/Publications/FICC Portfolio Monthly/2019_enhanced hedging/FX risk premia strategies/Code/FX_RPproject')
import HelperFunctions as fns

########UPDATE PATH############
data_dir = '//rschfiler2.baml.com/UK-NYCOMMODITIES/Jamal/Datasets'
out_dir = '//corp.bankofamerica.com/london/Researchshared/Share4/Sector/Commodities/Publications/FICC Portfolio Monthly/2019_enhanced hedging/FX risk premia strategies/Data'
###############################


#############################################################
#1. Load Data
#############################################################

    ### Optimal hedge ratios (rolling regression params)
    
    # instrument: USTW AFE (daily and monthly)
fname=os.path.join(out_dir, 'ustw_hedge_b.pickle')
ustw_hedge_b = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'ustw_hedge_b_m.pickle')
ustw_hedge_b_m = pickle.load( open( fname, "rb" ) )

    # instrument: USTW BROAD
fname=os.path.join(out_dir, 'ustwbroa_hedge_b.pickle')
ustwbroa_hedge_b = pickle.load( open( fname, "rb" ) )

    # instrument: DXY
fname=os.path.join(out_dir, 'dxy_hedge_b.pickle')
dxy_hedge_b = pickle.load( open( fname, "rb" ) )

    # instrument: FXRP IVol basket (daily and monthly) 
fname=os.path.join(out_dir, 'fxrp_ivol_hedge_b.pickle')
fxrp_ivol_hedge_b = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'fxrp_ivol_hedge_b_m.pickle')
fxrp_ivol_hedge_b_m = pickle.load( open( fname, "rb" ) )

    # instrument: FXRP Reg basket
fname=os.path.join(out_dir, 'fxrp_reg_hedge_b.pickle')
fxrp_reg_hedge_b = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'fxrp_reg_hedge_b_m.pickle')
fxrp_reg_hedge_b_m = pickle.load( open( fname, "rb" ) )

    # instrument: FXRP Ridge Reg basket
fname=os.path.join(out_dir, 'fxrp_reg_rid_hedge_b.pickle')
fxrp_reg_rid_hedge_b = pickle.load( open( fname, "rb" ) )

    ### Prices and Indices

    ### import asset prices
fname=os.path.join(data_dir, 'equity_fx_indices.pickle')
assets = pickle.load( open( fname, "rb" ) )

    ### import USTW AFE index
fname=os.path.join(data_dir, 'ustw_afe_index.pickle')
ustw_afe = pickle.load( open( fname, "rb" ) )

    ### import USTW BROAD index
fname=os.path.join(data_dir, 'ustw_index.pickle')
ustw_broa = pickle.load( open( fname, "rb" ) )

    ### import DXY index
fname=os.path.join(data_dir, 'dxy.pickle')
dxy = pickle.load( open( fname, "rb" ) )

    ### import FXRP IVol basket
fname=os.path.join(out_dir, 'fxrp_basket_ivol.pickle')
fxrp_basket_ivol = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'fxrp_basket_ivol_m.pickle')
fxrp_basket_ivol_m = pickle.load( open( fname, "rb" ) )

    ### import FXRP OLS Reg basket
fname=os.path.join(out_dir, 'fxrp_basket_reg.pickle')
fxrp_basket_reg = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'fxrp_basket_reg_m.pickle')
fxrp_basket_reg_m = pickle.load( open( fname, "rb" ) )

    ### import FXRP Ridge Reg basket
fname=os.path.join(out_dir, 'fxrp_basket_reg_rid.pickle')
fxrp_basket_reg_rid = pickle.load( open( fname, "rb" ) )

   ###Import FX RP indices (only used for correlations)
fname=os.path.join(out_dir, 'fx_factors.pickle')
fx_factors = pickle.load( open( fname, "rb" ) )

#Momentum
fname=os.path.join(data_dir, 'mom_dm.pickle')
mom_dm= pickle.load( open( fname, "rb" ) )
fname=os.path.join(data_dir, 'mom_em.pickle')
mom_em= pickle.load( open( fname, "rb" ) )

#Mean Reversion
fname=os.path.join(data_dir, 'meanrev_dm.pickle')
meanrev_dm= pickle.load( open( fname, "rb" ) )
fname=os.path.join(data_dir, 'meanrev_em.pickle')
meanrev_em= pickle.load( open( fname, "rb" ) )


##UPDATE: 
    ### import FXRP OLS Reg asset-specific basket
fname=os.path.join(out_dir, 'fxrp_basket_reg_m_fac.pickle')
fxrp_basket_reg_m_fac = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'fxrp_reg_hedge_b_m_fac.pickle')
fxrp_reg_hedge_b_m_fac = pickle.load( open( fname, "rb" ) )

#############################################################
#2. Create Date Series for Signal and Trade days
#############################################################

    ### 1-month holding period for strategy
        # generate signal on last trading day of month
        # trade/rebalance on first trading day of month

dateseries=pd.DataFrame(assets.Date)
dateseries['month']=dateseries['Date'].dt.month
    #signal on last day of month
dateseries['temp']=(dateseries.month)-(dateseries.month).shift(-1)
dateseries['signalDay']=0
dateseries.loc[dateseries['temp']!=0,'signalDay']=1
    #trade on first day of month
dateseries['temp']=(dateseries.month)-(dateseries.month).shift(1)
dateseries['tradeDay']=0
dateseries.loc[dateseries['temp']!=0,'tradeDay']=1

dateseries=dateseries[['Date','signalDay','tradeDay']]

#############################################################
#3. Generate Monthly Returns Series for realized returns
#############################################################

def monthly_ret(dat):
    dat=pd.merge(dateseries,dat,on='Date',how='left')
        #forward fill missing asset prices for monthly return calculation
    dat=dat.fillna(method='ffill')
        #subset to monthly observations
    dat_m=dat[dat['tradeDay']==1].drop(['signalDay','tradeDay'],axis=1)
        #calculate monthly returns for subsequent month (t+1)
    dat_m=fns.ret_next(dat_m).reset_index(drop=True)
    
    return dat_m

    # asset returns
assets_m = monthly_ret(assets)
    # ustw_afe returns
ustw_afe_m = monthly_ret(ustw_afe)
    # ustw_broad returns
ustw_broa_m = monthly_ret(ustw_broa)
    # DXY returns
dxy_m = monthly_ret(dxy)
    # fxrp ivol returns
fxrp_basket_ivol_m = monthly_ret(fxrp_basket_ivol_m)
    # fxrp reg returns
fxrp_basket_reg_m = monthly_ret(fxrp_basket_reg_m )
    # fxrp ridge reg returns
fxrp_basket_reg_rid_m = monthly_ret(fxrp_basket_reg_rid)
    # fxrp factor returns
fx_factors_m= monthly_ret(fx_factors)


def weekly_ret(dat):
    dat=pd.merge(dateseries,dat,on='Date',how='left')
        #forward fill missing asset prices for monthly return calculation
    dat=dat.fillna(method='ffill')
    dat['week']=dat['Date'].apply(lambda x: x.isocalendar()[1])
    dat['begin_week']=dat['week']-dat['week'].shift(1)    
        #subset to weekly observations
    dat_w=dat[dat['begin_week']!=0].drop(['signalDay','tradeDay','begin_week','week'],axis=1)
        #calculate weekly returns for subsequent week
    dat_w=fns.ret_next(dat_w).reset_index(drop=True)
    return dat_w

    # ustw_afe returns
ustw_afe_w = weekly_ret(ustw_afe)
    # ustw_broad returns
ustw_broa_w = weekly_ret(ustw_broa)
    # fxrp factor returns
fx_factors_w= weekly_ret(fx_factors)

#############################################################
#. Factor USTWER Correlations
#############################################################
    ### monthly returns
ustw_m=pd.merge(ustw_afe_m,ustw_broa_m,how='left')
ustw_factors_m=pd.merge(ustw_m,fx_factors_m,how='left')
corr_fx_all=ustw_factors_m.corr()

'''
from openpyxl import load_workbook
book= load_workbook(os.path.join(out_dir,'factor_correl.xlsx'))
writer=pd.ExcelWriter(os.path.join(out_dir,'factor_correl.xlsx'),engine='openpyxl')
writer.book=book
corr_fx_all.to_excel(writer,'corr_monthly')
writer.save()
'''

    ### weekly returns
ustw_w=pd.merge(ustw_afe_w,ustw_broa_w,how='left')
ustw_factors_w=pd.merge(ustw_w,fx_factors_w,how='left')
corr_fx_all=ustw_factors_w.corr()


'''
from openpyxl import load_workbook
book= load_workbook(os.path.join(out_dir,'factor_correl.xlsx'))
writer=pd.ExcelWriter(os.path.join(out_dir,'factor_correl.xlsx'),engine='openpyxl')
writer.book=book
corr_fx_all.to_excel(writer,'corr_weekly')
writer.save()
'''
    ### daily returns
ustw=pd.merge(ustw_afe,ustw_broa,how='left')
ustw_factors=pd.merge(fns.ret(ustw),fns.ret(fx_factors),how='left')
corr_fx_all=ustw_factors.corr()

'''
from openpyxl import load_workbook
book= load_workbook(os.path.join(out_dir,'factor_correl.xlsx'))
writer=pd.ExcelWriter(os.path.join(out_dir,'factor_correl.xlsx'),engine='openpyxl')
writer.book=book
corr_fx_all.to_excel(writer,'corr_daily')
writer.save()
'''

#############################################################
#############################################################
# ANALYSIS USING DAILY HEDGE RATIOS
#############################################################
#############################################################

#############################################################
#4. Calculate strategy returns
#############################################################

##########################################
###calculate asset returns in DXY terms###
###(what a foriegn investor buying #######
### with DXY currency basket would make)##
##########################################

#perfect hedge i.e. returns in USD
assets_m_ph=assets_m.copy()
temp=pd.DataFrame((1+assets_m.iloc[:,1:].values)*(1+dxy_m.iloc[:,1:].values)-1,columns=assets_m.columns[1:],index=assets_m.index)
assets_m.iloc[:,1:]=temp


def strategy_returns(assetreturns,hedge,hedge_wt,name):
    '''
    input assets returns (flexible to frequency of returns),
            hedging instrument returns (e.g. USTW$ or FXRP Basket),
            optimal hedge ratio dataframe (based on rolling regressions),
            & name of the hedging instrument variable in str format (e.g. 'USTW$ Index')
    output dictionary containing (asset: strategy returns) pairs 
            strategy returns computed as (asset return - 100.beta*hedge_instrument return)        
    '''
    strat_ret={}
    for p in assetreturns.columns.drop(['Date']):
        ### first merge rolling hedge ratios with date series
        signal=pd.merge(dateseries,hedge_wt[p],on='Date',how='left')
            ### next forward fill signals to ensure non-missing values on signal days
        signal=signal.fillna(method='ffill')
        signal=signal.dropna(subset=hedge_wt[p].columns.drop(['Date','const']),how='all')
            ### subset to signal days and create trade date series (t+1)
        signal=signal[(signal['signalDay']==1) | (signal['tradeDay']==1)]
        signal['tradeDate']=signal['Date'].shift(-1)
        signal=signal[(signal['signalDay']==1)]
        signal=signal.drop(['signalDay','tradeDay','const'],axis=1)
        signal.rename(columns = {name:'hedge_ratio'}, 
                      inplace=True)
            ### merge with hedging instrument returns on trade date (t+1)
        hedging_instrument=pd.merge(signal,hedge,left_on='tradeDate',right_on='Date',how='left',suffixes=('_x',''))
            ### invest $(beta*100) in hedging instrument on signal date with 1M holding period
        hedging_instrument['hedge_ins_ret']=(hedging_instrument['hedge_ratio'])*hedging_instrument[name]
        hedging_instrument=hedging_instrument[['Date','hedge_ins_ret','hedge_ratio']]
        
        strat_ret[p]=pd.merge(assetreturns[['Date',p]],hedging_instrument,how='right',on='Date')
        strat_ret[p]['strat_ret']=strat_ret[p][p]-strat_ret[p]['hedge_ins_ret']
    
    return strat_ret


    ### Compute strategy returns ###

    # fully hedged, hedging instrument: USTW$
    # create toy dictionary with 100% hedge ratios to pass to function
def f(dat,name):
    temp = dat.copy()
    temp[name]=1
    return temp


#############################################################
#############################################################
# ANALYSIS USING MONTHLY HEDGE RATIOS
#############################################################
#############################################################

#############################################################
#4. Calculate strategy returns
#############################################################

    # Fully hedged with USTW$
full_hedge_ustw_m={p:f(v,'USTW$ Index') for p,v in ustw_hedge_b_m.items()}
hedge_full_m=strategy_returns(assetreturns=assets_m,
                                hedge=ustw_afe_m,
                                hedge_wt=full_hedge_ustw_m,
                                name='USTW$ Index')

    # hedging instrument: USTW$
hedge_ustw_afe_m=strategy_returns(assetreturns=assets_m,
                                hedge=ustw_afe_m,
                                hedge_wt=ustw_hedge_b_m,
                                name='USTW$ Index')

    # hedging instrument: FXRP IVol BASKET
hedge_fxrp_ivol_m=strategy_returns(assetreturns=assets_m,
                                hedge=fxrp_basket_ivol_m,
                                hedge_wt=fxrp_ivol_hedge_b_m,
                                name='fxrp_basket_ivol')

    # hedging instrument: FXRP OLS Reg BASKET
hedge_fxrp_reg_m=strategy_returns(assetreturns=assets_m,
                                hedge=fxrp_basket_reg_m,
                                hedge_wt=fxrp_reg_hedge_b_m,
                                name='fxrp_basket_reg')


    # perfect hedge
hedge_perf_m=strategy_returns(assetreturns=assets_m_ph,
                                hedge=ustw_afe_m,
                                hedge_wt=full_hedge_ustw_m,
                                name='USTW$ Index')

#UPDATE: 
    # hedging instrument: FXRP OLS Reg BASKET Asset Specific
dummy=fxrp_basket_reg_m.copy()    
dummy['fxrp_basket_reg']=1
    
hedge_fxrp_reg_m_fac=strategy_returns(assetreturns=assets_m,
                                hedge=dummy,
                                hedge_wt=fxrp_basket_reg_m_fac,
                                name='fxrp_basket_reg')

#############################################################
#4. Compare strategy performance MONTHLY
#############################################################

compare_strategies_m={}
for p in hedge_ustw_afe_m.keys():
    
    ### merge all hedging strategies (FXRP, USTW AFE) -- UPDATE: EXCLUDE FIRST n OBS (n=60 rolling window for optimal hedge ratio regressions)
    compare_strategies_m[p] = pd.merge(hedge_full_m[p],hedge_ustw_afe_m[p],on=['Date'],how='inner',suffixes=('','_tw_afe')).iloc[60:,]
    compare_strategies_m[p] = pd.merge(compare_strategies_m[p],hedge_fxrp_ivol_m[p],on=['Date'],how='inner',suffixes=('','_fxrp_ivol'))
    compare_strategies_m[p] = pd.merge(compare_strategies_m[p],hedge_fxrp_reg_m[p],on=['Date'],how='inner',suffixes=('','_fxrp_reg'))
    compare_strategies_m[p] = pd.merge(compare_strategies_m[p],hedge_fxrp_reg_m_fac[p],on=['Date'],how='inner',suffixes=('','_fxrp_reg_fac'))
    compare_strategies_m[p] = pd.merge(compare_strategies_m[p],hedge_perf_m[p],on=['Date'],how='inner',suffixes=('','_perfect'))
    
    ### compute strategy returns
    
    # Unhedged
    compare_strategies_m[p]['_']=(pd.concat([pd.Series([100]),1+compare_strategies_m[p][p]])).cumprod().iloc[1:]

    # Full
    compare_strategies_m[p]['_full']=(pd.concat([pd.Series([100]),1+compare_strategies_m[p]['strat_ret']])).cumprod().iloc[1:]

    # USTW AFE
    compare_strategies_m[p]['_twd_afe']=(pd.concat([pd.Series([100]),1+compare_strategies_m[p]['strat_ret_tw_afe']])).cumprod().iloc[1:]

    # FXRP IVol
    compare_strategies_m[p]['_fxrp_ivol']=(pd.concat([pd.Series([100]),1+compare_strategies_m[p]['strat_ret_fxrp_ivol']])).cumprod().iloc[1:]

    # FXRP Reg
    compare_strategies_m[p]['_fxrp_reg']=(pd.concat([pd.Series([100]),1+compare_strategies_m[p]['strat_ret_fxrp_reg']])).cumprod().iloc[1:]

    # FXRP Reg Asset specific   
    compare_strategies_m[p]['_fxrp_reg_fac']=(pd.concat([pd.Series([100]),1+compare_strategies_m[p]['strat_ret_fxrp_reg_fac']])).cumprod().iloc[1:]

    # Perfect Hedge
    compare_strategies_m[p]['_perfect']=(pd.concat([pd.Series([100]),1+compare_strategies_m[p][p+'_perfect']])).cumprod().iloc[1:]

###TRACKING ERRORS
    compare_strategies_m[p]['strat_ret_tw_afe_te'] = (compare_strategies_m[p][p+'_perfect']-compare_strategies_m[p]['strat_ret_tw_afe'])    
    compare_strategies_m[p]['strat_ret_fxrp_ivol_te'] = (compare_strategies_m[p][p+'_perfect']-compare_strategies_m[p]['strat_ret_fxrp_ivol'])    
    compare_strategies_m[p]['strat_ret_fxrp_reg_te'] = (compare_strategies_m[p][p+'_perfect']-compare_strategies_m[p]['strat_ret_fxrp_reg'])
    compare_strategies_m[p]['strat_ret_fxrp_reg_fac_te'] = (compare_strategies_m[p][p+'_perfect']-compare_strategies_m[p]['strat_ret_fxrp_reg_fac'])
    
#############################################################
# Create barcharts for returns, vol, info ratios
#############################################################

from dateutil.relativedelta import relativedelta

    #storage dictionary containing dictionary for each asset with a dataframe for each period
compare_sstats_m={}
for p in hedge_ustw_afe_m.keys():
    #strategies to compare
    strats=[p,'strat_ret','strat_ret_tw_afe','strat_ret_fxrp_ivol','strat_ret_fxrp_reg','strat_ret_fxrp_reg_fac',p+'_perfect']
    #compute strategy return summary stats for last n years
    def last_n(n):
        _lastn=compare_strategies_m[p].loc[compare_strategies_m[p]['Date']>(dt.datetime.today()-relativedelta(years=n)),:][strats]
        mean_lastn = _lastn.mean()*12*100
        vol_lastn  = _lastn.std()*pow(12,0.5)*100
        info_lastn = mean_lastn / vol_lastn  
        lastn=pd.concat([mean_lastn,vol_lastn,info_lastn],axis=1)
        lastn.columns=[s + str(n) for s in ['mean_last','vol_last','info_last']]
        return lastn
    
    last_10=last_n(10)
    last_5 =last_n(5)
    last_3 =last_n(3)
    last_1 =last_n(1)

    def pre_n(n):
        _lastn=compare_strategies_m[p].loc[compare_strategies_m[p]['Date']<(dt.datetime.today()-relativedelta(years=n)),:][strats]
        mean_lastn = _lastn.mean()*12*100
        vol_lastn  = _lastn.std()*pow(12,0.5)*100
        info_lastn = mean_lastn / vol_lastn  
        lastn=pd.concat([mean_lastn,vol_lastn,info_lastn],axis=1)
        lastn.columns=[s + str(n) for s in ['mean_pre','vol_pre','info_pre']]
        return lastn

    pre_10=pre_n(10)


    hedge=['hedge_ratio_tw_afe','hedge_ratio_fxrp_ivol','hedge_ratio_fxrp_reg']
    #compute strategy return summary stats for last n years
    def hedge_n(n):
        _hedgen=compare_strategies_m[p].loc[compare_strategies_m[p]['Date']>(dt.datetime.today()-relativedelta(years=n)),:][hedge]
        hedge_lastn = pd.DataFrame(-1*_hedgen.median())
        hedge_lastn.columns=[s + str(n) for s in ['hedge_last']]
        return hedge_lastn 
    
    hedge_10=hedge_n(10)
    hedge_5 =hedge_n(5)
    hedge_3 =hedge_n(3)
    hedge_1 =hedge_n(1)
    
    #compute tracking errors for last n years
    te=['strat_ret_tw_afe_te','strat_ret_fxrp_ivol_te','strat_ret_fxrp_reg_te','strat_ret_fxrp_reg_fac_te']
    def te_n(n):
        _ten=compare_strategies_m[p].loc[compare_strategies_m[p]['Date']>(dt.datetime.today()-relativedelta(years=n)),:][te]
        te_lastn = pd.DataFrame(_ten.std())*pow(12,0.5)*100
        te_lastn.columns=[s + str(n) for s in ['te_last']]
        return te_lastn 
    
    te_10=te_n(10)
    te_5 =te_n(5)
    te_3 =te_n(3)
    te_1 =te_n(1)

    compare_sstats_m[p]={
            "mean":pd.concat([pre_10['mean_pre10'],last_10['mean_last10'],last_5['mean_last5'],last_3['mean_last3'],last_1['mean_last1']],axis=1),
            "vol":pd.concat([pre_10['vol_pre10'],last_10['vol_last10'],last_5['vol_last5'],last_3['vol_last3'],last_1['vol_last1']],axis=1),
            "info":pd.concat([pre_10['info_pre10'],last_10['info_last10'],last_5['info_last5'],last_3['info_last3'],last_1['info_last1']],axis=1),
            "hedge":pd.concat([hedge_10,hedge_5,hedge_3,hedge_1],axis=1),
            "te":pd.concat([te_10,te_5,te_3,te_1],axis=1)            
        }


with open(os.path.join(out_dir, 'compare_strategies_m.pickle'), 'wb') as output:
    pickle.dump(compare_strategies_m, output)
with open(os.path.join(out_dir, 'compare_sstats_m.pickle'), 'wb') as output:
    pickle.dump(compare_sstats_m, output)


#############################################################
#4. Compare strategy performance DAILY
#############################################################

#############################################################
#Calculate strategy returns
#############################################################
full_hedge_ustw={p:f(v,'USTW$ Index') for p,v in ustw_hedge_b.items()}

hedge_full=strategy_returns(assetreturns=assets_m,
                                hedge=ustw_afe_m,
                                hedge_wt=full_hedge_ustw,
                                name='USTW$ Index')

    # hedging instrument: USTW$
hedge_ustw_afe=strategy_returns(assetreturns=assets_m,
                                hedge=ustw_afe_m,
                                hedge_wt=ustw_hedge_b,
                                name='USTW$ Index')

    # hedging instrument: USTW Broad
hedge_ustw_broa=strategy_returns(assetreturns=assets_m,
                                hedge=ustw_broa_m,
                                hedge_wt=ustwbroa_hedge_b,
                                name='USTW Broad Index')

    # hedging instrument: DXY
hedge_dxy=strategy_returns(assetreturns=assets_m,
                                hedge=dxy_m,
                                hedge_wt=dxy_hedge_b,
                                name='DXY Curncy')

    # hedging instrument: FXRP IVol BASKET
hedge_fxrp_ivol=strategy_returns(assetreturns=assets_m,
                                hedge=fxrp_basket_ivol_m,
                                hedge_wt=fxrp_ivol_hedge_b,
                                name='fxrp_basket_ivol')

    # hedging instrument: FXRP OLS Reg BASKET
hedge_fxrp_reg=strategy_returns(assetreturns=assets_m,
                                hedge=fxrp_basket_reg_m,
                                hedge_wt=fxrp_reg_hedge_b,
                                name='fxrp_basket_reg')

    # hedging instrument: FXRP Ridge Reg BASKET
hedge_fxrp_reg_rid=strategy_returns(assetreturns=assets_m,
                                hedge=fxrp_basket_reg_rid_m,
                                hedge_wt=fxrp_reg_rid_hedge_b,
                                name='fxrp_basket_reg_rid')

    # perfect hedge
hedge_perf=strategy_returns(assetreturns=assets_m_ph,
                                hedge=ustw_afe_m,
                                hedge_wt=full_hedge_ustw,
                                name='USTW$ Index')

#############################################################
#4. Compare strategy performance
#       Merge USTW & FXRP hedging strategies and plot returns
#############################################################

compare_strategies={}
for p in hedge_ustw_afe.keys():
    
    ### merge all hedging strategies (Full, FXRP, USTW AFE, USTW BROAD) -- UPDATE: EXCLUDE FIRST n OBS (n=rolling window for optimal hedge ratio regressions)
    compare_strategies[p] = pd.merge(hedge_full[p],hedge_ustw_afe[p],on=['Date'],how='inner',suffixes=('','_tw_afe'))
    compare_strategies[p] = pd.merge(compare_strategies[p],hedge_ustw_broa[p],on=['Date'],how='inner',suffixes=('','_tw_broa'))
    compare_strategies[p] = pd.merge(compare_strategies[p],hedge_dxy[p],on=['Date'],how='inner',suffixes=('','_dxy'))
    compare_strategies[p] = pd.merge(compare_strategies[p],hedge_fxrp_ivol[p],on=['Date'],how='inner',suffixes=('','_fxrp_ivol'))
    compare_strategies[p] = pd.merge(compare_strategies[p],hedge_fxrp_reg[p],on=['Date'],how='inner',suffixes=('','_fxrp_reg'))
    compare_strategies[p] = pd.merge(compare_strategies[p],hedge_fxrp_reg_rid[p],on=['Date'],how='inner',suffixes=('','_fxrp_reg_rid'))
    compare_strategies[p] = pd.merge(compare_strategies[p],hedge_perf[p],on=['Date'],how='inner',suffixes=('','_perfect'))
    
    ### compute strategy returns
    
    # Unhedged
    compare_strategies[p]['_']=(pd.concat([pd.Series([100]),1+compare_strategies[p][p]])).cumprod().iloc[1:]

    # Full
    compare_strategies[p]['_full']=(pd.concat([pd.Series([100]),1+compare_strategies[p]['strat_ret']])).cumprod().iloc[1:]

    # USTW AFE
    compare_strategies[p]['_twd_afe']=(pd.concat([pd.Series([100]),1+compare_strategies[p]['strat_ret_tw_afe']])).cumprod().iloc[1:]

    # USTW BROAD
    compare_strategies[p]['_twd_broa']=(pd.concat([pd.Series([100]),1+compare_strategies[p]['strat_ret_tw_broa']])).cumprod().iloc[1:]

    # DXY
    compare_strategies[p]['_dxy']=(pd.concat([pd.Series([100]),1+compare_strategies[p]['strat_ret_dxy']])).cumprod().iloc[1:]

    # FXRP IVol
    compare_strategies[p]['_fxrp_ivol']=(pd.concat([pd.Series([100]),1+compare_strategies[p]['strat_ret_fxrp_ivol']])).cumprod().iloc[1:]

    # FXRP Reg
    compare_strategies[p]['_fxrp_reg']=(pd.concat([pd.Series([100]),1+compare_strategies[p]['strat_ret_fxrp_reg']])).cumprod().iloc[1:]

    # FXRP Ridge Reg
    compare_strategies[p]['_fxrp_reg_rid']=(pd.concat([pd.Series([100]),1+compare_strategies[p]['strat_ret_fxrp_reg_rid']])).cumprod().iloc[1:]

    # Perfect Hedge
    compare_strategies[p]['_perfect']=(pd.concat([pd.Series([100]),1+compare_strategies[p][p+'_perfect']])).cumprod().iloc[1:]


###TRACKING ERRORS
    compare_strategies[p]['strat_ret_tw_afe_te'] = (compare_strategies[p][p+'_perfect']-compare_strategies[p]['strat_ret_tw_afe'])    
    compare_strategies[p]['strat_ret_fxrp_ivol_te'] = (compare_strategies[p][p+'_perfect']-compare_strategies[p]['strat_ret_fxrp_ivol'])    
    compare_strategies[p]['strat_ret_fxrp_reg_te'] = (compare_strategies[p][p+'_perfect']-compare_strategies[p]['strat_ret_fxrp_reg'])
    compare_strategies[p]['strat_ret_fxrp_reg_rid_te'] = (compare_strategies[p][p+'_perfect']-compare_strategies[p]['strat_ret_fxrp_reg_rid'])

#### plot returns of unhedged strategy and hedged strategies using FXRP basket and USTW
    ### update: USTWBROA and USTWAFE yield same results so use USTW AFE

#############################################################
# Create barcharts for returns, vol, info ratios
#############################################################


p='S&P-500'

    #storage dictionary containing dictionary for each asset with a dataframe for each period
compare_sstats={}
for p in hedge_ustw_afe.keys():
    #strategies to compare
    strats=[p,'strat_ret','strat_ret_tw_afe','strat_ret_fxrp_ivol','strat_ret_fxrp_reg','strat_ret_fxrp_reg_rid',p+'_perfect']
    #compute strategy return summary stats for last n years
    def last_n(n):
        _lastn=compare_strategies[p].loc[compare_strategies[p]['Date']>(dt.datetime.today()-relativedelta(years=n)),:][strats]
        mean_lastn = _lastn.mean()*12*100
        vol_lastn  = _lastn.std()*pow(12,0.5)*100
        info_lastn = mean_lastn / vol_lastn  
        lastn=pd.concat([mean_lastn,vol_lastn,info_lastn],axis=1)
        lastn.columns=[s + str(n) for s in ['mean_last','vol_last','info_last']]
        return lastn
    
    last_10=last_n(10)
    last_5 =last_n(5)
    last_3 =last_n(3)
    last_1 =last_n(1)

    #compute hedge ratio median for last n years
    hedge=['hedge_ratio_tw_afe','hedge_ratio_fxrp_ivol','hedge_ratio_fxrp_reg','hedge_ratio_fxrp_reg_rid']
    def hedge_n(n):
        _hedgen=compare_strategies[p].loc[compare_strategies[p]['Date']>(dt.datetime.today()-relativedelta(years=n)),:][hedge]
        hedge_lastn = pd.DataFrame(-1*_hedgen.median())
        hedge_lastn.columns=[s + str(n) for s in ['hedge_last']]
        return hedge_lastn 
    
    hedge_10=hedge_n(10)
    hedge_5 =hedge_n(5)
    hedge_3 =hedge_n(3)
    hedge_1 =hedge_n(1)
    
    #compute tracking errors for last n years
    te=['strat_ret_tw_afe_te','strat_ret_fxrp_ivol_te','strat_ret_fxrp_reg_te','strat_ret_fxrp_reg_rid']
    def te_n(n):
        _ten=compare_strategies[p].loc[compare_strategies[p]['Date']>(dt.datetime.today()-relativedelta(years=n)),:][te]
        te_lastn = pd.DataFrame(_ten.std())*pow(12,0.5)*100
        te_lastn.columns=[s + str(n) for s in ['te_last']]
        return te_lastn 
    
    te_10=te_n(10)
    te_5 =te_n(5)
    te_3 =te_n(3)
    te_1 =te_n(1)

    compare_sstats[p]={
            "mean":pd.concat([last_10['mean_last10'],last_5['mean_last5'],last_3['mean_last3'],last_1['mean_last1']],axis=1),
            "vol":pd.concat([last_10['vol_last10'],last_5['vol_last5'],last_3['vol_last3'],last_1['vol_last1']],axis=1),
            "info":pd.concat([last_10['info_last10'],last_5['info_last5'],last_3['info_last3'],last_1['info_last1']],axis=1),
            "hedge":pd.concat([hedge_10,hedge_5,hedge_3,hedge_1],axis=1),
            "te":pd.concat([te_10,te_5,te_3,te_1],axis=1)            
        }
    

with open(os.path.join(out_dir, 'compare_strategies.pickle'), 'wb') as output:
    pickle.dump(compare_strategies, output)
with open(os.path.join(out_dir, 'compare_sstats.pickle'), 'wb') as output:
    pickle.dump(compare_sstats, output)


#############################################################
                # DRAWDOWN ANALYSIS
#############################################################

def drawdown(data):
    ret=data.copy()
    col=list(ret.columns.drop(['Date']))
# Use a trailing window equal to entire history
    window = len(ret)
# Calculate the max drawdown in the past window days for each day in the series.
# Use min_periods=1 to let the first window data to have an expanding window
    Roll_Max = ret[col].apply(lambda x: x.rolling(window, min_periods=1).max().reset_index(drop=True),axis=0)
#get index of rolling max to keep track of drawdown months
    temp=Roll_Max.drop_duplicates(keep='first')
    temp['maxindex']=temp.index
    Roll_Max = pd.merge(Roll_Max,temp,how='outer')
    _Drawdown = (ret[col]/Roll_Max.drop(['maxindex'],axis=1) - 1.0)
#Calculate a rolling series of MDD and the number of months over which it occurred
#    Roll_Min = _Drawdown.apply(lambda x: x.rolling(window, min_periods=1).min().reset_index(drop=True),axis=0)
#   _Drawdown=pd.concat([_Drawdown,Roll_Min[col]],axis=1,ignore_index=True)
#    _Drawdown.columns=['dd','mdd']
    MDD=pd.concat([_Drawdown,Roll_Max[['maxindex']]],axis=1,ignore_index=False)
    MDD['months']=MDD.index-MDD['maxindex']
#Number of days since last MDD
#    MDD=MDD.drop(['maxindex'],axis=1)
#    MDD=MDD.drop_duplicates(subset=Roll_Min.columns,keep='first')
    MDD['index']=MDD.index
    MDD=MDD.loc[MDD.groupby('maxindex')[col].idxmin().values.flatten()]

    MDD=MDD.sort_values(by=col).reset_index(drop=True)
    MDD=MDD[(MDD[col]!=0).values]
#    MDD['time_since_lastmdd']=np.append([window],MDD['index'].iloc[:-1].values-MDD['index'].iloc[1:].values)
#    MDD['unique']=0
#    MDD['unique']=(MDD['months']<MDD['time_since_lastmdd'])
#    MDD=MDD[MDD['unique']==True].reset_index(drop=True)[col+['months','index']].iloc[0:10,:]
    MDD=MDD[col+['months','index']].iloc[0:10,:]

    temp=ret[['Date']]
    temp['index']=temp.index
    MDD=pd.merge(temp,MDD,on='index',how='inner')

    MDD['mdd']=MDD[col]
    MDD['id']=col[0]
    MDD=MDD.sort_values(by='mdd').reset_index(drop=True)
    MDD['rank']=MDD.index+1
    MDD['unique']=(MDD['rank'].astype(str)+MDD['id']).apply(lambda x: x.strip())
    MDD=MDD[['unique','rank','id','mdd','Date','months']]
    return MDD

MDD_dict={}
for p in compare_strategies_m.keys():
    MDD=pd.DataFrame()
    #strategies to compare
    strats=['_','_twd_afe','_fxrp_ivol','_fxrp_reg','_fxrp_reg_fac','_perfect']
    for s in strats:
        temp=drawdown(compare_strategies_m[p][['Date',s]])
        MDD=pd.concat([MDD,temp],axis=0,ignore_index=False)
    MDD_dict[p]=MDD



writer=pd.ExcelWriter(os.path.join(out_dir,'drawdowns.xlsx'))
writer.save()
for p in MDD_dict.keys():
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(out_dir,'drawdowns.xlsx'))
    writer=pd.ExcelWriter(os.path.join(out_dir,'drawdowns.xlsx'),engine='openpyxl')
    writer.book=book
    MDD_dict[p].to_excel(writer,(p),index=False)
    writer.save()

#############################################################
                # average drawdown vs max drawdown
                # stdev of drawdowns
                # drawdowns over non-overlapping windows
                # max drawdown duration
                # calmar ratios by year / periods
#############################################################

def drawdown_by_year(data):
    ret=data.copy().dropna()
    ret['year'] = ret['Date'].dt.year
    temp = ret.drop(['Date'], axis = 1).groupby('year')
    px = 1 - ret[[s]] / (temp.rolling(window = 12, min_periods=1).max().reset_index(drop=True)[[s]])
    px['year'] = ret['year']
    px = px.groupby('year')

    ret = fns.ret(ret.drop(['year'], axis = 1))
    ret['year'] = ret['Date'].dt.year
    ret = ret.drop(['Date'], axis = 1).dropna().groupby('year')
    
    _dd  = px.max() 
    _ret = (ret.mean() * 12)
    _std = (ret.std() * (12**0.5))
    _ir  = _ret/_std
    _cr  = _ret/_dd

    stats = pd.concat([_ret, _std, _dd, _ir, _cr], axis = 1)
    stats.columns = ['ret', 'vol', 'drawdown', 'info', 'calmar']
    stats['basket'] = s
    stats.index = [i+j for i,j in zip(map(str,list(stats.index)),list(stats['basket']))]
    return stats

dd_yearly_dict={}
for p in compare_strategies_m.keys():
    dd=pd.DataFrame()
    #strategies to compare
    strats=['_','_twd_afe','_fxrp_ivol','_fxrp_reg','_fxrp_reg_fac','_perfect']
    for s in strats:
        temp=drawdown_by_year(compare_strategies_m[p][['Date',s]])
        dd=pd.concat([dd,temp],axis=0,ignore_index=False)
    dd_yearly_dict[p]=dd

writer=pd.ExcelWriter(os.path.join(out_dir,'drawdowns_yearly.xlsx'))
writer.save()
for p in dd_yearly_dict.keys():
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(out_dir,'drawdowns_yearly.xlsx'))
    writer=pd.ExcelWriter(os.path.join(out_dir,'drawdowns_yearly.xlsx'),engine='openpyxl')
    writer.book=book
    dd_yearly_dict[p].to_excel(writer,(p),index=True)
    writer.save()

#############################################################
                # Summary of Results -- for EXCEL
#############################################################

optimal_hr_10=pd.DataFrame()
optimal_hr_5=pd.DataFrame()
optimal_hr_3=pd.DataFrame()

ret_pre=pd.DataFrame()
ret_10=pd.DataFrame()
ret_5=pd.DataFrame()
ret_3=pd.DataFrame()

vol_pre=pd.DataFrame()
vol_10=pd.DataFrame()
vol_5=pd.DataFrame()
vol_3=pd.DataFrame()

info_pre=pd.DataFrame()
info_10=pd.DataFrame()
info_5=pd.DataFrame()
info_3=pd.DataFrame()

te_10=pd.DataFrame()
te_5=pd.DataFrame()
te_3=pd.DataFrame()

for p in compare_sstats_m.keys():
    optimal_hr_10[p]=compare_sstats_m[p]['hedge'].hedge_last10.reset_index(drop=True)
    optimal_hr_5[p]=compare_sstats_m[p]['hedge'].hedge_last5.reset_index(drop=True)
    optimal_hr_3[p]=compare_sstats_m[p]['hedge'].hedge_last3.reset_index(drop=True)
    ret_pre[p]=compare_sstats_m[p]['mean'].mean_pre10.reset_index(drop=True)
    ret_10[p]=compare_sstats_m[p]['mean'].mean_last10.reset_index(drop=True)
    ret_5[p]=compare_sstats_m[p]['mean'].mean_last5.reset_index(drop=True)
    ret_3[p]=compare_sstats_m[p]['mean'].mean_last3.reset_index(drop=True)
    vol_pre[p]=compare_sstats_m[p]['vol'].vol_pre10.reset_index(drop=True)
    vol_10[p]=compare_sstats_m[p]['vol'].vol_last10.reset_index(drop=True)
    vol_5[p]=compare_sstats_m[p]['vol'].vol_last5.reset_index(drop=True)
    vol_3[p]=compare_sstats_m[p]['vol'].vol_last3.reset_index(drop=True)
    info_pre[p]=compare_sstats_m[p]['info'].info_pre10.reset_index(drop=True)
    info_10[p]=compare_sstats_m[p]['info'].info_last10.reset_index(drop=True)
    info_5[p]=compare_sstats_m[p]['info'].info_last5.reset_index(drop=True)
    info_3[p]=compare_sstats_m[p]['info'].info_last3.reset_index(drop=True)
    te_10[p]=compare_sstats_m[p]['te'].te_last10.reset_index(drop=True)
    te_5[p] =compare_sstats_m[p]['te'].te_last5.reset_index(drop=True)
    te_3[p] =compare_sstats_m[p]['te'].te_last3.reset_index(drop=True)
    

optimal_hr_10.index=['USTW$','IVol','Optimal']
optimal_hr_5.index=['USTW$','IVol','Optimal']
optimal_hr_3.index=['USTW$','IVol','Optimal']
ret_pre.index=['Unhedged','Full','USTW$','IVol','Opt','Opt-Asset','USD']
ret_10.index=['Unhedged','Full','USTW$','IVol','Opt','Opt-Asset','USD']
ret_5.index=['Unhedged','Full','USTW$','IVol','Opt','Opt-Asset','USD']
ret_3.index=['Unhedged','Full','USTW$','IVol','Opt','Opt-Asset','USD']
vol_pre.index=['Unhedged','Full','USTW$','IVol','Opt','Opt-Asset','USD']
vol_10.index=['Unhedged','Full','USTW$','IVol','Opt','Opt-Asset','USD']
vol_5.index=['Unhedged','Full','USTW$','IVol','Opt','Opt-Asset','USD']
vol_3.index=['Unhedged','Full','USTW$','IVol','Opt','Opt-Asset','USD']
info_pre.index=['Unhedged','Full','USTW$','IVol','Opt','Opt-Asset','USD']
info_10.index=['Unhedged','Full','USTW$','IVol','Opt','Opt-Asset','USD']
info_5.index=['Unhedged','Full','USTW$','IVol','Opt','Opt-Asset','USD']
info_3.index=['Unhedged','Full','USTW$','IVol','Opt','Opt-Asset','USD']
te_10.index=['USTW$','IVol','Opt','Opt-Asset']
te_5.index=['USTW$','IVol','Opt','Opt-Asset']
te_3.index=['USTW$','IVol','Opt','Opt-Asset']

def get_df_name(df):
    name =[x for x in globals() if globals()[x] is df][0]
    return name

list_dfs=[optimal_hr_10,optimal_hr_5,optimal_hr_3,ret_pre,ret_10,ret_5,ret_3,vol_pre,vol_10,vol_5,vol_3,info_pre,info_10,info_5,info_3,te_10,te_5,te_3]
df_names=list(map(get_df_name,list_dfs))
list_dfs=list(map(lambda x: x.T,list_dfs))

def save_xls(list_dfs, xls_path):
    with pd.ExcelWriter(xls_path) as writer:
        for n, df in enumerate(list_dfs):
            df.to_excel(writer,df_names[n])
        writer.save()
        
save_xls(list_dfs, os.path.join(out_dir,'assetComparison.xlsx'))


#rolling betas
writer=pd.ExcelWriter(os.path.join(out_dir,'rollingB_m.xlsx'))
writer.save()
for p in compare_sstats_m.keys():
    temp=pd.merge(fxrp_ivol_hedge_b_m[p].drop(['const'],axis=1),fxrp_reg_hedge_b_m[p].drop(['const'],axis=1),on='Date',how='inner')
    temp=pd.merge(temp,ustw_hedge_b_m[p].drop(['const'],axis=1),on='Date',how='inner').iloc[60:,:]
    temp.iloc[:,1:]=-1*temp.iloc[:,1:]
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(out_dir,'rollingB_m.xlsx'))
    writer=pd.ExcelWriter(os.path.join(out_dir,'rollingB_m.xlsx'),engine='openpyxl')
    writer.book=book
    temp.to_excel(writer,(p))
    writer.save()

    #asset-specific hedge ratios
def save_xls_dict(dict_dfs, xls_path):
    with pd.ExcelWriter(xls_path) as writer:
        for name, df in dict_dfs.items():
            temp=df.iloc[60:,:]
            temp.iloc[:,1:]=-1*temp.iloc[:,1:]
            temp.to_excel(writer,name)
        writer.save()

save_xls_dict(fxrp_reg_hedge_b_m_fac, os.path.join(out_dir,'rollingB_m_faclevel.xlsx'))

#rolling betas box plots
writer=pd.ExcelWriter(os.path.join(out_dir,'box_rollingB_m.xlsx'))
writer.save()
for p in assets.keys()[1:]:
    temp=pd.merge(fxrp_ivol_hedge_b_m[p].drop(['const'],axis=1),fxrp_reg_hedge_b_m[p].drop(['const'],axis=1),on='Date',how='inner')
    temp=pd.merge(temp,ustw_hedge_b_m[p].drop(['const'],axis=1),on='Date',how='inner').iloc[60:,:]
    temp.iloc[:,1:]=-1*temp.iloc[:,1:]
    temp=temp.drop(['Date'],axis=1).reset_index(drop=True)
    temp=pd.concat([temp.min(),temp.std(),temp.mean(),temp.std(),temp.max()],axis=1)
    temp.columns=['min','bottom','mean','top','max']
    temp.index=['FXRP iVol','FXRP$','USTW$']
#    temp=temp.T
#    temp2=temp.diff(axis=0)
#    temp2.iloc[0,:]=temp.iloc[0,:]
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(out_dir,'box_rollingB_m.xlsx'))
    writer=pd.ExcelWriter(os.path.join(out_dir,'box_rollingB_m.xlsx'),engine='openpyxl')
    writer.book=book
    temp.to_excel(writer,(p))
    writer.save()


## info ratios
#rolling betas
writer=pd.ExcelWriter(os.path.join(out_dir,'inforatios.xlsx'))
writer.save()
for p in compare_sstats_m.keys():
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(out_dir,'inforatios.xlsx'))
    writer=pd.ExcelWriter(os.path.join(out_dir,'inforatios.xlsx'),engine='openpyxl')
    writer.book=book
    compare_sstats_m[p]['info'].to_excel(writer,(p))
    writer.save()

#strategy return correlations
writer=pd.ExcelWriter(os.path.join(out_dir,'strat_cor.xlsx'))
writer.save()
writer=pd.ExcelWriter(os.path.join(out_dir,'strat_cor_roll.xlsx'))
writer.save()
for p in compare_strategies_m.keys():
    temp=pd.merge(dxy_m,compare_strategies_m[p],on='Date',how='inner')
    temp['year']=temp['Date'].dt.year
    temp.dropna(inplace=True)

    cor_all=temp[['DXY Curncy',p,'strat_ret','strat_ret_tw_afe','strat_ret_fxrp_ivol','strat_ret_fxrp_reg','strat_ret_fxrp_reg_fac',p+'_perfect']].corr()    
    cor_all['strat']=['DXY','Unhedged','Full','USTW$','IVol','FXRP$','FXRP Asset','USD']
    cor_all['period']='Full Period'
    cor_all=cor_all.reset_index(drop=True)
    
    cor_year=temp.groupby(by=['year'])[['DXY Curncy',p,'strat_ret','strat_ret_tw_afe','strat_ret_fxrp_ivol','strat_ret_fxrp_reg','strat_ret_fxrp_reg_fac',p+'_perfect']].corr()
    cor_year['strat']=cor_year.index.get_level_values(1)
    cor_year['period']=cor_year.index.get_level_values(0)
    cor_year=cor_year.reset_index(drop=True)

    cor=pd.concat([cor_all,cor_year],axis=0)

###3 year rolling correlations
    rolling_corr=pd.concat([temp['Date'],temp['DXY Curncy'].rolling(36).corr(temp[p])],axis=1)
    rolling_corr=pd.concat([rolling_corr, temp['DXY Curncy'].rolling(36).corr(temp['strat_ret_fxrp_reg'])],axis=1)
    rolling_corr=pd.concat([rolling_corr, temp['DXY Curncy'].rolling(36).corr(temp['strat_ret_fxrp_reg_fac'])],axis=1)
    rolling_corr.columns=['Date','$-Unhedged','$-FXRP$','$-FXRP Asset']

    from openpyxl import load_workbook
    book= load_workbook(os.path.join(out_dir,'strat_cor.xlsx'))
    writer=pd.ExcelWriter(os.path.join(out_dir,'strat_cor.xlsx'),engine='openpyxl')
    writer.book=book
    cor.to_excel(writer,(p))
    writer.save()
    
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(out_dir,'strat_cor_roll.xlsx'))
    writer=pd.ExcelWriter(os.path.join(out_dir,'strat_cor_roll.xlsx'),engine='openpyxl')
    writer.book=book
    rolling_corr.to_excel(writer,(p))
    writer.save()
     
temp['strat_ret_tw_afe'].corr(temp[p])
temp['strat_ret_tw_afe'].corr(temp['strat_ret_fxrp_reg'])
temp['strat_ret_fxrp_reg'].corr(temp[p])
    
plt.plot(fxrp_basket_reg_m['Date'],fxrp_basket_reg_m['fxrp_basket_reg'])
plt.plot(ustw_afe_m['Date'],ustw_afe_m['USTW$ Index'])
fxrp_basket_reg_m['fxrp_basket_reg'].corr(ustw_afe_m['USTW$ Index'])
fxrp_basket_reg_m['fxrp_basket_reg'].mean()*12    
ustw_afe_m['USTW$ Index'].mean()*12    
  
## hedging instruments summary stats 
hedge_inst=pd.merge(dxy_m,pd.merge(pd.merge(pd.merge(ustw_afe_m,ustw_broa_m,on='Date'),fxrp_basket_ivol_m,on='Date'),fxrp_basket_reg_m,on='Date'),on='Date')
hedge_inst=hedge_inst.replace([np.inf, -np.inf], np.nan)
hedge_inst.dropna(inplace=True)
    
writer=pd.ExcelWriter(os.path.join(out_dir,'hedge_corr.xlsx'))
hedge_inst.corr().to_excel(writer,'monthly')
writer.save()

def last_n(n):
    _lastn=hedge_inst.loc[hedge_inst['Date']>(dt.datetime.today()-relativedelta(years=n)),:]
    mean_lastn = _lastn.mean()*12*100
    vol_lastn  = _lastn.std()*pow(12,0.5)*100
    info_lastn = mean_lastn / vol_lastn  
    lastn=pd.concat([mean_lastn,vol_lastn,info_lastn],axis=1)
    lastn.columns=[s + str(n) for s in ['mean_last','vol_last','info_last']]
    return lastn
    
h_full = last_n(30)
h_last_10=last_n(10)
h_last_5 =last_n(5)
h_last_3 =last_n(3)

hedge_inst_mean=pd.concat([h_full.iloc[:,0],h_last_10.iloc[:,0],h_last_5.iloc[:,0],h_last_3.iloc[:,0]],axis=1)
hedge_inst_vol =pd.concat([h_full.iloc[:,1],h_last_10.iloc[:,1],h_last_5.iloc[:,1],h_last_3.iloc[:,1]],axis=1)
hedge_inst_info=pd.concat([h_full.iloc[:,2],h_last_10.iloc[:,2],h_last_5.iloc[:,2],h_last_3.iloc[:,2]],axis=1)

writer=pd.ExcelWriter(os.path.join(out_dir,'hedge_instr_sstats.xlsx'))
hedge_inst_mean.to_excel(writer,'mean')
hedge_inst_vol.to_excel(writer,'vol')
hedge_inst_info.to_excel(writer,'info')
writer.save()


########FACTORS##########
fx_factors=fx_factors.drop(['Value_Global','Carry_Global','DollarCarry_Global'],axis=1)
fx_factors=pd.merge(pd.merge(pd.merge(pd.merge(fx_factors,mom_dm,on='Date',how='outer'),
                        mom_em,on='Date',how='outer'),
                        meanrev_dm,on='Date',how='outer'),
                        meanrev_em,on='Date',how='outer')

fx_factors_ret=monthly_ret(fx_factors).dropna(thresh=10)    
ustw_factors_m=pd.merge(ustw_m,fx_factors_ret,how='left')
corr_fx_m=ustw_factors_m.corr()

def strat(dat): 
    out=(pd.concat([pd.Series([100]),1+dat])).cumprod().iloc[1:]
    return out

backtest=pd.concat([fx_factors_ret[['Date']],fx_factors_ret.iloc[:,1:].apply(strat)],axis=1).reset_index(drop=True)

compare_factors={}
#compute strategy return summary stats for last n years
def last_n(n):
    _lastn=fx_factors_ret.loc[fx_factors_ret['Date']>(dt.datetime.today()-relativedelta(years=n)),:]
    mean_lastn = _lastn.mean()*12*100
    vol_lastn  = _lastn.std()*pow(12,0.5)*100
    info_lastn = mean_lastn / vol_lastn  
    lastn=pd.concat([mean_lastn,vol_lastn,info_lastn],axis=1)
    lastn.columns=[s + str(n) for s in ['mean_last','vol_last','info_last']]
    return lastn
    
last_20=last_n(20)
last_10=last_n(10)
last_5 =last_n(5)
last_3 =last_n(3)
last_1 =last_n(1)


compare_factors={
"mean":pd.concat([last_20['mean_last20'],last_10['mean_last10'],last_5['mean_last5'],last_3['mean_last3'],last_1['mean_last1']],axis=1),
"vol":pd.concat([last_20['vol_last20'],last_10['vol_last10'],last_5['vol_last5'],last_3['vol_last3'],last_1['vol_last1']],axis=1),
"info":pd.concat([last_20['info_last20'],last_10['info_last10'],last_5['info_last5'],last_3['info_last3'],last_1['info_last1']],axis=1),
"corr":corr_fx_m
}

writer=pd.ExcelWriter(os.path.join(out_dir,'factor_summary.xlsx'))
writer.save()
for p in compare_factors.keys():
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(out_dir,'factor_summary.xlsx'))
    writer=pd.ExcelWriter(os.path.join(out_dir,'factor_summary.xlsx'),engine='openpyxl')
    writer.book=book
    compare_factors[p].to_excel(writer,(p))
    writer.save()
from openpyxl import load_workbook
book= load_workbook(os.path.join(out_dir,'factor_summary.xlsx'))
writer=pd.ExcelWriter(os.path.join(out_dir,'factor_summary.xlsx'),engine='openpyxl')
writer.book=book
backtest.to_excel(writer,"backtest")
writer.save() 