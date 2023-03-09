# -*- coding: utf-8 -*-
"""
Created on Thu Nov  7 10:44:12 2019

@author: ZK463GK
"""


'''
To Update Optimal Hedge Ratio calculation, run script till part #4 &
add function call for additional hedging instruments at the end
'''


from IPython import get_ipython
get_ipython().magic('reset -sf')
import pandas as pd
import numpy as np
import os
import pickle
import statsmodels.api as sm
import datetime as dt
###import helper functions
os.chdir('//rschfiler2.baml.com/UK-NYCOMMODITIES/Jamal/FX risk premia strategies/Code/FX_RPproject')
import HelperFunctions as fns

########UPDATE PATH############
data_dir = '//rschfiler2.baml.com/UK-NYCOMMODITIES/Jamal/Datasets'
out_dir = '//corp.bankofamerica.com/london/Researchshared/Share4/Sector/Commodities/Publications/FICC Portfolio Monthly/2019_enhanced hedging/FX risk premia strategies/Data'
###############################


#############################################################
#. Load Data
#############################################################

    ###Import Equity and FI indices
fname=os.path.join(data_dir, 'equity_fx_indices.pickle')
assets = pickle.load( open( fname, "rb" ) )

    ###Import FX RP indices
fname=os.path.join(out_dir, 'fx_factors.pickle')
fx_factors = pickle.load( open( fname, "rb" ) )

    ###Import US trade weighted reconstructed indices
fname=os.path.join(data_dir, 'ustw_afe_index.pickle')
ustw_afe = pickle.load( open( fname, "rb" ) )
fname=os.path.join(data_dir, 'ustw_index.pickle')
ustw_broad = pickle.load( open( fname, "rb" ) )

ustw=pd.merge(ustw_afe,ustw_broad,how='outer')

    ###Import dollar spot
fname=os.path.join(data_dir, 'dxy.pickle')
dxy = pickle.load( open( fname, "rb" ) )

    ###Import FXmonopairs
fname=os.path.join(data_dir, 'fxforward_index_bb.pickle')
fxmono = pickle.load( open( fname, "rb" ) )

    ###import FX spot
fname=os.path.join(data_dir, 'fx_spot.pickle')
fx_spot=pickle.load( open( fname, "rb" ) )

    ###import FX futures contracts
fname=os.path.join(data_dir, 'fx_fwd.pickle')
fx_fwd=pickle.load( open( fname, "rb" ) )

    #load FXRP iVol
fname=os.path.join(out_dir, 'fxrp_basket_ivol_m.pickle')
fxrp_basket_ivol_m = pickle.load( open( fname, "rb" ) )
fxrp_basket_ivol_m=fns.ret(fxrp_basket_ivol_m)


#remove NAs from dfs
def dropna_(df):
    df=df.dropna(thresh=2)
    return df


############################################
    ### CALCULATE RETURNS
############################################

    ### 1-month holding period for strategy
        # generate signal on last trading day of month
        # trade/rebalance on first trading day of month

dateseries=pd.DataFrame(assets.Date)
dateseries['month']=dateseries['Date'].dt.month
    #signal on last day of month
dateseries['temp']=(dateseries.month)-(dateseries.month).shift(-1)
dateseries['signalDay']=0
dateseries.loc[dateseries['temp']!=0,'signalDay']=1
    #trade on first day of month
dateseries['temp']=(dateseries.month)-(dateseries.month).shift(1)
dateseries['tradeDay']=0
dateseries.loc[dateseries['temp']!=0,'tradeDay']=1

dateseries=dateseries[['Date','signalDay','tradeDay']]

def monthly_ret(dat):
    dat=pd.merge(dateseries,dat,on='Date',how='left')
        #forward fill missing asset prices for monthly return calculation
    dat=dat.fillna(method='ffill')
        #subset to monthly observations
    dat_m=dat[dat['signalDay']==1].drop(['signalDay','tradeDay'],axis=1)
        #calculate monthly returns for previous month (t-1 to t)
    dat_m=fns.ret(dat_m).reset_index(drop=True)
    
    return dat_m

    # asset returns
assets_m = monthly_ret(assets)
    # spot returns (USD_FOR => goes up when $ appreciates)
fx_spot_m=fx_spot.copy()
#fx_spot_m.iloc[:,1:]=1/fx_spot_m.iloc[:,1:]
fx_spot_m= monthly_ret(fx_spot_m)

'''
compute FX mono returns from persepective of non-US investor 
monopairs are based on USD as base currency 
i.e. long FOREIGN against USD
'''
fxmono_m= monthly_ret(fxmono)
fxmono_m.iloc[:,1:]=1/(fxmono_m.iloc[:,1:]+1)-1

   #get monthly prices
def monthly_px(dat):
    dat=pd.merge(dateseries,dat,on='Date',how='left')
        #forward fill missing asset prices for monthly return calculation
    dat=dat.fillna(method='ffill')
        #subset to monthly observations
    dat_m=dat[dat['signalDay']==1].drop(['signalDay','tradeDay'],axis=1).reset_index(drop=True)
    
    return dat_m

fx_factors_reg_m = monthly_px(fx_factors.drop(['Value_Global','Carry_Global','DollarCarry_Global'],axis=1))

##########################################
###calculate asset returns in foreign FX##
##########################################

#perfect hedge i.e. returns in USD
assets_ph_m=assets_m.copy()

################
#add wrapper for currency
################
    
def curr_hedge_ratios(nick,curr):
    fx_m=fx_spot_m[['Date',curr]]
    mono_m=fxmono_m[['Date',curr]]
    
    temp=pd.DataFrame((1+assets_ph_m.iloc[:,1:].values)*(1+fx_m.iloc[:,1:].values)-1,columns=assets_ph_m.columns[1:],index=assets_ph_m.index)
    assets_m.iloc[:,1:]=temp
    
    
    #############################################################
        #Construct FXRP$
    #############################################################
    
    #DETERMINE OPTIMAL FACTOR WEIGHTS FROM ROLLING REGRESSIONS
    
        ## 5 year rolling OLS regression: FX returns ~ FXRP returns 
    fx_fxrp_b_m,fx_fxrp_p_m=fns.roll_reg(ydat=fx_m,xdat=fns.ret(fx_factors_reg_m).reset_index(drop=True),name=curr,roll_window=60)
    
    fxrp_basket_reg_m=fx_fxrp_b_m.drop(['Date','const'],axis=1)
    
    #multiply weights by factor values
    fxrp_basket_reg_m=fns.ret(fx_factors_reg_m).iloc[:,1:].mul(fxrp_basket_reg_m,axis=1)
    fxrp_basket_reg_m=fxrp_basket_reg_m.sum(axis=1)
    fxrp_basket_reg_m=pd.concat([pd.Series([100]),1+fxrp_basket_reg_m]).cumprod().iloc[1:]
    fxrp_basket_reg_m=pd.concat([fx_factors_reg_m['Date'],fxrp_basket_reg_m],axis=1)
    fxrp_basket_reg_m.rename(columns={0:'fxrp_basket_reg'},inplace=True)
    
    fxrp_basket_reg_m['fxrp_basket_reg'].plot(subplots=True)
    
    #writer=pd.ExcelWriter(os.path.join(out_dir,'test2.xlsx'))
    #fx_factors_reg.to_excel(writer,'factors')
    #fxrp_basket_reg.to_excel(writer,'betas')
    #writer.save()
    
    with open(os.path.join(out_dir, (nick+'_fxrp_basket_reg_m.pickle')), 'wb') as output:
        pickle.dump(fxrp_basket_reg_m, output)
    
    #############################################################
    #Calculate Optimal Hedge Ratios --  
    #    Rolling regressions > Asset Returns ~ Hedging Instrument Returns
    #############################################################
    
    asset_tickers=list(assets_m.columns)[1:]
    
    ### 5-Year Rolling OLS Regression
    fxrp_reg_hedge_b_m,fxrp_reg_hedge_p_m = fns.run_reg(ydat=assets_m,xdat=fns.ret(fxrp_basket_reg_m),names=asset_tickers,roll_window=60)
    
    #remove NAs from dfs
    fxrp_reg_hedge_b_m = {k: dropna_(v) for k, v in fxrp_reg_hedge_b_m.items()}
    fxrp_reg_hedge_p_m  = {k: dropna_(v) for k, v in fxrp_reg_hedge_p_m.items()}
    
    with open(os.path.join(out_dir, (nick+'_fxrp_reg_hedge_b_m.pickle')), 'wb') as output:
        pickle.dump(fxrp_reg_hedge_b_m, output)
    with open(os.path.join(out_dir, (nick+'_fxrp_reg_hedge_p_m.pickle')), 'wb') as output:
        pickle.dump(fxrp_reg_hedge_p_m, output)
    
    
    ##########USTW_AFE##################
    
    ### 5-Year Rolling Regression
    fx_hedge_b_m,fx_hedge_p_m = fns.run_reg(ydat=assets_m,xdat=mono_m,names=asset_tickers,roll_window=60)
    
    fx_hedge_b_m = {k: fns.dropna_(v) for k, v in fx_hedge_b_m.items()}
    fx_hedge_p_m = {k: fns.dropna_(v) for k, v in fx_hedge_p_m.items()}
    
    with open(os.path.join(out_dir, (nick+'_hedge_b_m.pickle')), 'wb') as output:
        pickle.dump(fx_hedge_b_m, output)
    with open(os.path.join(out_dir, (nick+'_hedge_p_m.pickle')), 'wb') as output:
        pickle.dump(fx_hedge_p_m, output)
    
    
    #############################################################
        #Construct FXRP Asset
    #############################################################
    
    #DETERMINE OPTIMAL FACTOR WEIGHTS FROM ROLLING REGRESSIONS
    
        ## 5 year rolling OLS regression: FX returns ~ FXRP returns 
    fx_fxrp_b_m,fx_fxrp_p_m=fns.run_reg(ydat=assets_m,xdat=fns.ret(fx_factors_reg_m).reset_index(drop=True),names=asset_tickers,roll_window=60)
    
    #remove NAs from dfs
    fx_fxrp_b_m = {k: dropna_(v) for k, v in fx_fxrp_b_m.items()}
    fx_fxrp_p_m  = {k: dropna_(v) for k, v in fx_fxrp_p_m.items()}
    
    with open(os.path.join(out_dir, (nick+'_fxrp_reg_hedge_b_m_fac.pickle')), 'wb') as output:
        pickle.dump(fx_fxrp_b_m, output)
    with open(os.path.join(out_dir, (nick+'_fxrp_reg_hedge_p_m_fac.pickle')), 'wb') as output:
        pickle.dump(fx_fxrp_p_m, output)
        
        
    fxrp_basket_reg_m_fac={}
    for p in fx_fxrp_b_m.keys():
        basket=fx_fxrp_b_m[p].drop(['Date','const'],axis=1)
        #apply cap = 2x
        cap=2 
        basket=basket.clip_upper(cap)
        basket=basket.clip_lower(-cap)
    
        #multiply weights by factor values
        basket=fns.ret_next(fx_factors_reg_m).iloc[:,1:].mul(basket,axis=1)
        basket=basket.sum(axis=1)
        basket=pd.concat([fx_fxrp_b_m[p][['Date','const']],basket],axis=1).replace(0, np.nan)
        basket.rename(columns={0:'fxrp_basket_reg'},inplace=True)
        fxrp_basket_reg_m_fac[p]=basket
    
    
    #remove NAs from dfs
    fxrp_basket_reg_m_fac = {k: dropna_(v) for k, v in fxrp_basket_reg_m_fac.items()}
    
    with open(os.path.join(out_dir, (nick+'_fxrp_basket_reg_m_fac.pickle')), 'wb') as output:
        pickle.dump(fxrp_basket_reg_m_fac, output)


curr_hedge_ratios('aud','Australia')
curr_hedge_ratios('jpy','Japan')
curr_hedge_ratios('eur','European Union')
curr_hedge_ratios('gbp','United Kingdom')
curr_hedge_ratios('cad','Canada')
curr_hedge_ratios('sek','Sweden')
curr_hedge_ratios('chf','Switzerland')

curr_hedge_ratios('brl','Brazil')
curr_hedge_ratios('cny','China')
curr_hedge_ratios('mxn','Mexico')


##########################################
###Strategy returns
##########################################

## load currency specific hedge ratios
###AUD
fname=os.path.join(out_dir, 'aud_fxrp_basket_reg_m.pickle')
aud_fxrp_basket_reg_m=pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'aud_fxrp_reg_hedge_b_m.pickle')
aud_fxrp_reg_hedge_b_m = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'aud_hedge_b_m.pickle')
aud_hedge_b_m = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'aud_fxrp_reg_hedge_b_m_fac.pickle')
aud_fxrp_reg_hedge_b_m_fac = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'aud_fxrp_basket_reg_m_fac.pickle')
aud_fxrp_basket_reg_m_fac = pickle.load( open( fname, "rb" ) )

###JPY
fname=os.path.join(out_dir, 'jpy_fxrp_basket_reg_m.pickle')
jpy_fxrp_basket_reg_m=pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'jpy_fxrp_reg_hedge_b_m.pickle')
jpy_fxrp_reg_hedge_b_m = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'jpy_hedge_b_m.pickle')
jpy_hedge_b_m = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'jpy_fxrp_reg_hedge_b_m_fac.pickle')
jpy_fxrp_reg_hedge_b_m_fac = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'jpy_fxrp_basket_reg_m_fac.pickle')
jpy_fxrp_basket_reg_m_fac = pickle.load( open( fname, "rb" ) )

###EUR
fname=os.path.join(out_dir, 'eur_fxrp_basket_reg_m.pickle')
eur_fxrp_basket_reg_m=pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'eur_fxrp_reg_hedge_b_m.pickle')
eur_fxrp_reg_hedge_b_m = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'eur_hedge_b_m.pickle')
eur_hedge_b_m = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'eur_fxrp_reg_hedge_b_m_fac.pickle')
eur_fxrp_reg_hedge_b_m_fac = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'eur_fxrp_basket_reg_m_fac.pickle')
eur_fxrp_basket_reg_m_fac = pickle.load( open( fname, "rb" ) )

###GBP
fname=os.path.join(out_dir, 'gbp_fxrp_basket_reg_m.pickle')
gbp_fxrp_basket_reg_m=pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'gbp_fxrp_reg_hedge_b_m.pickle')
gbp_fxrp_reg_hedge_b_m = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'gbp_hedge_b_m.pickle')
gbp_hedge_b_m = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'gbp_fxrp_reg_hedge_b_m_fac.pickle')
gbp_fxrp_reg_hedge_b_m_fac = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'gbp_fxrp_basket_reg_m_fac.pickle')
gbp_fxrp_basket_reg_m_fac = pickle.load( open( fname, "rb" ) )

###CAD
fname=os.path.join(out_dir, 'cad_fxrp_basket_reg_m.pickle')
cad_fxrp_basket_reg_m=pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'cad_fxrp_reg_hedge_b_m.pickle')
cad_fxrp_reg_hedge_b_m = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'cad_hedge_b_m.pickle')
cad_hedge_b_m = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'cad_fxrp_reg_hedge_b_m_fac.pickle')
cad_fxrp_reg_hedge_b_m_fac = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'cad_fxrp_basket_reg_m_fac.pickle')
cad_fxrp_basket_reg_m_fac = pickle.load( open( fname, "rb" ) )

###SEK
fname=os.path.join(out_dir, 'sek_fxrp_basket_reg_m.pickle')
sek_fxrp_basket_reg_m=pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'sek_fxrp_reg_hedge_b_m.pickle')
sek_fxrp_reg_hedge_b_m = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'sek_hedge_b_m.pickle')
sek_hedge_b_m = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'sek_fxrp_reg_hedge_b_m_fac.pickle')
sek_fxrp_reg_hedge_b_m_fac = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'sek_fxrp_basket_reg_m_fac.pickle')
sek_fxrp_basket_reg_m_fac = pickle.load( open( fname, "rb" ) )

###CHF
fname=os.path.join(out_dir, 'chf_fxrp_basket_reg_m.pickle')
chf_fxrp_basket_reg_m=pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'chf_fxrp_reg_hedge_b_m.pickle')
chf_fxrp_reg_hedge_b_m = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'chf_hedge_b_m.pickle')
chf_hedge_b_m = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'chf_fxrp_reg_hedge_b_m_fac.pickle')
chf_fxrp_reg_hedge_b_m_fac = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'chf_fxrp_basket_reg_m_fac.pickle')
chf_fxrp_basket_reg_m_fac = pickle.load( open( fname, "rb" ) )

###BRL
fname=os.path.join(out_dir, 'brl_fxrp_basket_reg_m.pickle')
brl_fxrp_basket_reg_m=pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'brl_fxrp_reg_hedge_b_m.pickle')
brl_fxrp_reg_hedge_b_m = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'brl_hedge_b_m.pickle')
brl_hedge_b_m = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'brl_fxrp_reg_hedge_b_m_fac.pickle')
brl_fxrp_reg_hedge_b_m_fac = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'brl_fxrp_basket_reg_m_fac.pickle')
brl_fxrp_basket_reg_m_fac = pickle.load( open( fname, "rb" ) )

###CNY
fname=os.path.join(out_dir, 'cny_fxrp_basket_reg_m.pickle')
cny_fxrp_basket_reg_m=pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'cny_fxrp_reg_hedge_b_m.pickle')
cny_fxrp_reg_hedge_b_m = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'cny_hedge_b_m.pickle')
cny_hedge_b_m = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'cny_fxrp_reg_hedge_b_m_fac.pickle')
cny_fxrp_reg_hedge_b_m_fac = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'cny_fxrp_basket_reg_m_fac.pickle')
cny_fxrp_basket_reg_m_fac = pickle.load( open( fname, "rb" ) )

###MXN
fname=os.path.join(out_dir, 'mxn_fxrp_basket_reg_m.pickle')
mxn_fxrp_basket_reg_m=pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'mxn_fxrp_reg_hedge_b_m.pickle')
mxn_fxrp_reg_hedge_b_m = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'mxn_hedge_b_m.pickle')
mxn_hedge_b_m = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'mxn_fxrp_reg_hedge_b_m_fac.pickle')
mxn_fxrp_reg_hedge_b_m_fac = pickle.load( open( fname, "rb" ) )
fname=os.path.join(out_dir, 'mxn_fxrp_basket_reg_m_fac.pickle')
mxn_fxrp_basket_reg_m_fac = pickle.load( open( fname, "rb" ) )

def strategy_returns(assetreturns,hedge,hedge_wt,name):
    '''
    input assets returns (flexible to frequency of returns),
            hedging instrument returns (e.g. USTW$ or FXRP Basket),
            optimal hedge ratio dataframe (based on rolling regressions),
            & name of the hedging instrument variable in str format (e.g. 'USTW$ Index')
    output dictionary containing (asset: strategy returns) pairs 
            strategy returns computed as (asset return - 100.beta*hedge_instrument return)        
    '''
    strat_ret={}
    for p in assetreturns.columns.drop(['Date']):
        ### first merge rolling hedge ratios with date series
        signal=pd.merge(dateseries,hedge_wt[p],on='Date',how='left')
            ### next forward fill signals to ensure non-missing values on signal days
        signal=signal.fillna(method='ffill')
        signal=signal.dropna(subset=hedge_wt[p].columns.drop(['Date','const']),how='all')
            ### subset to signal days and create trade date series (t+1)
        signal=signal[(signal['signalDay']==1) | (signal['tradeDay']==1)]
        signal['tradeDate']=signal['Date'].shift(-1)
        signal=signal[(signal['signalDay']==1)]
        signal=signal.drop(['signalDay','tradeDay','const'],axis=1)
        signal.rename(columns = {name:'hedge_ratio'}, 
                      inplace=True)
            ### merge with hedging instrument returns on trade date (t+1)
        hedging_instrument=pd.merge(signal,hedge,left_on='tradeDate',right_on='Date',how='left',suffixes=('_x',''))
            ### invest $(beta*100) in hedging instrument on signal date with 1M holding period
        hedging_instrument['hedge_ins_ret']=(hedging_instrument['hedge_ratio'])*hedging_instrument[name]
        hedging_instrument=hedging_instrument[['Date','hedge_ins_ret','hedge_ratio']]
        
        strat_ret[p]=pd.merge(assetreturns[['Date',p]],hedging_instrument,how='right',on='Date')
        strat_ret[p]['strat_ret']=strat_ret[p][p]-strat_ret[p]['hedge_ins_ret']
    
    return strat_ret


    ### Compute strategy returns ###

    # fully hedged, hedging instrument: USTW$
    # create toy dictionary with 100% hedge ratios to pass to function
def f(dat,name):
    temp = dat.copy()
    temp[name]=1
    return temp


#############################################################
#############################################################
# ANALYSIS USING MONTHLY HEDGE RATIOS
#############################################################
#############################################################

def monthly_ret(dat):
    dat=pd.merge(dateseries,dat,on='Date',how='left')
        #forward fill missing asset prices for monthly return calculation
    dat=dat.fillna(method='ffill')
        #subset to monthly observations
    dat_m=dat[dat['tradeDay']==1].drop(['signalDay','tradeDay'],axis=1)
        #calculate monthly returns for subsequent month (t+1)
    dat_m=fns.ret_next(dat_m).reset_index(drop=True)
    
    return dat_m

assets_m = monthly_ret(assets)
fxmono_m= monthly_ret(fxmono)

#perfect hedge i.e. returns in USD
assets_ph_m=assets_m.copy()

#############################################################
#4. Calculate strategy returns
#############################################################

def strats(nick,curr):
    fx_m=fx_spot_m[['Date',curr]]
    mono_m=fxmono_m[['Date',curr]]
    
    temp=pd.DataFrame((1+assets_ph_m.iloc[:,1:].values)*(1+fx_m.iloc[:,1:].values)-1,columns=assets_ph_m.columns[1:],index=assets_ph_m.index)
    assets_m.iloc[:,1:]=temp

    _hedge_b_m=eval(nick+'_hedge_b_m')
    _fxrp_basket_reg_m=eval(nick+'_fxrp_basket_reg_m')
    _fxrp_reg_hedge_b_m=eval(nick+'_fxrp_reg_hedge_b_m')
    _fxrp_basket_reg_m_fac=eval(nick+'_fxrp_basket_reg_m_fac')
    _fxrp_reg_hedge_b_m_fac=eval(nick+'_fxrp_reg_hedge_b_m_fac')
    
    
        # Fully hedged with USTW$
    full_hedge_m={p:f(v,curr) for p,v in _hedge_b_m.items()}
    hedge_full_m=strategy_returns(assetreturns=assets_m,
                                    hedge=mono_m,
                                    hedge_wt=full_hedge_m,
                                    name=curr)
        # hedging instrument: USTW$
    hedge_ustw_afe_m=strategy_returns(assetreturns=assets_m,
                                    hedge=mono_m,
                                    hedge_wt=_hedge_b_m,
                                    name=curr)
    
        # hedging instrument: FXRP OLS Reg BASKET
    hedge_fxrp_reg_m=strategy_returns(assetreturns=assets_m,
                                    hedge=monthly_ret(_fxrp_basket_reg_m),
                                    hedge_wt=_fxrp_reg_hedge_b_m,
                                    name='fxrp_basket_reg')
    
        # hedging instrument: FXRP OLS Reg BASKET Asset Specific
    dummy=monthly_ret(_fxrp_basket_reg_m).copy()    
    dummy['fxrp_basket_reg']=1
        
    hedge_fxrp_reg_m_fac=strategy_returns(assetreturns=assets_m,
                                    hedge=dummy,
                                    hedge_wt=_fxrp_basket_reg_m_fac,
                                    name='fxrp_basket_reg')
    
    
        # perfect hedge
    hedge_perf_m=strategy_returns(assetreturns=assets_ph_m,
                                    hedge=mono_m,
                                    hedge_wt=full_hedge_m,
                                    name=curr)
    
    #############################################################
    #4. Compare strategy performance MONTHLY
    #############################################################
    
    compare_strategies_m={}
    for p in hedge_ustw_afe_m.keys():
        
        ### merge all hedging strategies (FXRP, USTW AFE) -- UPDATE: EXCLUDE FIRST n OBS (n=60 rolling window for optimal hedge ratio regressions)
        compare_strategies_m[p] = pd.merge(hedge_full_m[p],hedge_ustw_afe_m[p],on=['Date'],how='inner',suffixes=('','_tw_afe')).iloc[60:,]
        compare_strategies_m[p] = pd.merge(compare_strategies_m[p],hedge_fxrp_reg_m[p],on=['Date'],how='inner',suffixes=('','_fxrp_reg'))
        compare_strategies_m[p] = pd.merge(compare_strategies_m[p],hedge_fxrp_reg_m_fac[p],on=['Date'],how='inner',suffixes=('','_fxrp_reg_fac'))
        compare_strategies_m[p] = pd.merge(compare_strategies_m[p],hedge_perf_m[p],on=['Date'],how='inner',suffixes=('','_perfect'))
          
        ### compute strategy returns
        
        # Unhedged
        compare_strategies_m[p]['_']=(pd.concat([pd.Series([100]),1+compare_strategies_m[p][p]])).cumprod().iloc[1:]
    
        # Full
        compare_strategies_m[p]['_full']=(pd.concat([pd.Series([100]),1+compare_strategies_m[p]['strat_ret']])).cumprod().iloc[1:]
    
        # USTW AFE
        compare_strategies_m[p]['_twd_afe']=(pd.concat([pd.Series([100]),1+compare_strategies_m[p]['strat_ret_tw_afe']])).cumprod().iloc[1:]
     
        # FXRP Reg
        compare_strategies_m[p]['_fxrp_reg']=(pd.concat([pd.Series([100]),1+compare_strategies_m[p]['strat_ret_fxrp_reg']])).cumprod().iloc[1:]
    
        # FXRP Reg Asset specific   
        compare_strategies_m[p]['_fxrp_reg_fac']=(pd.concat([pd.Series([100]),1+compare_strategies_m[p]['strat_ret_fxrp_reg_fac']])).cumprod().iloc[1:]
    
        # Perfect Hedge
        compare_strategies_m[p]['_perfect']=(pd.concat([pd.Series([100]),1+compare_strategies_m[p][p+'_perfect']])).cumprod().iloc[1:]
    
    ###TRACKING ERRORS
        compare_strategies_m[p]['strat_ret_tw_afe_te'] = (compare_strategies_m[p][p+'_perfect']-compare_strategies_m[p]['strat_ret_tw_afe'])    
        compare_strategies_m[p]['strat_ret_fxrp_reg_te'] = (compare_strategies_m[p][p+'_perfect']-compare_strategies_m[p]['strat_ret_fxrp_reg'])
        compare_strategies_m[p]['strat_ret_fxrp_reg_fac_te'] = (compare_strategies_m[p][p+'_perfect']-compare_strategies_m[p]['strat_ret_fxrp_reg_fac'])
        
    #############################################################
    # Create barcharts for returns, vol, info ratios
    #############################################################
    
    from dateutil.relativedelta import relativedelta
    
        #storage dictionary containing dictionary for each asset with a dataframe for each period
    compare_sstats_m={}
    for p in hedge_ustw_afe_m.keys():
        #strategies to compare
        strats=[p,'strat_ret','strat_ret_tw_afe','strat_ret_fxrp_reg','strat_ret_fxrp_reg_fac',p+'_perfect']
        #compute strategy return summary stats for last n years
        def last_n(n):
            _lastn=compare_strategies_m[p].loc[compare_strategies_m[p]['Date']>(dt.datetime.today()-relativedelta(years=n)),:][strats]
            mean_lastn = _lastn.mean()*12*100
            vol_lastn  = _lastn.std()*pow(12,0.5)*100
            info_lastn = mean_lastn / vol_lastn  
            lastn=pd.concat([mean_lastn,vol_lastn,info_lastn],axis=1)
            lastn.columns=[s + str(n) for s in ['mean_last','vol_last','info_last']]
            return lastn
        
        last_10=last_n(10)
        last_5 =last_n(5)
        last_3 =last_n(3)
        last_1 =last_n(1)
    
        def pre_n(n):
            _lastn=compare_strategies_m[p].loc[compare_strategies_m[p]['Date']<(dt.datetime.today()-relativedelta(years=n)),:][strats]
            mean_lastn = _lastn.mean()*12*100
            vol_lastn  = _lastn.std()*pow(12,0.5)*100
            info_lastn = mean_lastn / vol_lastn  
            lastn=pd.concat([mean_lastn,vol_lastn,info_lastn],axis=1)
            lastn.columns=[s + str(n) for s in ['mean_pre','vol_pre','info_pre']]
            return lastn
    
        pre_10=pre_n(10)
    
    
        hedge=['hedge_ratio_tw_afe','hedge_ratio_fxrp_reg']
        #compute strategy return summary stats for last n years
        def hedge_n(n):
            _hedgen=compare_strategies_m[p].loc[compare_strategies_m[p]['Date']>(dt.datetime.today()-relativedelta(years=n)),:][hedge]
            hedge_lastn = pd.DataFrame(-1*_hedgen.median())
            hedge_lastn.columns=[s + str(n) for s in ['hedge_last']]
            return hedge_lastn 
        
        hedge_10=hedge_n(10)
        hedge_5 =hedge_n(5)
        hedge_3 =hedge_n(3)
        hedge_1 =hedge_n(1)
        
        #compute tracking errors for last n years
        te=['strat_ret_tw_afe_te','strat_ret_fxrp_reg_te','strat_ret_fxrp_reg_fac_te']
        def te_n(n):
            _ten=compare_strategies_m[p].loc[compare_strategies_m[p]['Date']>(dt.datetime.today()-relativedelta(years=n)),:][te]
            te_lastn = pd.DataFrame(_ten.std())*pow(12,0.5)*100
            te_lastn.columns=[s + str(n) for s in ['te_last']]
            return te_lastn 
        
        te_10=te_n(10)
        te_5 =te_n(5)
        te_3 =te_n(3)
        te_1 =te_n(1)
    
        compare_sstats_m[p]={
                "mean":pd.concat([pre_10['mean_pre10'],last_10['mean_last10'],last_5['mean_last5'],last_3['mean_last3'],last_1['mean_last1']],axis=1),
                "vol":pd.concat([pre_10['vol_pre10'],last_10['vol_last10'],last_5['vol_last5'],last_3['vol_last3'],last_1['vol_last1']],axis=1),
                "info":pd.concat([pre_10['info_pre10'],last_10['info_last10'],last_5['info_last5'],last_3['info_last3'],last_1['info_last1']],axis=1),
                "hedge":pd.concat([hedge_10,hedge_5,hedge_3,hedge_1],axis=1),
                "te":pd.concat([te_10,te_5,te_3,te_1],axis=1)            
            }
        
    #############################################################
                    # Summary of Results -- for EXCEL
    #############################################################
    
    optimal_hr_10=pd.DataFrame()
    optimal_hr_5=pd.DataFrame()
    optimal_hr_3=pd.DataFrame()
    
    ret_pre=pd.DataFrame()
    ret_10=pd.DataFrame()
    ret_5=pd.DataFrame()
    ret_3=pd.DataFrame()
    
    vol_pre=pd.DataFrame()
    vol_10=pd.DataFrame()
    vol_5=pd.DataFrame()
    vol_3=pd.DataFrame()
    
    info_pre=pd.DataFrame()
    info_10=pd.DataFrame()
    info_5=pd.DataFrame()
    info_3=pd.DataFrame()
    
    te_10=pd.DataFrame()
    te_5=pd.DataFrame()
    te_3=pd.DataFrame()
    
    for p in compare_sstats_m.keys():
        optimal_hr_10[p]=compare_sstats_m[p]['hedge'].hedge_last10.reset_index(drop=True)
        optimal_hr_5[p]=compare_sstats_m[p]['hedge'].hedge_last5.reset_index(drop=True)
        optimal_hr_3[p]=compare_sstats_m[p]['hedge'].hedge_last3.reset_index(drop=True)
        ret_pre[p]=compare_sstats_m[p]['mean'].mean_pre10.reset_index(drop=True)
        ret_10[p]=compare_sstats_m[p]['mean'].mean_last10.reset_index(drop=True)
        ret_5[p]=compare_sstats_m[p]['mean'].mean_last5.reset_index(drop=True)
        ret_3[p]=compare_sstats_m[p]['mean'].mean_last3.reset_index(drop=True)
        vol_pre[p]=compare_sstats_m[p]['vol'].vol_pre10.reset_index(drop=True)
        vol_10[p]=compare_sstats_m[p]['vol'].vol_last10.reset_index(drop=True)
        vol_5[p]=compare_sstats_m[p]['vol'].vol_last5.reset_index(drop=True)
        vol_3[p]=compare_sstats_m[p]['vol'].vol_last3.reset_index(drop=True)
        info_pre[p]=compare_sstats_m[p]['info'].info_pre10.reset_index(drop=True)
        info_10[p]=compare_sstats_m[p]['info'].info_last10.reset_index(drop=True)
        info_5[p]=compare_sstats_m[p]['info'].info_last5.reset_index(drop=True)
        info_3[p]=compare_sstats_m[p]['info'].info_last3.reset_index(drop=True)
        te_10[p]=compare_sstats_m[p]['te'].te_last10.reset_index(drop=True)
        te_5[p] =compare_sstats_m[p]['te'].te_last5.reset_index(drop=True)
        te_3[p] =compare_sstats_m[p]['te'].te_last3.reset_index(drop=True)
        
    
    optimal_hr_10.index=['USTW$','Optimal']
    optimal_hr_5.index=['USTW$','Optimal']
    optimal_hr_3.index=['USTW$','Optimal']
    ret_10.index=['Unhedged','Full','USTW$','Opt','Opt-Asset','USD']
    ret_5.index=['Unhedged','Full','USTW$','Opt','Opt-Asset','USD']
    ret_3.index=['Unhedged','Full','USTW$','Opt','Opt-Asset','USD']
    vol_pre.index=['Unhedged','Full','USTW$','Opt','Opt-Asset','USD']
    vol_10.index=['Unhedged','Full','USTW$','Opt','Opt-Asset','USD']
    vol_5.index=['Unhedged','Full','USTW$','Opt','Opt-Asset','USD']
    vol_3.index=['Unhedged','Full','USTW$','Opt','Opt-Asset','USD']
    info_pre.index=['Unhedged','Full','USTW$','Opt','Opt-Asset','USD']
    info_10.index=['Unhedged','Full','USTW$','Opt','Opt-Asset','USD']
    info_5.index=['Unhedged','Full','USTW$','Opt','Opt-Asset','USD']
    info_3.index=['Unhedged','Full','USTW$','Opt','Opt-Asset','USD']
    te_10.index=['USTW$','Opt','Opt-Asset']
    te_5.index=['USTW$','Opt','Opt-Asset']
    te_3.index=['USTW$','Opt','Opt-Asset']

    
    list_dfs=[optimal_hr_10,optimal_hr_5,optimal_hr_3,ret_pre,ret_10,ret_5,ret_3,vol_pre,vol_10,vol_5,vol_3,info_pre,info_10,info_5,info_3,te_10,te_5,te_3]
 
    #rolling betas
    writer=pd.ExcelWriter(os.path.join(out_dir,(nick+'_rollingB_m.xlsx')))
    writer.save()
    for p in compare_sstats_m.keys():
        temp=pd.merge(_fxrp_reg_hedge_b_m[p].drop(['const'],axis=1),_hedge_b_m[p].drop(['const'],axis=1),on='Date',how='inner').iloc[60:,:]
        temp.iloc[:,1:]=-1*temp.iloc[:,1:]
        from openpyxl import load_workbook
        book= load_workbook(os.path.join(out_dir,(nick+'_rollingB_m.xlsx')))
        writer=pd.ExcelWriter(os.path.join(out_dir,(nick+'_rollingB_m.xlsx')),engine='openpyxl')
        writer.book=book
        temp.to_excel(writer,(p))
        writer.save()
    
        #asset-specific hedge ratios
    def save_xls_dict(dict_dfs, xls_path):
        with pd.ExcelWriter(xls_path) as writer:
            for name, df in dict_dfs.items():
                temp=df.iloc[60:,:]
                temp.iloc[:,1:]=-1*temp.iloc[:,1:]
                temp.to_excel(writer,name)
            writer.save()
    
    save_xls_dict(_fxrp_reg_hedge_b_m_fac, os.path.join(out_dir,(nick+'_rollingB_m_faclevel.xlsx')))
    
    
    ## info ratios
    #rolling betas
    writer=pd.ExcelWriter(os.path.join(out_dir,(nick+'_inforatios.xlsx')))
    writer.save()
    for p in compare_sstats_m.keys():
        from openpyxl import load_workbook
        book= load_workbook(os.path.join(out_dir,(nick+'_inforatios.xlsx')))
        writer=pd.ExcelWriter(os.path.join(out_dir,(nick+'_inforatios.xlsx')),engine='openpyxl')
        writer.book=book
        compare_sstats_m[p]['info'].to_excel(writer,(p))
        writer.save()
    

    return compare_strategies_m, compare_sstats_m, list_dfs


aud_compare_strategies,aud_compare_sstats_m,aud_list_dfs = strats('aud','Australia')
jpy_compare_strategies,jpy_compare_sstats_m,jpy_list_dfs = strats('jpy','Japan')
eur_compare_strategies,eur_compare_sstats_m,eur_list_dfs = strats('eur','European Union')
gbp_compare_strategies,gbp_compare_sstats_m,gbp_list_dfs = strats('gbp','United Kingdom')
cad_compare_strategies,cad_compare_sstats_m,cad_list_dfs = strats('cad','Canada')
sek_compare_strategies,sek_compare_sstats_m,sek_list_dfs = strats('sek','Sweden')
chf_compare_strategies,chf_compare_sstats_m,chf_list_dfs = strats('chf','Switzerland')

brl_compare_strategies,chf_compare_sstats_m,brl_list_dfs = strats('brl','Brazil')
cny_compare_strategies,chf_compare_sstats_m,cny_list_dfs = strats('cny','China')
mxn_compare_strategies,chf_compare_sstats_m,mxn_list_dfs = strats('mxn','Mexico')

def save_xls(list_dfs, xls_path):
    with pd.ExcelWriter(xls_path) as writer:
        for n, df in enumerate(list_dfs):
            df.to_excel(writer,df_names[n])
        writer.save()
    
df_names=['optimal_hr_10','optimal_hr_5','optimal_hr_3','ret_pre','ret_10','ret_5','ret_3','vol_pre','vol_10','vol_5','vol_3','info_pre','info_10','info_5','info_3','te_10','te_5','te_3']

aud_list_dfs=list(map(lambda x: x.T,aud_list_dfs))   
save_xls(aud_list_dfs, os.path.join(out_dir,('aud_assetComparison.xlsx')))

jpy_list_dfs=list(map(lambda x: x.T,jpy_list_dfs))   
save_xls(jpy_list_dfs, os.path.join(out_dir,('jpy_assetComparison.xlsx')))

eur_list_dfs=list(map(lambda x: x.T,eur_list_dfs))   
save_xls(eur_list_dfs, os.path.join(out_dir,('eur_assetComparison.xlsx')))
            
gbp_list_dfs=list(map(lambda x: x.T,gbp_list_dfs))   
save_xls(gbp_list_dfs, os.path.join(out_dir,('gbp_assetComparison.xlsx')))

cad_list_dfs=list(map(lambda x: x.T,cad_list_dfs))   
save_xls(cad_list_dfs, os.path.join(out_dir,('cad_assetComparison.xlsx')))

sek_list_dfs=list(map(lambda x: x.T,sek_list_dfs))   
save_xls(sek_list_dfs, os.path.join(out_dir,('sek_assetComparison.xlsx')))

chf_list_dfs=list(map(lambda x: x.T,chf_list_dfs))   
save_xls(chf_list_dfs, os.path.join(out_dir,('chf_assetComparison.xlsx')))
    
brl_list_dfs=list(map(lambda x: x.T,brl_list_dfs))   
save_xls(brl_list_dfs, os.path.join(out_dir,('brl_assetComparison.xlsx')))

cny_list_dfs=list(map(lambda x: x.T,cny_list_dfs))   
save_xls(cny_list_dfs, os.path.join(out_dir,('cny_assetComparison.xlsx')))

mxn_list_dfs=list(map(lambda x: x.T,mxn_list_dfs))   
save_xls(mxn_list_dfs, os.path.join(out_dir,('mxn_assetComparison.xlsx')))






















