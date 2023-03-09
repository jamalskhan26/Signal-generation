# -*- coding: utf-8 -*-
"""
Created on Wed Jan 29 09:11:34 2020

@author: ZK463GK
"""

from IPython import get_ipython
get_ipython().magic('reset -sf')
import pandas as pd
import numpy as np
import os
import _pickle as pickle
import time

data_dir = '//rschfiler2.baml.com/UK-NYCOMMODITIES/Jamal/CnE/Data'
out_dir = '//rschfiler2.baml.com/UK-NYCOMMODITIES/Jamal/CnE/Output'

wd = os.getcwd()
os.chdir('//rschfiler2.baml.com/UK-NYCOMMODITIES/Jamal')
from regression_models import regression_models
os.chdir(wd)

energy =      ['Brent','WTI','Natural Gas','Heating Oil','Gasoline','Gasoil']
ags =         ['Cocoa','Coffee','Corn','Cotton','Sugar','Wheat','Kansas Wheat','Soybean Oil','Soybeans','Soybean Meal','Orange Juice']
industrials = ['Aluminum','Nickel','Zinc','Tin','Copper']
precious =    ['Gold','Silver','Palladium','Platinum']
livestock =   ['Lean Hogs','Live Cattle','Feeder Cattle'] 
_all =        energy + ags + industrials + precious + livestock + ['Softs','Grains','Industrial Metals','Livestock','Precious Metals','Petroleum']

##################
#output hit ratios
##################

fname=os.path.join(data_dir, "encModels_1Ylookback.pickle")
model = pickle.load( open( fname, "rb" ))
model.ensemble()
model.performance_stats()

writer=pd.ExcelWriter(os.path.join(data_dir, 'actual_returns.xlsx'))
writer.save() 
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir, 'actual_returns.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir, 'actual_returns.xlsx'),engine='openpyxl')
writer.book=book
model.dates.to_excel(writer,'dates')
model.actual_returns.to_excel(writer,'CO')
model.X.to_excel(writer,'FX')
writer.save()


writer=pd.ExcelWriter(os.path.join(data_dir, 'enc_1Y_hit_summary.xlsx'))
writer.save() 
for p in model.hit_ratio.keys(): 
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, 'enc_1Y_hit_summary.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, 'enc_1Y_hit_summary.xlsx'),engine='openpyxl')
    writer.book=book
    model.hit_ratio[p].to_excel(writer,p)
    writer.save()
    

sectors = [energy, ags, industrials, precious, livestock, _all]
nick = ["Energy","Agricultural","Industrial metals","Precious metals","Livestock","All"]
hit_year_sector = {}

for s,n in zip(sectors,nick):
    temp = pd.DataFrame(index=list(range(2005,2021)),columns=model.models).replace(np.nan,0)
    for p in s:
        temp = temp.add(model.hit_ratio_by_year[p])
    hit_year_sector[n] = temp / len(s)
        
writer=pd.ExcelWriter(os.path.join(data_dir, 'enc_1Y_hit_yearly_summary.xlsx'))
writer.save() 
for p in hit_year_sector.keys(): 
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, 'enc_1Y_hit_yearly_summary.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, 'enc_1Y_hit_yearly_summary.xlsx'),engine='openpyxl')
    writer.book=book
    hit_year_sector[p].to_excel(writer,p)
    writer.save()
    

best_models = pd.DataFrame()
for p in model.hit_ratio.keys(): 
    best_models.loc[p,0] = model.hit_ratio[p].loc[['LR','RR','SVD','DT','SVM'],'Hit_Ratio'].idxmax(axis=1)
    best_models.loc[p,1] = model.hit_ratio[p].loc[['LR','RR','SVD','DT','SVM'],'Hit_Ratio'].max()

best_models.columns = ['best','hr']

from scipy.stats import skew, kurtosis 
best_models['vol'] = model.actual_returns.std()*(52**0.5)
best_models['skew'] = model.actual_returns.apply(lambda x: skew(x))
best_models['kurt'] = model.actual_returns.apply(lambda x: kurtosis(x))

best_models = best_models.sort_values(by='best')
best_models = best_models[['best','vol','skew','kurt','hr']] 

writer=pd.ExcelWriter(os.path.join(data_dir, 'best_models.xlsx'))
writer.save() 
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir, 'best_models.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir, 'best_models.xlsx'),engine='openpyxl')
writer.book=book
best_models.to_excel(writer,'best_models')
writer.save()
        
##################

#############################################################
# Function calls
#############################################################
#EnC
#######################################################################################################################################################################################

def generate_charts_enc(datestart, fname, version, pdfname):
    # load model object
    fname=os.path.join(data_dir, fname)
    model = pickle.load( open( fname, "rb" ))
    
    # generate performance stats -- no signal weighting
    model.ensemble = regression_models.ensemble
    model.ensemble(self = model)
    model.performance_stats()
    
    for p in model.hit_ratio_by_year.keys(): 
        model.hit_ratio_by_year[p] =  model.hit_ratio_by_year[p][model.hit_ratio_by_year[p].index!=2020]
        model.info_ratio_by_year[p] = model.info_ratio_by_year[p][model.info_ratio_by_year[p].index!=2020]
        model.ann_returns_by_year[p] = model.ann_returns_by_year[p][model.ann_returns_by_year[p].index!=2020]
      
    model.charting =       regression_models.charting
    model.charting_strat = regression_models.charting_strat

    model.charting(self = model, specs = version + ", no signal weighting", fname = pdfname + "_nw.pdf", out_dir = out_dir)

    #model.actual_returns = model.actual_returns[model.actual_returns[model.mergevar] > datestart]
    model.actual_returns = model.actual_returns.iloc[datestart:, :]
   
    #construct strategies
    #1 - ALL Commodities
    model.construct_strategy(basket = _all)
    model.charting_strat(self = model, strat_nick = "All Commodities", specs = version + ", no signal weighting", fname = pdfname + "_nw_all_strat.pdf", out_dir = out_dir)
    #2 - Energy
    model.construct_strategy(basket = energy)
    model.charting_strat(self = model, strat_nick = "Energy", specs = version + ", no signal weighting", fname = pdfname + "_nw_co1_strat.pdf", out_dir = out_dir)
    #3 - Agriculture
    model.construct_strategy(basket = ags)
    model.charting_strat(self = model, strat_nick = "Agriculture", specs = version + ", no signal weighting", fname = pdfname + "_nw_co2_strat.pdf", out_dir = out_dir)
    #4 - Industrial Metals
    model.construct_strategy(basket = industrials)
    model.charting_strat(self = model, strat_nick = "Industrial Metals", specs = version + ", no signal weighting", fname = pdfname + "_nw_co3_strat.pdf", out_dir = out_dir)
    #4 - Precious Metals
    model.construct_strategy(basket = precious)
    model.charting_strat(self = model, strat_nick = "Precious Metals", specs = version + ", no signal weighting", fname = pdfname + "_nw_co4_strat.pdf", out_dir = out_dir)
    #5 - Livestock
    model.construct_strategy(basket = livestock)
    model.charting_strat(self = model, strat_nick = "Livestock", specs = version + ", no signal weighting", fname = pdfname + "_nw_co5_strat.pdf", out_dir = out_dir)


### 1: 5Y lookback
generate_charts_enc(datestart = 260,
                    fname = 'encModels_5Ylookback.pickle',
                    version = '5Y lookback',
                    pdfname = 'EnC_5Yroll')

### 2: 3Y lookback
generate_charts_enc(datestart = 160,
                fname = 'encModels_3Ylookback.pickle',
                version = '3Y lookback',
                pdfname = 'EnC_3Yroll')

### 3: 1Y lookback
generate_charts_enc(datestart = 52,
                fname = 'encModels_1Ylookback.pickle',
                version = '1Y lookback',
                pdfname = 'EnC_1Yroll')


### 4: 1Y lookback - daily returns
generate_charts_enc(datestart = 252,
                fname = 'encModels_1Ydaily.pickle',
                version = '1Y daily',
                pdfname = 'EnC_1Yroll_daily')

### 5: 4Y lookback - daily returns
generate_charts_enc(datestart = 1000,
                fname = 'encModels_4Ydaily.pickle',
                version = '4Y daily',
                pdfname = 'EnC_4Yroll_daily')

### 5: 6M lookback - daily returns
generate_charts_enc(datestart = 126,
                fname = 'encModels_6mdaily.pickle',
                version = '6m daily',
                pdfname = 'EnC_6mroll_daily')

### 5: 3M lookback - daily returns
generate_charts_enc(datestart = 63,
                fname = 'encModels_3mdaily.pickle',
                version = '3m daily',
                pdfname = 'EnC_3mroll_daily')


#############output for excel#################


def output_results(model, version):

    # load model object
    fname = os.path.join(data_dir, model)
    model = pickle.load( open( fname, "rb" ))
    
    # generate performance stats -- no signal weighting
    model.ensemble = regression_models.ensemble
    model.ensemble(self = model)
    model.performance_stats()

    fx_cor = model.X_noconst.corr()
    fx_cor = fx_cor.mask(np.triu(np.ones(fx_cor.shape, dtype=np.bool_)))
    
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'fx_corr.xlsx'))
    writer.save()
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'fx_corr.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'fx_corr.xlsx'),engine='openpyxl')
    writer.book=book
    fx_cor .to_excel(writer,'corr_weekly')
    writer.save()

    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'co_predictions.xlsx'))
    writer.save()
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'co_predictions.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'co_predictions.xlsx'),engine='openpyxl')
    writer.book=book
    model.actual_returns.to_excel(writer,'actual')
    model.predict_ols['Brent'].to_excel(writer,'OLS')
    model.predict_corr_f['Brent'].to_excel(writer,'Corr_f')
    model.predict_corr_d['Brent'].to_excel(writer,'Corr_d')
    model.predict_lasso['Brent'].to_excel(writer,'lasso')
    model.predict_ridge['Brent'].to_excel(writer,'ridge')
    model.predict_pcr['Brent'].to_excel(writer,'pcr')
    model.predict_pls['Brent'].to_excel(writer,'pls')
    model.predict_rf['Brent'].to_excel(writer,'rf')
    model.predict_gb['Brent'].to_excel(writer,'gb')
    model.predict_psvm['Brent'].to_excel(writer,'psvm')
    model.predict_rsvm['Brent'].to_excel(writer,'rsvm')
    writer.save()

    model.construct_strategy = regression_models.construct_strategy
    model.charting =           regression_models.charting
    model.charting_strat =     regression_models.charting_strat
    
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'hit_by_year.xlsx'))
    writer.save()
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'info_by_year.xlsx'))
    writer.save()
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'annret_by_year.xlsx'))
    writer.save()
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_eq.xlsx'))
    writer.save()
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_iv.xlsx'))
    writer.save()
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'rf_feature_importance.xlsx'))
    writer.save()

    
    for p in model.hit_ratio_by_year.keys(): 
        rf_imp = model.beta_rf[p].copy()
        rf_imp['year'] = rf_imp['date'].dt.year
        rf_imp = rf_imp.iloc[:,1:].astype(float).groupby('year').mean().dropna().rank(axis = 1, ascending = False)
        from openpyxl import load_workbook
        book= load_workbook(os.path.join(data_dir, version + 'rf_feature_importance.xlsx'))
        writer=pd.ExcelWriter(os.path.join(data_dir, version + 'rf_feature_importance.xlsx'),engine='openpyxl')
        writer.book=book
        rf_imp.to_excel(writer,(p))
        writer.save()

        model.hit_ratio_by_year[p] =  model.hit_ratio_by_year[p][model.hit_ratio_by_year[p].index!=2005]
        model.info_ratio_by_year[p] = model.info_ratio_by_year[p][model.info_ratio_by_year[p].index!=2005]
        model.ann_returns_by_year[p] = model.ann_returns_by_year[p][model.ann_returns_by_year[p].index!=2005]
 
        from openpyxl import load_workbook
        book= load_workbook(os.path.join(data_dir, version + 'hit_by_year.xlsx'))
        writer=pd.ExcelWriter(os.path.join(data_dir, version + 'hit_by_year.xlsx'),engine='openpyxl')
        writer.book=book
        model.hit_ratio_by_year[p].to_excel(writer,(p))
        writer.save()

        from openpyxl import load_workbook
        book= load_workbook(os.path.join(data_dir, version + 'info_by_year.xlsx'))
        writer=pd.ExcelWriter(os.path.join(data_dir, version + 'info_by_year.xlsx'),engine='openpyxl')
        writer.book=book
        model.info_ratio_by_year[p].to_excel(writer,(p))
        writer.save()

        from openpyxl import load_workbook
        book= load_workbook(os.path.join(data_dir, version + 'annret_by_year.xlsx'))
        writer=pd.ExcelWriter(os.path.join(data_dir, version + 'annret_by_year.xlsx'),engine='openpyxl')
        writer.book=book
        model.ann_returns_by_year[p].to_excel(writer,(p))
        writer.save()

        
            #construct strategies
    #1 - ALL Commodities
    model.construct_strategy(self = model, basket = _all)
    
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_eq.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_eq.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_eq.to_excel(writer, 'all_ts')
    model.strategy_eq_ret.to_excel(writer,'all_ret')
    model.strategy_eq_info.to_excel(writer,'all_info')    
    writer.save()
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_iv.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_iv.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_iv.to_excel(writer, 'all_ts')
    model.strategy_iv_ret.to_excel(writer,'all_ret')
    model.strategy_iv_info.to_excel(writer,'all_info')    
    writer.save()

    #2 - Energy
    model.construct_strategy(self = model, basket = energy)

    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_eq.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_eq.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_eq.to_excel(writer, 'energy_ts')
    model.strategy_eq_ret.to_excel(writer,'energy_ret')
    model.strategy_eq_info.to_excel(writer,'energy_info')    
    writer.save()
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_iv.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_iv.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_iv.to_excel(writer, 'energy_ts')
    model.strategy_iv_ret.to_excel(writer,'energy_ret')
    model.strategy_iv_info.to_excel(writer,'energy_info')    
    writer.save()
    
    #3 - Agriculture
    model.construct_strategy(self = model, basket = ags)

    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_eq.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_eq.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_eq.to_excel(writer, 'ags_ts')
    model.strategy_eq_ret.to_excel(writer,'ags_ret')
    model.strategy_eq_info.to_excel(writer,'ags_info')    
    writer.save()
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_iv.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_iv.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_iv.to_excel(writer, 'ags_ts')
    model.strategy_iv_ret.to_excel(writer,'ags_ret')
    model.strategy_iv_info.to_excel(writer,'ags_info')    
    writer.save()
    
    #4 - Industrial Metals
    model.construct_strategy(self = model, basket = industrials)

    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_eq.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_eq.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_eq.to_excel(writer, 'indus_ts')
    model.strategy_eq_ret.to_excel(writer,'indus_ret')
    model.strategy_eq_info.to_excel(writer,'indus_info')    
    writer.save()
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_iv.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_iv.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_iv.to_excel(writer, 'indus_ts')
    model.strategy_iv_ret.to_excel(writer,'indus_ret')
    model.strategy_iv_info.to_excel(writer,'indus_info')    
    writer.save()
    
    #5 - Precious Metals
    model.construct_strategy(self = model, basket = precious)

    book= load_workbook(os.path.join(data_dir, version + 'strat_eq.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_eq.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_eq.to_excel(writer, 'prec_ts')
    model.strategy_eq_ret.to_excel(writer,'prec_ret')
    model.strategy_eq_info.to_excel(writer,'prec_info')    
    writer.save()
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_iv.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_iv.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_iv.to_excel(writer, 'prec_ts')
    model.strategy_iv_ret.to_excel(writer,'prec_ret')
    model.strategy_iv_info.to_excel(writer,'prec_info')    
    writer.save()
    
    #6 - Livestock
    model.construct_strategy(self = model, basket = livestock)

    book= load_workbook(os.path.join(data_dir, version + 'strat_eq.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_eq.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_eq.to_excel(writer, 'live_ts')
    model.strategy_eq_ret.to_excel(writer,'live_ret')
    model.strategy_eq_info.to_excel(writer,'live_info')    
    writer.save()
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_iv.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_iv.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_iv.to_excel(writer, 'live_ts')
    model.strategy_iv_ret.to_excel(writer,'live_ret')
    model.strategy_iv_info.to_excel(writer,'live_info')    
    writer.save()
  


#output_results(model = 'encModels_5Ymonthly.pickle', version = 'enc_5Y_monthly_')
#output_results(model = 'encModels_1Ylookback.pickle', version = 'enc_1Y_yearly_')
#output_results(model = 'encModels_3Ylookback.pickle', version = 'enc_3Y_yearly_')
output_results(model = 'encModels_1Ydaily.pickle', version = 'enc_1Y_daily_')
output_results(model = 'encModels_4Ydaily.pickle', version = 'enc_4Y_daily_')
output_results(model = 'encModels_6mdaily.pickle', version = 'enc_6m_daily_')
output_results(model = 'encModels_3mdaily.pickle', version = 'enc_3m_daily_')

output_results(model = 'encModels_latest.pickle', version = 'enc_latest_')

def avg_info(datestart, fname, version, sw):
    # load model object
    fname=os.path.join(data_dir, fname)
    model = pickle.load( open( fname, "rb" ))
    
    # generate performance stats -- no signal weighting
    model.ensemble()

    model.performance_stats = regression_models.performance_stats
    model.performance_stats(self = model, signal_weighting = sw)
    model.actual_returns = model.actual_returns.iloc[datestart:, :]
         
    #construct strategies
    #1 - ALL Commodities
    model.construct_strategy(basket = _all, signal_weighting = sw)
    _al = model.strategy_eq_info[(model.strategy_eq_info.index!=2005) & (model.strategy_eq_info.index!=2020)].mean()  
    _al_ret = model.strategy_eq_ret[(model.strategy_eq_ret.index!=2005) & (model.strategy_eq_ret.index!=2020)].mean()  
    #2 - Energy
    model.construct_strategy(basket = energy, signal_weighting = sw)
    _en = model.strategy_eq_info[(model.strategy_eq_info.index!=2005) & (model.strategy_eq_info.index!=2020)].mean()
    _en_ret = model.strategy_eq_ret[(model.strategy_eq_ret.index!=2005) & (model.strategy_eq_ret.index!=2020)].mean()  
    #3 - Agriculture
    model.construct_strategy(basket = ags, signal_weighting = sw)
    _ag = model.strategy_eq_info[(model.strategy_eq_info.index!=2005) & (model.strategy_eq_info.index!=2020)].mean()
    _ag_ret = model.strategy_eq_ret[(model.strategy_eq_ret.index!=2005) & (model.strategy_eq_ret.index!=2020)].mean()  
    #4 - Industrial Metals
    model.construct_strategy(basket = industrials, signal_weighting = sw)
    _im = model.strategy_eq_info[(model.strategy_eq_info.index!=2005) & (model.strategy_eq_info.index!=2020)].mean()
    _im_ret = model.strategy_eq_ret[(model.strategy_eq_ret.index!=2005) & (model.strategy_eq_ret.index!=2020)].mean()  
    #4 - Precious Metals
    model.construct_strategy(basket = precious, signal_weighting = sw)
    _pm = model.strategy_eq_info[(model.strategy_eq_info.index!=2005) & (model.strategy_eq_info.index!=2020)].mean()
    _pm_ret = model.strategy_eq_ret[(model.strategy_eq_ret.index!=2005) & (model.strategy_eq_ret.index!=2020)].mean()  
    #5 - Livestock
    model.construct_strategy(basket = livestock, signal_weighting = sw)
    _ls = model.strategy_eq_info[(model.strategy_eq_info.index!=2005) & (model.strategy_eq_info.index!=2020)].mean()
    _ls_ret = model.strategy_eq_ret[(model.strategy_eq_ret.index!=2005) & (model.strategy_eq_ret.index!=2020)].mean()  

    avg_info = pd.concat([_al, _en, _ag, _im, _pm, _ls], axis = 1)   
    avg_info.columns = ['All', 'Energy', 'Agricultural', 'Industrial metals', 'Precious metals', 'Livestock'] 
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'avg_info.xlsx'))
    avg_info.to_excel(writer,'avg_info')
    writer.save()

    avg_ret = pd.concat([_al_ret, _en_ret, _ag_ret, _im_ret, _pm_ret, _ls_ret], axis = 1)   
    avg_ret.columns = ['All', 'Energy', 'Agricultural', 'Industrial metals', 'Precious metals', 'Livestock'] 
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'avg_ret.xlsx'))
    avg_ret.to_excel(writer,'avg_ret')
    writer.save()

'''
avg_info(datestart = 52, fname = 'encModels_1Ylookback.pickle', version = 'enc_1Y_yearly_',    sw = False)
avg_info(datestart = 52, fname = 'encModels_1Ylookback.pickle', version = 'enc_1Y_yearly_sw_', sw = True)

avg_info(datestart = 60, fname = 'encModels_5Ymonthly.pickle', version = 'enc_5Y_monthly_',  sw = False)
avg_info(datestart = 260, fname = 'encModels_1Ylookback.pickle', version = 'enc_1Y_weekly_', sw = False)
avg_info(datestart = 260, fname = 'encModels_3Ylookback.pickle', version = 'enc_3Y_weekly_', sw = False)
avg_info(datestart = 260, fname = 'encModels_5Ylookback.pickle', version = 'enc_5Y_weekly_', sw = False)
'''
avg_info(datestart = 1000, fname = 'encModels_4Ydaily.pickle', version = 'enc_4Y_daily_', sw = False)
avg_info(datestart = 1000, fname = 'encModels_1Ydaily.pickle', version = 'enc_1Y_daily_', sw = False)
avg_info(datestart = 1000, fname = 'encModels_6mdaily.pickle', version = 'enc_6m_daily_', sw = False)
avg_info(datestart = 1000, fname = 'encModels_3mdaily.pickle', version = 'enc_3m_daily_', sw = False)


def avg_info_iv(datestart, fname, version, sw):
    # load model object
    fname=os.path.join(data_dir, fname)
    model = pickle.load( open( fname, "rb" ))
    
    # generate performance stats -- no signal weighting
    model.ensemble()
    model.performance_stats(self = model, signal_weighting = sw)
    model.actual_returns = model.actual_returns.iloc[datestart:, :]
         
    #construct strategies
    #1 - ALL Commodities

    model.performance_stats = regression_models.performance_stat
    model.construct_strategy(basket = _all, signal_weighting = sw)
    _al = model.strategy_iv_info[(model.strategy_iv_info.index!=2005) & (model.strategy_iv_info.index!=2020)].mean()  
    #2 - Energy
    model.construct_strategy(basket = energy, signal_weighting = sw)
    _en = model.strategy_iv_info[(model.strategy_iv_info.index!=2005) & (model.strategy_iv_info.index!=2020)].mean()
    #3 - Agriculture
    model.construct_strategy(basket = ags, signal_weighting = sw)
    _ag = model.strategy_iv_info[(model.strategy_iv_info.index!=2005) & (model.strategy_iv_info.index!=2020)].mean()
    #4 - Industrial Metals
    model.construct_strategy(basket = industrials, signal_weighting = sw)
    _im = model.strategy_iv_info[(model.strategy_iv_info.index!=2005) & (model.strategy_iv_info.index!=2020)].mean()
    #4 - Precious Metals
    model.construct_strategy(basket = precious, signal_weighting = sw)
    _pm = model.strategy_iv_info[(model.strategy_iv_info.index!=2005) & (model.strategy_iv_info.index!=2020)].mean()
    #5 - Livestock
    model.construct_strategy(basket = livestock, signal_weighting = sw)
    _ls = model.strategy_iv_info[(model.strategy_iv_info.index!=2005) & (model.strategy_iv_info.index!=2020)].mean()

    avg_info = pd.concat([_al, _en, _ag, _im, _pm, _ls], axis = 1)   
    avg_info.columns = ['All', 'Energy', 'Agricultural', 'Industrial metals', 'Precious metals', 'Livestock'] 
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'avg_info_iv.xlsx'))
    avg_info.to_excel(writer,'avg_info_iv')
    writer.save()

avg_info_iv(datestart = 52, fname = 'encModels_1Ylookback.pickle', version = 'enc_1Y_yearly_',    sw = False)



fname=os.path.join(data_dir, 'encModels_latest.pickle')
model = pickle.load( open( fname, "rb" ))
    
    # generate performance stats -- no signal weighting
model.ensemble()

for q,r in zip(model.predictions, model.models):
    writer=pd.ExcelWriter(os.path.join(data_dir,'2020', r + '.xlsx'))
    writer.save()
    for p in q.keys():
        from openpyxl import load_workbook
        book= load_workbook(os.path.join(data_dir,'2020', r + '.xlsx'))
        writer=pd.ExcelWriter(os.path.join(data_dir,'2020', r + '.xlsx'),engine='openpyxl')
        writer.book=book
        q[p].to_excel(writer,(p))
        writer.save()


fname=os.path.join(data_dir, 'trade_returns_co.pickle')
trade_co = pickle.load( open( fname, "rb" ) )

writer=pd.ExcelWriter(os.path.join(data_dir,'2020/Actual.xlsx'))
writer.save()
from openpyxl import load_workbook
book= load_workbook(os.path.join(data_dir,'2020/Actual.xlsx'))
writer=pd.ExcelWriter(os.path.join(data_dir,'2020/Actual.xlsx'),engine='openpyxl')
writer.book=book
trade_co.to_excel(writer,'Actual')
writer.save()



#############################################################
#CnE
#######################################################################################################################################################################################

dmfx = ['Australia','European Union','Canada','Switzerland','United Kingdom','Japan','Norway','New Zealand','Sweden']
emfx = ['Czech','Hungary','Poland','Romania',
        'South Africa','Israel','Russia','Turkey',
        'Korea','Indonesia','Philippines','Thailand','Taiwan','India','Singapore','China',
        'Brazil','Chile','Peru','Colombia','Mexico']
emfx_1 = ['Czech','Hungary','Poland','Romania']
emfx_2 = ['Korea','Indonesia','Philippines','Thailand','Taiwan','India','Singapore','China', 'Turkey']
emfx_3 = ['Brazil','Chile','Peru','Colombia','Mexico']
emfx_4 = ['Brazil','Russia','India','China','South Africa']
emfx_5 = np.unique(emfx_2 + emfx_3)
emfx_6 = np.unique(emfx_2 + emfx_3 + emfx_4)

def generate_charts(datestart, fname, version, pdfname):
    # load model object
    fname=os.path.join(data_dir, fname)
    model = pickle.load( open( fname, "rb" ))
    
    # generate performance stats -- no signal weighting
    model.ensemble()
    model.performance_stats()
    
    for p in model.hit_ratio_by_year.keys(): 
        model.hit_ratio_by_year[p] =  model.hit_ratio_by_year[p][model.hit_ratio_by_year[p].index!=2020]
        model.info_ratio_by_year[p] = model.info_ratio_by_year[p][model.info_ratio_by_year[p].index!=2020]
        model.ann_returns_by_year[p] = model.ann_returns_by_year[p][model.ann_returns_by_year[p].index!=2020]
      
    model.charting =       regression_models.charting
    model.charting_strat = regression_models.charting_strat
'''
    model.charting(self = model, specs = version + ", no signal weighting", fname = pdfname + "_nw.pdf", out_dir = out_dir)
'''
    #model.actual_returns = model.actual_returns[model.actual_returns[model.mergevar] > datestart]
    model.actual_returns = model.actual_returns.iloc[datestart:, :]
    
    #construct strategies
    #1 - ALL EMFX
    model.construct_strategy(basket = emfx)
    model.charting_strat(self = model, strat_nick = "All EMFX", specs = version + ", no signal weighting", fname = pdfname + "_nw_all_strat.pdf", out_dir = out_dir)
    #2 - EastEUR + ME
    model.construct_strategy(basket = emfx_1)
    model.charting_strat(self = model, strat_nick = "Eastern European FX", specs = version + ", no signal weighting", fname = pdfname + "_nw_em1_strat.pdf", out_dir = out_dir)
    #3 - Asia 
    model.construct_strategy(basket = emfx_2)
    model.charting_strat(self = model, strat_nick = "Asian FX", specs = version + ", no signal weighting", fname = pdfname + "_nw_em2_strat.pdf", out_dir = out_dir)
    #4 - SouthAMR 
    model.construct_strategy(basket = emfx_3)
    model.charting_strat(self = model, strat_nick = "LatAm FX", specs = version + ", no signal weighting", fname = pdfname + "_nw_em3_strat.pdf", out_dir = out_dir)
    #4 - BRICS 
    model.construct_strategy(basket = emfx_4)
    model.charting_strat(self = model, strat_nick = "BRICS FX", specs = version + ", no signal weighting", fname = pdfname + "_nw_em4_strat.pdf", out_dir = out_dir)
    #5 - Asia + South AMR 
    model.construct_strategy(basket = emfx_5)
    model.charting_strat(self = model, strat_nick = "Asian + LatAm FX", specs = version + ", no signal weighting", fname = pdfname + "_nw_em5_strat.pdf", out_dir = out_dir)
    #6 - Asia + South AMR + BRICS 
    model.construct_strategy(basket = emfx_6)
    model.charting_strat(self = model, strat_nick = "BRICS + Asian + LatAm FX", specs = version + ", no signal weighting", fname = pdfname + "_nw_em6_strat.pdf", out_dir = out_dir)
    #7 - DMFX
    model.construct_strategy(basket = dmfx)
    model.charting_strat(self = model, strat_nick = "G-10 FX", specs = version + ", no signal weighting", fname = pdfname + "_nw_dm_strat.pdf", out_dir = out_dir)
'''   
    # generate performance stats -- signal weighting
    model.performance_stats(signal_weighting = True)
    model.charting(self = model, specs = version + ", signal weighting", fname = pdfname + "_sw.pdf", out_dir = out_dir)
    
    #construct strategies
    #1 - ALL EMFX
    model.construct_strategy(basket = emfx, signal_weighting = True)
    model.charting_strat(self = model, strat_nick = "All EMFX", specs = version + ", signal weighting", fname = pdfname + "_sw_all_strat.pdf", out_dir = out_dir)
    #2 - EastEUR + ME
    model.construct_strategy(basket = emfx_1, signal_weighting = True)
    model.charting_strat(self = model, strat_nick = "Eastern European FX", specs = version + ", signal weighting", fname = pdfname + "_sw_em1_strat.pdf", out_dir = out_dir)
    #3 - Asia 
    model.construct_strategy(basket = emfx_2, signal_weighting = True)
    model.charting_strat(self = model, strat_nick = "Asian FX", specs = version + ", signal weighting", fname = pdfname + "_sw_em2_strat.pdf", out_dir = out_dir)
    #4 - SouthAMR 
    model.construct_strategy(basket = emfx_3, signal_weighting = True)
    model.charting_strat(self = model, strat_nick = "LatAm FX", specs = version + ", signal weighting", fname = pdfname + "_sw_em3_strat.pdf", out_dir = out_dir)
    #4 - BRICS 
    model.construct_strategy(basket = emfx_4, signal_weighting = True)
    model.charting_strat(self = model, strat_nick = "BRICS FX", specs = version + ", signal weighting", fname = pdfname + "_sw_em4_strat.pdf", out_dir = out_dir)
    #5 - Asia + South AMR 
    model.construct_strategy(basket = emfx_5, signal_weighting = True)
    model.charting_strat(self = model, strat_nick = "Asian + LatAm FX", specs = version + ", signal weighting", fname = pdfname + "_sw_em5_strat.pdf", out_dir = out_dir)
    #6 - Asia + South AMR + BRICS 
    model.construct_strategy(basket = emfx_6, signal_weighting = True)
    model.charting_strat(self = model, strat_nick = "BRICS + Asian + LatAm FX", specs = version + ", signal weighting", fname = pdfname + "_sw_em6_strat.pdf", out_dir = out_dir)
    #7 - DMFX
    model.construct_strategy(basket = dmfx, signal_weighting = True)
    model.charting_strat(self = model, strat_nick = "G-10 FX", specs = version + ", signal weighting", fname = pdfname + "_sw_dm_strat.pdf", out_dir = out_dir)
'''

### 1: 5Y lookback
generate_charts(datestart = 260,
                fname = 'cneModels_5Ylookback.pickle',
                version = '5Y lookback',
                pdfname = 'CnE_5Yroll')
    
### 1: 3Y lookback
generate_charts(datestart = 160,
                fname = 'cneModels_3Ylookback.pickle',
                version = '3Y lookback',
                pdfname = 'CnE_3Yroll')

### 1: 1Y lookback
generate_charts(datestart = 52,
                fname = 'cneModels_default.pickle',
                version = '1Y lookback',
                pdfname = 'CnE_1Yroll')


### 1: 1Y lookback - daily returns
generate_charts(datestart = 252,
                fname = 'cneModels_1Ydaily.pickle',
                version = '1Y daily',
                pdfname = 'CnE_1Yroll_daily')

#############output for excel#################


def output_results(model, version):

    # load model object
    fname = os.path.join(data_dir, model)
    model = pickle.load( open( fname, "rb" ))
    
    # generate performance stats -- no signal weighting
    model.ensemble = regression_models.ensemble
    model.ensemble(self = model)
    model.performance_stats()

    model.construct_strategy = regression_models.construct_strategy
    model.charting =           regression_models.charting
    model.charting_strat =     regression_models.charting_strat
    
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'hit_by_year.xlsx'))
    writer.save()
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'info_by_year.xlsx'))
    writer.save()
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'annret_by_year.xlsx'))
    writer.save()
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_eq.xlsx'))
    writer.save()
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_iv.xlsx'))
    writer.save()
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'rf_feature_importance.xlsx'))
    writer.save()

    
    for p in model.hit_ratio_by_year.keys(): 
        rf_imp = model.beta_rf[p].copy()
        rf_imp['year'] = rf_imp['date'].dt.year
        rf_imp = rf_imp.iloc[:,1:].astype(float).groupby('year').mean().dropna().rank(axis = 1, ascending = False)
        from openpyxl import load_workbook
        book= load_workbook(os.path.join(data_dir, version + 'rf_feature_importance.xlsx'))
        writer=pd.ExcelWriter(os.path.join(data_dir, version + 'rf_feature_importance.xlsx'),engine='openpyxl')
        writer.book=book
        rf_imp.to_excel(writer,(p))
        writer.save()

        model.hit_ratio_by_year[p] =  model.hit_ratio_by_year[p][model.hit_ratio_by_year[p].index!=2020]
        model.info_ratio_by_year[p] = model.info_ratio_by_year[p][model.info_ratio_by_year[p].index!=2020]
        model.ann_returns_by_year[p] = model.ann_returns_by_year[p][model.ann_returns_by_year[p].index!=2020]
 
        from openpyxl import load_workbook
        book= load_workbook(os.path.join(data_dir, version + 'hit_by_year.xlsx'))
        writer=pd.ExcelWriter(os.path.join(data_dir, version + 'hit_by_year.xlsx'),engine='openpyxl')
        writer.book=book
        model.hit_ratio_by_year[p].to_excel(writer,(p))
        writer.save()

        from openpyxl import load_workbook
        book= load_workbook(os.path.join(data_dir, version + 'info_by_year.xlsx'))
        writer=pd.ExcelWriter(os.path.join(data_dir, version + 'info_by_year.xlsx'),engine='openpyxl')
        writer.book=book
        model.info_ratio_by_year[p].to_excel(writer,(p))
        writer.save()

        from openpyxl import load_workbook
        book= load_workbook(os.path.join(data_dir, version + 'annret_by_year.xlsx'))
        writer=pd.ExcelWriter(os.path.join(data_dir, version + 'annret_by_year.xlsx'),engine='openpyxl')
        writer.book=book
        model.ann_returns_by_year[p].to_excel(writer,(p))
        writer.save()

        
            #construct strategies
    #1 - ALL EM
    model.construct_strategy(self = model, basket = emfx)
    
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_eq.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_eq.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_eq.to_excel(writer, 'emfx_ts')
    model.strategy_eq_ret.to_excel(writer,'emfx_ret')
    model.strategy_eq_info.to_excel(writer,'emfx_info')    
    writer.save()
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_iv.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_iv.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_iv.to_excel(writer, 'emfx_ts')
    model.strategy_iv_ret.to_excel(writer,'emfx_ret')
    model.strategy_iv_info.to_excel(writer,'emfx_info')    
    writer.save()

    #2 - 
    model.construct_strategy(self = model, basket = emfx_1)

    book= load_workbook(os.path.join(data_dir, version + 'strat_eq.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_eq.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_eq.to_excel(writer, 'emfx_1_ts')
    from openpyxl import load_workbook
    model.strategy_eq_ret.to_excel(writer,'emfx_1_ret')
    model.strategy_eq_info.to_excel(writer,'emfx_1_info')    
    writer.save()
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_iv.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_iv.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_iv.to_excel(writer, 'emfx_1_ts')
    model.strategy_iv_ret.to_excel(writer,'emfx_1_ret')
    model.strategy_iv_info.to_excel(writer,'emfx_1_info')    
    writer.save()
    
    #3 - 
    model.construct_strategy(self = model, basket = emfx_2)

    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_eq.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_eq.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_eq.to_excel(writer, 'emfx_2_ts')
    model.strategy_eq_ret.to_excel(writer,'emfx_2_ret')
    model.strategy_eq_info.to_excel(writer,'emfx_2_info')    
    writer.save()
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_iv.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_iv.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_iv.to_excel(writer, 'emfx_2_ts')
    model.strategy_iv_ret.to_excel(writer,'emfx_2_ret')
    model.strategy_iv_info.to_excel(writer,'emfx_2_info')    
    writer.save()
    
    #4 - 
    model.construct_strategy(self = model, basket = emfx_3)

    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_eq.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_eq.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_eq.to_excel(writer, 'emfx_3_ts')
    model.strategy_eq_ret.to_excel(writer,'emfx_3_ret')
    model.strategy_eq_info.to_excel(writer,'emfx_3_info')    
    writer.save()
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_iv.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_iv.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_iv.to_excel(writer, 'emfx_3_ts')
    model.strategy_iv_ret.to_excel(writer,'emfx_3_ret')
    model.strategy_iv_info.to_excel(writer,'emfx_3_info')    
    writer.save()
    
    #5 - 
    model.construct_strategy(self = model, basket = emfx_4)

    book= load_workbook(os.path.join(data_dir, version + 'strat_eq.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_eq.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_eq.to_excel(writer, 'emfx_4_ts')
    model.strategy_eq_ret.to_excel(writer,'emfx_4_ret')
    model.strategy_eq_info.to_excel(writer,'emfx_4_info')    
    writer.save()
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_iv.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_iv.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_iv.to_excel(writer, 'emfx_4_ts')
    model.strategy_iv_ret.to_excel(writer,'emfx_4_ret')
    model.strategy_iv_info.to_excel(writer,'emfx_4_info')    
    writer.save()
    
    #6 - 
    model.construct_strategy(self = model, basket = emfx_5)

    book= load_workbook(os.path.join(data_dir, version + 'strat_eq.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_eq.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_eq.to_excel(writer, 'emfx_5_ts')
    model.strategy_eq_ret.to_excel(writer,'emfx_5_ret')
    model.strategy_eq_info.to_excel(writer,'emfx_5_info')    
    writer.save()
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_iv.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_iv.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_iv.to_excel(writer, 'emfx_5_ts')
    model.strategy_iv_ret.to_excel(writer,'emfx_5_ret')
    model.strategy_iv_info.to_excel(writer,'emfx_5_info')    
    writer.save()
  
    #7 - 
    model.construct_strategy(self = model, basket = emfx_6)

    book= load_workbook(os.path.join(data_dir, version + 'strat_eq.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_eq.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_eq.to_excel(writer, 'emfx_6_ts')
    model.strategy_eq_ret.to_excel(writer,'emfx_6_ret')
    model.strategy_eq_info.to_excel(writer,'emfx_6_info')    
    writer.save()
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_iv.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_iv.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_iv.to_excel(writer, 'emfx_6_ts')
    model.strategy_iv_ret.to_excel(writer,'emfx_6_ret')
    model.strategy_iv_info.to_excel(writer,'emfx_6_info')    
    writer.save()

    #7 - 
    model.construct_strategy(self = model, basket = dmfx)

    book= load_workbook(os.path.join(data_dir, version + 'strat_eq.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_eq.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_eq.to_excel(writer, 'dmfx_ts')
    model.strategy_eq_ret.to_excel(writer,'dmfx_ret')
    model.strategy_eq_info.to_excel(writer,'dmfx_info')    
    writer.save()
    from openpyxl import load_workbook
    book= load_workbook(os.path.join(data_dir, version + 'strat_iv.xlsx'))
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'strat_iv.xlsx'),engine='openpyxl')
    writer.book=book
    model.strategy_iv.to_excel(writer, 'dmfx_ts')
    model.strategy_iv_ret.to_excel(writer,'dmfx_ret')
    model.strategy_iv_info.to_excel(writer,'dmfx_info')    
    writer.save()

output_results(model = 'cneModels_default.pickle', version = 'cne_1Y_weekly_')



def avg_info(datestart, fname, version, sw):
    # load model object
    fname=os.path.join(data_dir, fname)
    model = pickle.load( open( fname, "rb" ))
    
    # generate performance stats -- no signal weighting
    model.ensemble()

    model.performance_stats = regression_models.performance_stats
    model.performance_stats(self = model, signal_weighting = sw)
    model.actual_returns = model.actual_returns.iloc[datestart:, :]
         
    #construct strategies
    #1 - ALL EM
    model.construct_strategy(basket = emfx, signal_weighting = sw)
    _emfx = model.strategy_eq_info[(model.strategy_eq_info.index!=2005) & (model.strategy_eq_info.index!=2020)].mean()  
    _emfx_ret = model.strategy_eq_ret[(model.strategy_eq_ret.index!=2005) & (model.strategy_eq_ret.index!=2020)].mean()  
    #2 - EMFX1
    model.construct_strategy(basket = emfx_1, signal_weighting = sw)
    _emfx_1 = model.strategy_eq_info[(model.strategy_eq_info.index!=2005) & (model.strategy_eq_info.index!=2020)].mean()
    _emfx_1_ret = model.strategy_eq_ret[(model.strategy_eq_ret.index!=2005) & (model.strategy_eq_ret.index!=2020)].mean()  
    #3 - EMFX2
    model.construct_strategy(basket = emfx_2, signal_weighting = sw)
    _emfx_2 = model.strategy_eq_info[(model.strategy_eq_info.index!=2005) & (model.strategy_eq_info.index!=2020)].mean()
    _emfx_2_ret = model.strategy_eq_ret[(model.strategy_eq_ret.index!=2005) & (model.strategy_eq_ret.index!=2020)].mean()  
    #4 - EMFX3
    model.construct_strategy(basket = emfx_3, signal_weighting = sw)
    _emfx_3 = model.strategy_eq_info[(model.strategy_eq_info.index!=2005) & (model.strategy_eq_info.index!=2020)].mean()
    _emfx_3_ret = model.strategy_eq_ret[(model.strategy_eq_ret.index!=2005) & (model.strategy_eq_ret.index!=2020)].mean()  
    #5 - EMFX4
    model.construct_strategy(basket = emfx_4, signal_weighting = sw)
    _emfx_4 = model.strategy_eq_info[(model.strategy_eq_info.index!=2005) & (model.strategy_eq_info.index!=2020)].mean()
    _emfx_4_ret = model.strategy_eq_ret[(model.strategy_eq_ret.index!=2005) & (model.strategy_eq_ret.index!=2020)].mean()  
    #6 - EMFX5
    model.construct_strategy(basket = emfx_5, signal_weighting = sw)
    _emfx_5 = model.strategy_eq_info[(model.strategy_eq_info.index!=2005) & (model.strategy_eq_info.index!=2020)].mean()
    _emfx_5_ret = model.strategy_eq_ret[(model.strategy_eq_ret.index!=2005) & (model.strategy_eq_ret.index!=2020)].mean()  
    #7 - EMFX6
    model.construct_strategy(basket = emfx_6, signal_weighting = sw)
    _emfx_6 = model.strategy_eq_info[(model.strategy_eq_info.index!=2005) & (model.strategy_eq_info.index!=2020)].mean()
    _emfx_6_ret = model.strategy_eq_ret[(model.strategy_eq_ret.index!=2005) & (model.strategy_eq_ret.index!=2020)].mean()  
    #8 - DMFX
    model.construct_strategy(basket = dmfx, signal_weighting = sw)
    _dmfx = model.strategy_eq_info[(model.strategy_eq_info.index!=2005) & (model.strategy_eq_info.index!=2020)].mean()
    _dmfx_ret = model.strategy_eq_ret[(model.strategy_eq_ret.index!=2005) & (model.strategy_eq_ret.index!=2020)].mean()  

    avg_info = pd.concat([_emfx, _emfx_1, _emfx_2, _emfx_3, _emfx_4, _emfx_5, _emfx_6, _dmfx], axis = 1)   
    avg_info.columns = ['All EMFX', 'Eastern European FX', 'Asian FX', 'LatAm FX', 'BRICS FX', 'Asian + LatAm FX', 'BRICS + Asian + LatAm FX', 'G-10 FX'] 
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'avg_info.xlsx'))
    avg_info.to_excel(writer,'avg_info')
    writer.save()

    avg_ret = pd.concat([_emfx_ret, _emfx_1_ret, _emfx_2_ret, _emfx_3_ret, _emfx_4_ret, _emfx_5_ret, _emfx_6_ret, _dmfx_ret], axis = 1)   
    avg_ret.columns = ['All EMFX', 'Eastern European FX', 'Asian FX', 'LatAm FX', 'BRICS FX', 'Asian + LatAm FX', 'BRICS + Asian + LatAm FX', 'G-10 FX'] 
    writer=pd.ExcelWriter(os.path.join(data_dir, version + 'avg_ret.xlsx'))
    avg_ret.to_excel(writer,'avg_ret')
    writer.save()


avg_info(datestart = 52, fname = 'cneModels_default.pickle', version = 'cne_1Y_weekly_',    sw = False)


